from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, redirect, url_for, session, g
from flask_cors import CORS
from werkzeug.utils import secure_filename
from itsdangerous import BadSignature, URLSafeSerializer
import json, os, requests, sys, time, threading
import re, secrets, string
import pandas as pd
from datetime import datetime, timedelta
import concurrent.futures
from openpyxl import load_workbook
from barcode import Code128
from barcode.writer import ImageWriter
from io import BytesIO, StringIO
from openpyxl.drawing.image import Image as XLImage
from db_compat import database_api as sqlite3, using_mysql
from ha_runtime import ha_config, is_standby as ha_is_standby
import base64
import qrcode
import pandas as pd
import zipfile
from copy import copy
import tempfile
import shutil
from pathlib import Path
from urllib.parse import quote
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import weakref
import gzip


DB_LOCK = threading.RLock()


class ThreadLocalSqliteCursor:

    def __init__(self, connection_proxy):
        self._connection_proxy = connection_proxy
        self._local = threading.local()

    def _set_cursor(self, cursor_obj):
        self._local.cursor = cursor_obj
        return cursor_obj

    def _cursor(self):
        cursor_obj = getattr(self._local, "cursor", None)
        if cursor_obj is None:
            cursor_obj = self._connection_proxy._connection().cursor()
            self._set_cursor(cursor_obj)
        return cursor_obj

    def execute(self, *args, **kwargs):
        cursor_obj = self._connection_proxy._connection().cursor()
        cursor_obj.execute(*args, **kwargs)
        self._set_cursor(cursor_obj)
        return self

    def executemany(self, *args, **kwargs):
        cursor_obj = self._connection_proxy._connection().cursor()
        cursor_obj.executemany(*args, **kwargs)
        self._set_cursor(cursor_obj)
        return self

    def fetchone(self, *args, **kwargs):
        return self._cursor().fetchone(*args, **kwargs)

    def fetchall(self, *args, **kwargs):
        return self._cursor().fetchall(*args, **kwargs)

    def fetchmany(self, *args, **kwargs):
        return self._cursor().fetchmany(*args, **kwargs)

    def close(self):
        cursor_obj = getattr(self._local, "cursor", None)
        if cursor_obj is not None:
            cursor_obj.close()
            self._local.cursor = None

    def __iter__(self):
        return iter(self._cursor())

    def __next__(self):
        return next(self._cursor())

    def __getattr__(self, name):
        return getattr(self._cursor(), name)


class ThreadLocalSqliteConnection:

    def __init__(self, db_file):
        self._db_file = db_file
        self._local = threading.local()

    def _connection(self):
        db_conn = getattr(self._local, "connection", None)
        if db_conn is None:
            db_conn = configure_sqlite_connection(
                sqlite3.connect(
                    self._db_file,
                    check_same_thread=False,
                    timeout=15
                )
            )
            self._local.connection = db_conn
        return db_conn

    def cursor(self, *args, **kwargs):
        return ThreadLocalSqliteCursor(self)

    def execute(self, *args, **kwargs):
        return self._connection().execute(*args, **kwargs)

    def executemany(self, *args, **kwargs):
        return self._connection().executemany(*args, **kwargs)

    def commit(self):
        return self._connection().commit()

    def rollback(self):
        return self._connection().rollback()

    def close(self):
        db_conn = getattr(self._local, "connection", None)
        if db_conn is not None:
            db_conn.close()
            self._local.connection = None

    def _release_tracked_cursors(self):
        return None

    def close_thread_connection(self):
        self.close()

    def __getattr__(self, name):
        return getattr(self._connection(), name)


REQUEST_LOCAL = threading.local()

def get_req_session():

    session_obj = getattr(REQUEST_LOCAL, "session", None)

    if session_obj is None:
        session_obj = requests.Session()
        session_obj.headers.update({
            "Connection": "keep-alive"
        })
        REQUEST_LOCAL.session = session_obj

    return session_obj

class ThreadLocalRequests:

    def get(self, *args, **kwargs):
        return get_req_session().get(*args, **kwargs)

    def post(self, *args, **kwargs):
        return get_req_session().post(*args, **kwargs)

req_session = ThreadLocalRequests()
EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=10)
# Scan normal disimpan berurutan di worker terpisah agar respons scanner tidak
# menunggu round-trip database VPS.
SCAN_SAVE_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=1)
# Lookup flag scan dipisahkan dari worker API agar dapat berjalan sambil menunggu
# respons J&T dan tidak memperpanjang respons scanner.
SCAN_LOOKUP_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=1)
AUTH_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=4)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # folder aplikasi saat ini
SCAN_OUTBOX_DIR = Path(BASE_DIR) / "scan_outbox"
    

TEMPLATE_RETUR = "Output.xlsx"
DB_FILE = os.path.join(BASE_DIR, "scanner.db")

def configure_sqlite_connection(db_conn):

    db_conn.execute("PRAGMA journal_mode=WAL")
    db_conn.execute("PRAGMA synchronous=NORMAL")
    db_conn.execute("PRAGMA busy_timeout=15000")

    return db_conn

conn = ThreadLocalSqliteConnection(DB_FILE)
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS scans (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    resi TEXT,
    dp_out TEXT,
    seller TEXT,
    last_station TEXT,
    last_status TEXT,
    harga INTEGER,
    barang TEXT,
    waktu_scan TEXT,
    waktu TEXT,
    status TEXT,
    spot TEXT,
    kode TEXT,
    collect_staff TEXT,
    badges TEXT,
    examine_time TEXT,
    scan_pack_seller_time TEXT,
    scan_delivery_time TEXT,
    scan_pack_time TEXT,
    scan_tracer_station TEXT,
    scan_by TEXT
    
)
""")

try:
    cursor.execute("""
        ALTER TABLE scans
        ADD COLUMN collect_staff TEXT
    """)
    conn.commit()
    

except:
    pass

for tracking_column in (
    "examine_time",
    "scan_pack_seller_time",
    "scan_delivery_time",
    "scan_pack_time",
    "scan_tracer_station",
    "scan_by",
    "scan_pack_code",
    "scan_delivery_code",
    "scan_pack_seller_code"
):
    try:
        cursor.execute(
            f"ALTER TABLE scans ADD COLUMN {tracking_column} TEXT"
        )
        conn.commit()
    except sqlite3.OperationalError:
        pass

try:
    cursor.execute("ALTER TABLE scans ADD COLUMN bulk_import_id TEXT")
    conn.commit()
except sqlite3.OperationalError:
    pass

cursor.execute("""
CREATE TABLE IF NOT EXISTS bulk_import_batches (
    batch_id TEXT PRIMARY KEY,
    import_date TEXT NOT NULL,
    import_mode TEXT,
    created_at TEXT NOT NULL,
    status TEXT NOT NULL,
    total_items INTEGER DEFAULT 0,
    processed INTEGER DEFAULT 0,
    undone_at TEXT
)
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS bulk_import_changes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    batch_id TEXT NOT NULL,
    scan_id INTEGER NOT NULL,
    action TEXT NOT NULL,
    old_data TEXT,
    UNIQUE(batch_id, scan_id)
)
""")
cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_bulk_import_batches_date
ON bulk_import_batches(import_date, created_at)
""")
conn.commit()
    
try:
    cursor.execute("""
        ALTER TABLE scans
        ADD COLUMN badges TEXT
    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE scans
        ADD COLUMN received_at TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE scans
        ADD COLUMN received_photo TEXT

    """)
    conn.commit()
except:
    pass
    
try:

    cursor.execute("""

        ALTER TABLE scans
        ADD COLUMN kode TEXT

    """)
    conn.commit()

except:
    pass

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_waktu
ON scans(waktu)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_waktu_status_id
ON scans(waktu, status, id DESC)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_resi
ON scans(resi)
""")


cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_resi_waktu_id
ON scans(resi, waktu, id DESC)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_kode_waktu
ON scans(kode, waktu)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_scan_pack_time
ON scans(scan_pack_time, waktu)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_scan_delivery_time
ON scans(scan_delivery_time, waktu)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_scan_pack_seller_time
ON scans(scan_pack_seller_time, waktu)
""")

cursor.execute("""
CREATE INDEX IF NOT EXISTS idx_scans_scan_tracer_station_time
ON scans(scan_tracer_station, waktu)
""")
conn.commit()
cursor.execute("""
CREATE TABLE IF NOT EXISTS manual_flags (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    resi TEXT,

    type TEXT,

    seller TEXT,

    nominal INTEGER,

    tanggal TEXT,

    photo TEXT,

    created_at TEXT

)
""")

cursor.execute("""

CREATE TABLE IF NOT EXISTS retur_batches (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    batch_code TEXT,

    seller TEXT,

    tanggal TEXT,

    total_resi INTEGER,

    resi_data TEXT,

    created_at TEXT,

    received_at TEXT,

    received_photo TEXT,

    status TEXT

)

""")

conn.commit()

try:
    cursor.execute("""

        ALTER TABLE retur_batches
        ADD COLUMN status TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE retur_batches
        ADD COLUMN received_at TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE retur_batches
        ADD COLUMN received_photo TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""
        ALTER TABLE manual_flags
        ADD COLUMN keterangan TEXT
    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""
        ALTER TABLE manual_flags
        ADD COLUMN updated_at TEXT
    """)
    conn.commit()
except:
    pass

conn.commit()

# =========================
# USERS
# =========================

cursor.execute("""

CREATE TABLE IF NOT EXISTS users (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    fullname TEXT,

    username TEXT UNIQUE,

    password TEXT,

    level TEXT,

    permissions TEXT

)

""")

conn.commit()

try:

    cursor.execute("""

        ALTER TABLE users
        ADD COLUMN permissions TEXT

    """)

    conn.commit()

except:
    pass

try:

    cursor.execute("""

        ALTER TABLE users
        ADD COLUMN must_change_credentials INTEGER DEFAULT 0

    """)

    conn.commit()

except:
    pass

try:

    cursor.execute("""

        ALTER TABLE users
        ADD COLUMN attendance_bypass INTEGER DEFAULT 0

    """)

    conn.commit()

except:
    pass

try:

    cursor.execute("""

        ALTER TABLE users
        ADD COLUMN iphone_user INTEGER DEFAULT 0

    """)

    conn.commit()

except:
    pass

try:

    cursor.execute("""

        ALTER TABLE users
        ADD COLUMN disable_location_lock INTEGER DEFAULT 0

    """)

    conn.commit()

except:
    pass

for _app_profile_column_sql in [
    "ALTER TABLE users ADD COLUMN app_display_name TEXT",
    "ALTER TABLE users ADD COLUMN department TEXT",
    "ALTER TABLE users ADD COLUMN profile_photo TEXT"
]:
    try:
        cursor.execute(_app_profile_column_sql)
        conn.commit()
    except Exception:
        pass

# =========================
# ACCOUNT MESSAGES
# =========================

cursor.execute("""

CREATE TABLE IF NOT EXISTS account_messages (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    sender_username TEXT,

    sender_fullname TEXT,

    recipient_username TEXT,

    recipient_fullname TEXT,

    message TEXT,

    image_path TEXT,

    created_at TEXT,

    read_at TEXT

)

""")

conn.commit()
for _account_message_column_sql in [
    "ALTER TABLE account_messages ADD COLUMN message_type TEXT DEFAULT 'message'",
    "ALTER TABLE account_messages ADD COLUMN action_url TEXT",
    "ALTER TABLE account_messages ADD COLUMN action_label TEXT",
    "ALTER TABLE account_messages ADD COLUMN app_version TEXT"
]:
    try:
        cursor.execute(_account_message_column_sql)
        conn.commit()
    except Exception:
        pass

# =========================
# ATTENDANCE
# =========================

cursor.execute("""

CREATE TABLE IF NOT EXISTS attendance (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    username TEXT,

    fullname TEXT,

    tanggal TEXT,

    jam TEXT,

    photo TEXT,

    latitude TEXT,

    longitude TEXT,

    address TEXT,

    created_at TEXT

)

""")

conn.commit()

cursor.execute("""

CREATE TABLE IF NOT EXISTS attendance_sync_events (
    event_id TEXT PRIMARY KEY,
    username TEXT NOT NULL,
    mode TEXT NOT NULL,
    attendance_id INTEGER,
    captured_at TEXT NOT NULL,
    synced_at TEXT NOT NULL
)

""")
conn.commit()

cursor.execute("""

CREATE TABLE IF NOT EXISTS attendance_leave (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    username TEXT,

    fullname TEXT,

    tanggal TEXT,

    type TEXT,

    keterangan TEXT,

    photo TEXT,

    created_at TEXT,

    updated_at TEXT,

    UNIQUE(username, tanggal)

)

""")

conn.commit()

cursor.execute("""

CREATE TABLE IF NOT EXISTS attendance_settings (

    key TEXT PRIMARY KEY,

    value TEXT,

    updated_at TEXT

)

""")

conn.commit()

try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN address TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN clock_out TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN clock_out_at TEXT

    """)
    conn.commit()
except:
    pass
try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN clock_out_photo TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN clock_out_latitude TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN clock_out_longitude TEXT

    """)
    conn.commit()
except:
    pass

try:
    cursor.execute("""

        ALTER TABLE attendance
        ADD COLUMN clock_out_address TEXT

    """)
    conn.commit()
except:
    pass

for _attendance_column_sql in [
    "ALTER TABLE attendance ADD COLUMN manual_entry INTEGER DEFAULT 0",
    "ALTER TABLE attendance ADD COLUMN shift_id TEXT",
    "ALTER TABLE attendance ADD COLUMN face_score INTEGER DEFAULT 0",
    "ALTER TABLE attendance ADD COLUMN clock_out_face_score INTEGER DEFAULT 0",
    "ALTER TABLE attendance ADD COLUMN device_info TEXT",
    "ALTER TABLE attendance ADD COLUMN clock_out_device_info TEXT",
    "ALTER TABLE attendance ADD COLUMN ip_address TEXT",
    "ALTER TABLE attendance ADD COLUMN clock_out_ip_address TEXT",
    "ALTER TABLE attendance ADD COLUMN app_version TEXT"
]:
    try:
        cursor.execute(_attendance_column_sql)
        conn.commit()
    except Exception:
        pass

try:
    cursor.execute("ALTER TABLE attendance_leave ADD COLUMN status TEXT DEFAULT 'Tercatat'")
    conn.commit()
except Exception:
    pass

try:
    cursor.execute("""

        CREATE TABLE IF NOT EXISTS server_nodes (
            node_id VARCHAR(80) PRIMARY KEY,
            node_name VARCHAR(120),
            node_kind VARCHAR(20),
            priority INTEGER DEFAULT 50,
            hostname VARCHAR(160),
            local_ip VARCHAR(80),
            app_url TEXT,
            app_version VARCHAR(40),
            flask_ok INTEGER DEFAULT 0,
            db_tunnel_ok INTEGER DEFAULT 0,
            cloudflared_ok INTEGER DEFAULT 0,
            last_mode VARCHAR(20),
            last_seen VARCHAR(32),
            started_at VARCHAR(32),
            updated_at VARCHAR(32)
        )

    """)
    cursor.execute("""

        CREATE TABLE IF NOT EXISTS server_cluster_state (
            cluster_id VARCHAR(40) PRIMARY KEY,
            active_node_id VARCHAR(80),
            lease_until VARCHAR(32),
            generation INTEGER DEFAULT 0,
            updated_at VARCHAR(32)
        )

    """)
    conn.commit()
except Exception:
    pass

for _attendance_index_sql in (
    "CREATE INDEX IF NOT EXISTS idx_attendance_tanggal_username ON attendance(tanggal, username)",
    "CREATE INDEX IF NOT EXISTS idx_attendance_username_tanggal ON attendance(username, tanggal)",
    "CREATE INDEX IF NOT EXISTS idx_attendance_leave_tanggal_username ON attendance_leave(tanggal, username)",
):
    try:
        cursor.execute(_attendance_index_sql)
        conn.commit()
    except Exception:
        pass

cursor.execute("""

CREATE TABLE IF NOT EXISTS attendance_offline_sync (

    event_id TEXT PRIMARY KEY,

    username TEXT,

    mode TEXT,

    attendance_id INTEGER,

    synced_at TEXT

)

""")
conn.commit()

try:
    cursor.execute("""

        CREATE TABLE IF NOT EXISTS attendance_shifts (
            username TEXT PRIMARY KEY,
            total_shift INTEGER DEFAULT 1,
            waktu_shift TEXT,
            shift1_clock_in TEXT,
            shift1_clock_out TEXT,
            shift2_clock_in TEXT,
            shift2_clock_out TEXT,
            updated_at TEXT
        )

    """)
    conn.commit()
except Exception:
    pass
# =========================
# DEFAULT ACCOUNT
# =========================

cursor.execute("""

SELECT * FROM users

WHERE username=?

""", ("superman",))

user = cursor.fetchone()

if not user:

    cursor.execute("""

    INSERT INTO users (

        fullname,
        username,
        password,
        level

    )

    VALUES (?,?,?,?)

    """, (

        "Superman",
        "superman",
        os.environ.get("TRACER_INITIAL_SUPERMAN_PASSWORD", "CHANGE_ME_BEFORE_FIRST_RUN"),
        "SUPERMAN"

    ))

    conn.commit()

def _load_flask_secret():
    configured = str(os.environ.get("TRACER_FLASK_SECRET_KEY", "")).strip()
    if configured:
        return configured

    secret_path = Path(__file__).with_name(".flask_secret")
    try:
        existing = secret_path.read_text(encoding="utf-8").strip()
        if existing:
            return existing
    except FileNotFoundError:
        pass

    generated = secrets.token_hex(32)
    try:
        with secret_path.open("x", encoding="utf-8") as secret_file:
            secret_file.write(generated)
    except FileExistsError:
        pass
    return secret_path.read_text(encoding="utf-8").strip() or generated


app = Flask(__name__)
app.secret_key = _load_flask_secret()
app.permanent_session_lifetime = timedelta(days=2)
MONITOR_URL = "http://localhost:5000/api/notify_access"  # alamat dashboard kamu
CORS(app)


@app.after_request
def compress_large_text_response(response):
    """Compress dashboard/API payloads for browsers without touching APK/files."""
    if (
        "gzip" not in str(request.headers.get("Accept-Encoding", "")).lower()
        or response.status_code < 200
        or response.status_code >= 300
        or response.direct_passthrough
        or response.headers.get("Content-Encoding")
        or response.mimetype not in {
            "application/json", "text/html", "text/css", "application/javascript", "text/javascript"
        }
    ):
        return response
    payload = response.get_data()
    if len(payload) < 2048:
        return response
    compressed = gzip.compress(payload, compresslevel=5)
    if len(compressed) >= len(payload):
        return response
    response.set_data(compressed)
    response.headers["Content-Encoding"] = "gzip"
    response.headers["Content-Length"] = str(len(compressed))
    response.vary.add("Accept-Encoding")
    return response


@app.route('/api/ha/health', methods=['GET'])
def ha_health():
    """Small health endpoint for the future domain failover monitor."""
    try:
        check_cursor = conn.cursor()
        check_cursor.execute("SELECT 1")
        check_cursor.fetchone()
        database_ok = True
    except Exception:
        database_ok = False
    config = ha_config()
    return jsonify({
        "healthy": database_ok,
        "node": config.get("HA_NODE_NAME", "unknown"),
        "role": config.get("HA_NODE_ROLE", "primary"),
    }), (200 if database_ok else 503)


SERVER_CLUSTER_ID = "tracer-main"
SERVER_HEARTBEAT_TIMEOUT_SECONDS = 40
SERVER_LEASE_SECONDS = 45


def parse_cluster_datetime(value):
    try:
        return datetime.strptime(str(value or ""), "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def server_control_authorized():
    configured = str(ha_config().get("HA_SHARED_SECRET", "")).strip()
    supplied = str(request.headers.get("X-Tracer-Server-Key", "")).strip()
    return bool(
        configured and
        supplied and
        secrets.compare_digest(configured, supplied)
    )


def normalize_server_node_payload(payload):
    node_id = str(payload.get("node_id", "")).strip()[:80]
    node_name = str(payload.get("node_name", "")).strip()[:120]
    node_kind = str(payload.get("node_kind", "backup")).strip().lower()
    if node_kind not in {"primary", "backup"}:
        node_kind = "backup"
    if not re.fullmatch(r"[A-Za-z0-9._-]{3,80}", node_id):
        raise ValueError("node_id server tidak valid")
    if not node_name:
        node_name = node_id
    return {
        "node_id": node_id,
        "node_name": node_name,
        "node_kind": node_kind,
        "priority": 100 if node_kind == "primary" else 50,
        "hostname": str(payload.get("hostname", "")).strip()[:160],
        "local_ip": str(payload.get("local_ip", "")).strip()[:80],
        "app_url": str(payload.get("app_url", "")).strip()[:500],
        "app_version": str(payload.get("app_version", "")).strip()[:40],
        "flask_ok": 1 if truthy_flag(payload.get("flask_ok", 0)) else 0,
        "db_tunnel_ok": 1 if truthy_flag(payload.get("db_tunnel_ok", 0)) else 0,
        "cloudflared_ok": 1 if truthy_flag(payload.get("cloudflared_ok", 0)) else 0,
        "started_at": str(payload.get("started_at", "")).strip()[:32]
    }


def upsert_server_node(conn_local, node, now_text):
    existing = conn_local.execute(
        "SELECT node_id FROM server_nodes WHERE node_id=? LIMIT 1",
        (node["node_id"],)
    ).fetchone()
    values = (
        node["node_name"], node["node_kind"], node["priority"],
        node["hostname"], node["local_ip"], node["app_url"],
        node["app_version"], node["flask_ok"], node["db_tunnel_ok"],
        node["cloudflared_ok"], now_text, node["started_at"], now_text,
        node["node_id"]
    )
    if existing:
        conn_local.execute("""
            UPDATE server_nodes
            SET node_name=?, node_kind=?, priority=?, hostname=?, local_ip=?,
                app_url=?, app_version=?, flask_ok=?, db_tunnel_ok=?,
                cloudflared_ok=?, last_seen=?,
                started_at=COALESCE(NULLIF(started_at, ''), ?), updated_at=?
            WHERE node_id=?
        """, values)
    else:
        conn_local.execute("""
            INSERT INTO server_nodes (
                node_name, node_kind, priority, hostname, local_ip, app_url,
                app_version, flask_ok, db_tunnel_ok, cloudflared_ok,
                last_seen, started_at, updated_at, node_id
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, values)


def coordinate_server_cluster(node):
    now_dt = datetime.now()
    now_text = now_dt.strftime("%Y-%m-%d %H:%M:%S")
    cutoff_text = (
        now_dt - timedelta(seconds=SERVER_HEARTBEAT_TIMEOUT_SECONDS)
    ).strftime("%Y-%m-%d %H:%M:%S")
    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
    try:
        if using_mysql():
            # Row-level transaction lock is released automatically on
            # commit/rollback. Unlike GET_LOCK, it cannot remain attached to
            # an idle pooled connection and stall later heartbeats.
            conn_local._raw.begin()

        upsert_server_node(conn_local, node, now_text)
        state_query = """
            SELECT active_node_id, lease_until, generation
            FROM server_cluster_state
            WHERE cluster_id=?
            LIMIT 1
        """
        if using_mysql():
            state_query += " FOR UPDATE"
        state = conn_local.execute(
            state_query, (SERVER_CLUSTER_ID,)
        ).fetchone()

        eligible_rows = conn_local.execute("""
            SELECT node_id, node_name, node_kind, priority, app_url
            FROM server_nodes
            WHERE last_seen>=?
              AND flask_ok=1
              AND db_tunnel_ok=1
            ORDER BY priority DESC, node_id ASC
        """, (cutoff_text,)).fetchall()
        eligible = {
            str(row[0]): {
                "node_id": str(row[0]),
                "node_name": row[1] or row[0],
                "node_kind": row[2] or "backup",
                "priority": int(row[3] or 0),
                "app_url": row[4] or ""
            }
            for row in eligible_rows
        }

        active_node_id = str(state[0] or "") if state else ""
        lease_until_dt = parse_cluster_datetime(state[1] if state else "")
        generation = int(state[2] or 0) if state else 0
        current_valid = bool(
            active_node_id in eligible and
            lease_until_dt and
            lease_until_dt > now_dt
        )
        best_node_id = str(eligible_rows[0][0]) if eligible_rows else ""

        if not current_valid:
            selected_node_id = best_node_id
        elif (
            best_node_id and
            eligible[best_node_id]["priority"] > eligible[active_node_id]["priority"]
        ):
            selected_node_id = best_node_id
        else:
            selected_node_id = active_node_id

        if selected_node_id:
            new_lease_until = (
                now_dt + timedelta(seconds=SERVER_LEASE_SECONDS)
            ).strftime("%Y-%m-%d %H:%M:%S")
        else:
            new_lease_until = ""

        if not state:
            generation = 1
            conn_local.execute("""
                INSERT INTO server_cluster_state (
                    cluster_id, active_node_id, lease_until, generation, updated_at
                ) VALUES (?,?,?,?,?)
            """, (
                SERVER_CLUSTER_ID, selected_node_id, new_lease_until,
                generation, now_text
            ))
        elif selected_node_id != active_node_id:
            generation += 1
            conn_local.execute("""
                UPDATE server_cluster_state
                SET active_node_id=?, lease_until=?, generation=?, updated_at=?
                WHERE cluster_id=?
            """, (
                selected_node_id, new_lease_until, generation,
                now_text, SERVER_CLUSTER_ID
            ))
        elif selected_node_id == node["node_id"]:
            conn_local.execute("""
                UPDATE server_cluster_state
                SET lease_until=?, updated_at=?
                WHERE cluster_id=?
            """, (new_lease_until, now_text, SERVER_CLUSTER_ID))
        else:
            new_lease_until = str(state[1] or "")

        mode = "write" if selected_node_id == node["node_id"] else "read_only"
        conn_local.execute(
            "UPDATE server_nodes SET last_mode=? WHERE node_id=?",
            (mode, node["node_id"])
        )
        conn_local.commit()
        active = eligible.get(selected_node_id, {})
        return {
            "success": True,
            "mode": mode,
            "standby": mode != "write",
            "active_node_id": selected_node_id,
            "active_node_name": active.get("node_name", selected_node_id),
            "active_node_kind": active.get("node_kind", ""),
            "active_app_url": active.get("app_url", ""),
            "lease_until": new_lease_until,
            "generation": generation,
            "heartbeat_interval": 5
        }
    except Exception:
        try:
            conn_local.rollback()
        except Exception:
            pass
        raise
    finally:
        conn_local.close()


@app.route('/api/ha/heartbeat', methods=['POST'])
def server_ha_heartbeat():
    if not server_control_authorized():
        return jsonify({"success": False, "error": "Kunci server tidak valid"}), 401
    try:
        node = normalize_server_node_payload(request.get_json(silent=True) or {})
        return jsonify(coordinate_server_cluster(node))
    except ValueError as error:
        return jsonify({"success": False, "error": str(error)}), 400
    except Exception as error:
        app.logger.exception("server heartbeat failed")
        return jsonify({"success": False, "error": str(error)}), 503


def cluster_guard_enabled():
    return truthy_flag(ha_config().get("HA_CLUSTER_GUARD", "0"))


def cluster_node_can_write():
    config = ha_config()
    node_id = str(config.get("HA_NODE_ID", "")).strip()
    if not node_id:
        return False, ""
    now_dt = datetime.now()
    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=8))
    try:
        row = conn_local.execute("""
            SELECT active_node_id, lease_until
            FROM server_cluster_state
            WHERE cluster_id=?
            LIMIT 1
        """, (SERVER_CLUSTER_ID,)).fetchone()
        if not row:
            return False, ""
        active_node_id = str(row[0] or "")
        lease_until = parse_cluster_datetime(row[1])
        allowed = bool(active_node_id == node_id and lease_until and lease_until > now_dt)
        return allowed, active_node_id
    finally:
        conn_local.close()

def create_attendance_sync_token(username):
    serializer = URLSafeSerializer(
        app.secret_key,
        salt="tracer-attendance-sync"
    )
    return serializer.dumps({
        "username": str(username or "").strip()
    })


def read_attendance_sync_token(token):
    serializer = URLSafeSerializer(
        app.secret_key,
        salt="tracer-attendance-sync"
    )
    try:
        payload = serializer.loads(str(token or ""))
    except BadSignature:
        return ""
    return str(payload.get("username", "")).strip()


UPLOAD_FOLDER = os.path.join(
    BASE_DIR,
    "bulk_upload"
)

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)

MESSAGE_UPLOAD_FOLDER = os.path.join(
    BASE_DIR,
    "static",
    "message_uploads"
)

os.makedirs(
    MESSAGE_UPLOAD_FOLDER,
    exist_ok=True
)
PROFILE_UPLOAD_FOLDER = os.path.join(
    BASE_DIR,
    "static",
    "profile_uploads"
)

os.makedirs(
    PROFILE_UPLOAD_FOLDER,
    exist_ok=True
)

CONFIG_DIR = os.path.join(BASE_DIR, 'config')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
SCAN_BUFFER = []
LAST_SCAN_TIME = None
BUFFER_LOCK = threading.Lock()
BULK_JOB_LOCK = threading.Lock()
BULK_IMPORT_WORKERS = 2
BULK_PACKAGE_WORKERS = 2
# Jangan timeout paksa future yang sudah running; Python tidak bisa
# membunuh thread itu. Timeout tetap dikontrol di requests.* timeout.
BULK_ITEM_TIMEOUT_SECONDS = 0
BULK_WAIT_POLL_SECONDS = 1
BULK_STAGE_DIR = Path(BASE_DIR) / "bulk_staging"
BULK_STAGE_DIR.mkdir(parents=True, exist_ok=True)
BULK_JOB = {

    "running":False,

    "total":0,

    "processed":0,

    "success":0,

    "failed":0,

    "skipped":0,

    "skip_details":[],

    "import_date":"",

    "current_item":"",

    "current_step":"",

    "last_update":"",

    "started_at":"",

    "done":False,

    "cancel":False

}

def update_bulk_job(**values):

    values["last_update"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    with BULK_JOB_LOCK:
        BULK_JOB.update(values)

def bulk_job_add(key, amount=1):

    with BULK_JOB_LOCK:
        BULK_JOB[key] = BULK_JOB.get(key, 0) + amount
        BULK_JOB["last_update"] = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

def bulk_job_skip(resi, network, reason):

    with BULK_JOB_LOCK:
        BULK_JOB["skipped"] = BULK_JOB.get("skipped", 0) + 1
        BULK_JOB.setdefault("skip_details", []).append({
            "resi": resi,
            "network": network,
            "reason": reason
        })
        BULK_JOB["skip_details"] = BULK_JOB["skip_details"][-50:]
        BULK_JOB["last_update"] = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )


def bulk_job_get(key, default=None):

    with BULK_JOB_LOCK:
        return BULK_JOB.get(key, default)


def bulk_job_snapshot():

    with BULK_JOB_LOCK:
        return dict(BULK_JOB)


def _write_json_atomic(destination, payload):
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name("." + destination.name + ".tmp")
    with temporary.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, separators=(",", ":"))
        file.flush()
        os.fsync(file.fileno())
    os.replace(temporary, destination)


def _bulk_stage_paths(batch_id):
    folder = BULK_STAGE_DIR / str(batch_id)
    return folder, folder / "manifest.json"


def write_bulk_checkpoint(batch_id, items, next_index, state="RUNNING"):
    folder, manifest = _bulk_stage_paths(batch_id)
    _write_json_atomic(manifest, {
        "batch_id": batch_id, "items": items, "next_index": int(next_index),
        "state": state, "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })


def stage_bulk_scan(batch_id, data):
    folder, _ = _bulk_stage_paths(batch_id)
    name = f"scan_{int(time.time() * 1000)}_{secrets.token_hex(6)}.json"
    _write_json_atomic(folder / name, {"data": data, "bulk_import_id": batch_id})


def flush_staged_bulk_scans(batch_id):
    folder, _ = _bulk_stage_paths(batch_id)
    staged_files = sorted(folder.glob("scan_*.json"))
    if not staged_files:
        return True, 0
    completed = []
    try:
        with DB_LOCK:
            for staged_file in staged_files:
                with staged_file.open("r", encoding="utf-8") as file:
                    payload = json.load(file)
                if not save_sqlite(payload["data"], payload.get("bulk_import_id"), commit=False):
                    raise RuntimeError("Simpan data bulk ke VPS gagal")
                completed.append(staged_file)
            conn.commit()
        for staged_file in completed:
            staged_file.unlink(missing_ok=True)
        return True, len(completed)
    except Exception as error:
        try:
            conn.rollback()
        except Exception:
            pass
        print("BULK VPS SYNC ERROR:", error)
        return False, 0

HEADERS_CACHE = None
HEADERS_LOCK = threading.RLock()
MAPPING_CACHE = None
MAPPING_INDEX_CACHE = None
MAPPING_CACHE_MTIME = None
MAPPING_CACHE_CHECKED_AT = 0.0
MAPPING_LOCK = threading.RLock()


os.makedirs(CONFIG_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADERS_FILE = os.path.join(BASE_DIR, 'headers.json')  # simpan di root
MAPPING_FILE = os.path.join(CONFIG_DIR, 'mapping_implant.json')
SCAN_SETTINGS_FILE = os.path.join(CONFIG_DIR, 'scan_settings.json')

DEFAULT_SCAN_SETTINGS = {
    "badges": {
        "OUTGOING": True,
        "INCOMING": True,
        "COMPLAINT": True,
        "AUTOCLAIM": True,
        "CLAIM INTERNAL": True,
        "HIGH VALUE": True
    },
    "fields": {
        "resi": True,
        "dp_out": True,
        "implant": True,
        "seller": True,
        "collect_staff": True,
        "last_station": True,
        "last_status": True,
        "harga": True,
        "waktu_scan": True,
        "barang": True
    }
}




def normalize_scan_station(value):

    station = str(
        value or ""
    ).strip()

    station_map = {
        "Gudang Modern Park": "Gudang Modern Park (Drop Off)",
        "Gudang Modern Park (Drop Off)": "Gudang Modern Park (Drop Off)",
        "Gudang Tanah Tinggi": "Gudang Tanah Tinggi",
        "Gudang Retur 188": "Gudang Retur 188",
        "Gudang 188": "Gudang Retur 188",
        "Retur 188": "Gudang Retur 188"
    }

    return station_map.get(
        station,
        ""
    )

def dashboard_package_station(
    scan_tracer_station,
    scan_delivery_time="",
    scan_pack_seller_time=""
):

    station = normalize_scan_station(
        scan_tracer_station
    )

    if station:

        return station

    if (
        str(scan_delivery_time or "").strip()
        and
        str(scan_delivery_time or "").strip() != "-"
    ):

        return "Gudang Retur 188"

    if (
        str(scan_pack_seller_time or "").strip()
        and
        str(scan_pack_seller_time or "").strip() != "-"
    ):

        return "Gudang Retur 188"

    return "Gudang Modern Park (Drop Off)"
def check_login():

    if not session.get("logged_in"):

        return False

    return True
    
def has_permission(permission):

    if (
        str(session.get("level", "")).upper() == "SUPERMAN" and
        str(session.get("username", "")).lower() == "superman"
    ):

        return True

    perms = session.get(
        "permissions",
        []
    )

    return permission in normalize_permissions(
        perms
    )

def is_superman_session():

    return (
        str(
            session.get("level", "")
        ).upper() == "SUPERMAN" and
        str(
            session.get("username", "")
        ).lower() == "superman"
    )

def is_reserved_superman_account(username=None, level=None):

    return (
        str(username or "").strip().lower() == "superman" or
        str(level or "").strip().upper() == "SUPERMAN"
    )

ATTENDANCE_ONLY_PERMISSION = "ATTENDANCE_ONLY"
ATTENDANCE_BYPASS_PERMISSION = "ATTENDANCE_BYPASS"
ATTENDANCE_QR_TOKENS = {}
ATTENDANCE_QR_TTL_SECONDS = 600
ATTENDANCE_APP_SCHEME = "attendanceapp"
ATTENDANCE_APK_DOWNLOAD_URL = "/attendance_apk"

def normalize_permissions(permissions):

    if isinstance(
        permissions,
        str
    ):

        permissions = permissions.split(",")

    return [
        str(item).strip()
        for item in (permissions or [])
        if str(item).strip()
    ]

def is_attendance_only_permissions(permissions):

    return set(
        normalize_permissions(
            permissions
        )
    ) == {
        ATTENDANCE_ONLY_PERMISSION
    }

def has_attendance_bypass_permission(permissions):

    return (
        ATTENDANCE_BYPASS_PERMISSION in normalize_permissions(
            permissions
        )
    )



def cleanup_attendance_qr_tokens():

    now = time.time()

    expired_tokens = [
        token
        for token, payload in ATTENDANCE_QR_TOKENS.items()
        if payload.get("expires_at", 0) < now
    ]

    for token in expired_tokens:

        ATTENDANCE_QR_TOKENS.pop(
            token,
            None
        )


def create_attendance_qr_token(mode="clock_in"):

    cleanup_attendance_qr_tokens()

    token = secrets.token_urlsafe(
        32
    )

    if mode == "clock_out":

        username = session.get(
            "username",
            ""
        )
        fullname = session.get(
            "fullname",
            username
        )
        level = session.get(
            "level",
            ""
        )
        permissions = session.get(
            "permissions",
            []
        )
        attendance_bypass = session.get(
            "attendance_bypass",
            0
        )

    else:

        username = session.get(
            "pending_username",
            ""
        )
        fullname = session.get(
            "pending_fullname",
            username
        )
        level = session.get(
            "pending_level",
            ""
        )
        permissions = session.get(
            "pending_permissions",
            []
        )
        attendance_bypass = session.get(
            "pending_attendance_bypass",
            0
        )

    ATTENDANCE_QR_TOKENS[token] = {
        "mode": mode,
        "username": username,
        "fullname": fullname,
        "level": level,
        "permissions": permissions,
        "attendance_bypass": attendance_bypass,
        "iphone_user": session.get("iphone_user", session.get("pending_iphone_user", 0)),
        "expires_at": time.time() + ATTENDANCE_QR_TTL_SECONDS
    }

    return token

def build_attendance_app_links(token):

    web_url = url_for(
        'attendance_mobile_qr',
        token=token,
        source="app",
        _external=True
    )
    launcher_url = url_for(
        'attendance_app_launcher',
        token=token,
        _external=True
    )
    payload = ATTENDANCE_QR_TOKENS.get(
        token,
        {}
    )
    mode = payload.get(
        "mode",
        "clock_in"
    )
    username = payload.get(
        "username",
        ""
    )
    fullname = payload.get(
        "fullname",
        username
    )
    level = payload.get(
        "level",
        ""
    )
    last_photo_url = ""
    try:
        last_photo_url = get_last_attendance_photo_url(
            username
        )
        if last_photo_url:
            last_photo_url = url_for(
                "static",
                filename=last_photo_url.replace(
                    "/static/",
                    ""
                ),
                _external=True
            )
    except Exception:
        last_photo_url = ""

    app_url = "{}://attendance/open?token={}&mode={}&web_url={}&username={}&fullname={}&level={}&last_photo_url={}".format(
        ATTENDANCE_APP_SCHEME,
        quote(token, safe=""),
        quote(mode, safe=""),
        quote(web_url, safe=""),
        quote(username, safe=""),
        quote(fullname, safe=""),
        quote(level, safe=""),
        quote(last_photo_url, safe="")
    )

    return {
        "web_url": web_url,
        "launcher_url": launcher_url,
        "app_url": app_url,
        "install_url": ATTENDANCE_APK_DOWNLOAD_URL
    }

def is_attendance_only_session():

    return is_attendance_only_permissions(
        session.get(
            "permissions",
            []
        )
    )

def is_pending_attendance_only_session():

    return is_attendance_only_permissions(
        session.get(
            "pending_permissions",
            []
        )
    )


def current_attendance_date():

    return datetime.now().strftime(
        "%Y-%m-%d"
    )


def mark_attendance_only_session_date():

    session.permanent = True
    session["attendance_session_date"] = current_attendance_date()


def expire_stale_attendance_only_session():

    if not session.get("logged_in"):

        return

    if not is_attendance_only_session():

        return

    username = session.get(
        "username",
        ""
    )

    if not username:

        session.clear()
        return

    today = current_attendance_date()
    session_date = session.get(
        "attendance_session_date",
        ""
    )

    pending_clock_out = get_pending_clock_out_attendance(
        username
    )

    if pending_clock_out:

        session["last_attendance_id"] = pending_clock_out.get(
            "id"
        )

        if (
            session_date and
            session_date != today
        ):

            session["attendance_notice"] = "Belum clock out di sesi sebelumnya. Clock out dulu untuk clock in tanggal hari ini."

        session["attendance_session_date"] = today
        return

    attendance = get_today_attendance(
        username
    )

    if (
        session_date and
        session_date != today
    ):

        session.clear()
        return

    if not attendance:

        session.clear()
        return

    if not session_date:

        session["attendance_session_date"] = today


def has_pending_clock_out(attendance):

    return bool(
        attendance and
        not str(
            attendance.get(
                "clock_out",
                ""
            ) or ""
        ).strip()
    )
def sanitize_user_permissions(permissions):

    permission_list = normalize_permissions(
        permissions
    )
    bypass_enabled = (
        ATTENDANCE_BYPASS_PERMISSION in permission_list
    )

    if ATTENDANCE_ONLY_PERMISSION in permission_list:

        sanitized = [
            ATTENDANCE_ONLY_PERMISSION
        ]

        if bypass_enabled:
            sanitized.append(
                ATTENDANCE_BYPASS_PERMISSION
            )

        return sanitized

    return permission_list

def build_username_base(fullname):

    cleaned = re.sub(
        r"[^a-z0-9]+",
        ".",
        str(fullname or "").lower()
    ).strip(".")

    parts = [
        item
        for item in cleaned.split(".")
        if item
    ]

    if len(parts) >= 2:

        return (
            parts[0] +
            "." +
            parts[-1]
        )[:24].strip(".")

    if parts:

        return parts[0][:24].strip(".")

    return "user"

def generate_import_password():

    alphabet = string.ascii_letters + string.digits

    return "".join(
        secrets.choice(alphabet)
        for _ in range(8)
    )

def generate_unique_username(fullname, cursor_obj):

    base = build_username_base(
        fullname
    )

    username = base
    counter = 1

    while cursor_obj.execute("""

        SELECT 1
        FROM users
        WHERE username=?

    """,(username,)).fetchone():

        counter += 1

        username = (
            f"{base}{counter}"
        )

    return username

PROTECTED_LOGIN_ENDPOINTS = {
    "admin",
    "admin_panel",
    "manual_list",
    "manual_add",
    "manual_delete",
    "manual_update",
    "dashboard_data",
    "attendance_dashboard_data",
    "save_attendance_settings",
    "search_resi",
    "export_preview",
    "test_har",
    "scan_resi",
    "scan_common",
    "get_complaints",
    "get_ip",
    "debug_scan",
    "save_spot_photo",
    "upload_spot_photo",
    "check_resi",
    "serve_image",
    "export_zip",
    "admin_complaint",
    "admin_claim",
    "save_manual_flag",
    "get_retur_batch",
    "lookup_retur_receipt",
    "submit_retur",
    "bulk_headers",
    "bulk_upload",
    "bulk_progress",
    "bulk_cancel",
    "bulk_undo",
    "refresh_live_tracking",
    "scanner_debug",
    "send_account_message",
    "send_account_notification",
    "api_account_bulk_settings",
    "api_attendance_locations",
    "api_delete_attendance_location",
    "account_message_poll",
    "account_message_ack",
    "add_user",
    "import_users",
    "delete_user",
    "delete_attendance_user",
    "update_user",
    "check_attendance_status",
    "api_permissions",
}

PROTECTED_PAGE_ENDPOINTS = {
    "admin",
    "admin_panel",
    "admin_complaint",
    "admin_claim",
    "account_management",
    "scan",
    "dashboard",
    "attendance_dashboard",
    "upload_har",
    "scan_settings",
    "mapping_implant",
    "retur_page",
}

ENDPOINT_PERMISSIONS = {
    "admin": "ACCOUNT_MANAGEMENT",
    "admin_panel": "ADMIN_PANEL",
    "account_management": "ACCOUNT_MANAGEMENT",
    "send_account_message": "ACCOUNT_MANAGEMENT",
    "send_account_notification": "ACCOUNT_MANAGEMENT",
    "api_account_bulk_settings": "ACCOUNT_MANAGEMENT",
    "api_attendance_locations": "ACCOUNT_MANAGEMENT",
    "api_delete_attendance_location": "ACCOUNT_MANAGEMENT",
    "add_user": "ACCOUNT_MANAGEMENT",
    "import_users": "ACCOUNT_MANAGEMENT",
    "delete_user": "ACCOUNT_MANAGEMENT",
    "delete_attendance_user": "MANAGE_ABSEN",
    "save_attendance_manual": "MANAGE_ABSEN",
    "save_attendance_shift": "DASHBOARD_ABSEN",
    "update_user": "ACCOUNT_MANAGEMENT",
    "manual_list": "MANUAL_FLAG",
    "manual_add": "MANUAL_FLAG",
    "manual_delete": "MANUAL_FLAG",
    "manual_update": "MANUAL_FLAG",
    "save_manual_flag": "MANUAL_FLAG",
    "admin_complaint": "MANUAL_FLAG",
    "admin_claim": "MANUAL_FLAG",
    "scan_resi": "SCAN_SORTIR",
    "scan": "SCAN_SORTIR",
    "scan_common": "SCAN_SORTIR",
    "get_complaints": "SCAN_SORTIR",
    "save_spot_photo": "SCAN_SORTIR",
    "upload_spot_photo": "SCAN_SORTIR",
    "check_resi": "SCAN_SORTIR",
    "dashboard_data": "DASHBOARD_RETUR",
    "dashboard": "DASHBOARD_RETUR",
    "search_resi": "DASHBOARD_RETUR",
    "attendance_dashboard_data": "DASHBOARD_ABSEN",
    "save_attendance_settings": "DASHBOARD_ABSEN",
    "attendance_dashboard": "DASHBOARD_ABSEN",
    "check_attendance_status": "DASHBOARD_ABSEN",
    "save_attendance_leave": "MANAGE_ABSEN",
    "upload_har": "SCAN_SETTINGS",
    "test_har": "SCAN_SETTINGS",
    "scan_settings": "SCAN_SETTINGS",
    "mapping_implant": "SCAN_SETTINGS",
    "mapping_implant_import": "SCAN_SETTINGS",
    "mapping_implant_export": "SCAN_SETTINGS",
    "export_preview": "EXPORT_DATA",
    "export_zip": "EXPORT_DATA",
    "serve_image": "EXPORT_DATA",
    "retur_page": "RETUR_PROCESS",
    "get_retur_batch": "RETUR_PROCESS",
    "lookup_retur_receipt": "RETUR_PROCESS",
    "submit_retur": "RETUR_PROCESS",
    "bulk_headers": "IMPORT_DATA",
    "bulk_upload": "IMPORT_DATA",
    "bulk_cancel": "IMPORT_DATA",
    "bulk_undo": "IMPORT_DATA",
    "bulk_progress": "DASHBOARD_RETUR",
    "refresh_live_tracking": "DASHBOARD_RETUR",
    "get_ip": "DEBUG_TOOLS",
    "debug_scan": "DEBUG_TOOLS",
    "scanner_debug": "DEBUG_TOOLS",
}

@app.before_request
def expire_attendance_only_session_daily():

    expire_stale_attendance_only_session()

    return None


@app.before_request
def block_writes_while_standby():
    """A standby can be inspected, but it must not create diverging scan data."""
    if (
        request.method in {"GET", "HEAD", "OPTIONS"} or
        request.path.startswith("/api/ha/")
    ):
        return None

    active_node_id = ""
    if cluster_guard_enabled():
        try:
            write_allowed, active_node_id = cluster_node_can_write()
        except Exception:
            write_allowed = False
        standby = not write_allowed
    else:
        standby = ha_is_standby()

    if (
        standby
    ):
        return jsonify({
            "success": False,
            "error": "Server ini sedang mode read-only/siaga. Data hanya dapat ditulis pada server aktif.",
            "active_node_id": active_node_id
        }), 503
    return None


@app.before_request
def protect_login_required_routes():

    endpoint = request.endpoint

    if (
        endpoint not in PROTECTED_LOGIN_ENDPOINTS and
        endpoint not in ENDPOINT_PERMISSIONS
    ):

        return None

    if check_login():

        required_permission = (
            ENDPOINT_PERMISSIONS.get(
                endpoint
            )
        )

        if (
            not required_permission or
            has_permission(required_permission)
        ):

            return None

        if (
            endpoint in PROTECTED_PAGE_ENDPOINTS and
            request.method == "GET"
        ):

            return redirect("/?access_denied=1")

        return jsonify({
            "success":False,
            "error":"Tidak punya akses"
        }), 403

    if (
        endpoint in PROTECTED_PAGE_ENDPOINTS and
        request.method == "GET"
    ):

        return redirect("/")

    return jsonify({
        "success":False,
        "error":"Login required"
    }), 401


@app.teardown_request
def cleanup_sqlite_operation_locks(error=None):

    try:
        conn._release_tracked_cursors()
    except Exception:
        pass

# === LOAD CONFIG ===
def load_headers():
    global HEADERS_CACHE
    with HEADERS_LOCK:
        if HEADERS_CACHE is None:
            with open(HEADERS_FILE, 'r', encoding='utf-8') as f:
                HEADERS_CACHE = json.load(f)
        return dict(HEADERS_CACHE)

RETURN_DATE_FOR_PAGE_URL = "https://jmsgw.jntexpress.id/operatingplatform/returnsManagement/returnDateForPage"
PACKAGE_WAYBILLS_URL = "https://jmsgw.jntexpress.id/operatingplatform/packScanList/waybillIdsByPackageNumber"
TRACKING_NEW_URL = "https://jmsgw.jntexpress.id/operatingplatform/podTrackingNew/inner/query/keywordList"


def collect_audit_network_names(value):

    names = []

    if isinstance(value, dict):

        network_name = value.get("auditNetworkName")

        if network_name:

            names.append(str(network_name).strip())

        for child in value.values():

            names.extend(collect_audit_network_names(child))

    elif isinstance(value, list):

        for child in value:

            names.extend(collect_audit_network_names(child))

    return names

def find_first_audit_network_name(value):

    if isinstance(value, dict):

        network_name = value.get("auditNetworkName")

        if network_name:

            return str(network_name).strip()

        for child in value.values():

            found = find_first_audit_network_name(child)

            if found:

                return found

    elif isinstance(value, list):

        for child in value:

            found = find_first_audit_network_name(child)

            if found:

                return found

    return ""


def normalize_import_date(import_date=None):

    selected_date = str(import_date or "").strip()

    try:
        return datetime.strptime(selected_date, "%Y-%m-%d").strftime("%Y-%m-%d")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d")


def check_bulk_return_network(resi, import_date=None):

    selected_date = normalize_import_date(import_date)

    payload = {
        "current": 1,
        "size": 20,
        "waybillId": "",
        "waybillIds": [resi],
        "applyNetworkCode": "",
        "applyNetworkCodes": [],
        "countryId": "1",
        "dateType": 1,
        "startTime": f"{selected_date} 00:00:00",
        "endTime": f"{selected_date} 23:59:59",
        "exportType": 3,
        "expressTypeCodes": [],
        "orderSourceCodes": [],
        "selectTime": "",
        "status": "",
        "totalCount": 1,
        "type": 3
    }

    try:

        response = req_session.post(
            RETURN_DATE_FOR_PAGE_URL,
            headers=load_headers(),
            json=payload,
            timeout=10
        )

        if response.status_code == 401:

            return {
                "allowed": False,
                "network": "-",
                "reason": "Token expired"
            }

        if response.status_code != 200:

            return {
                "allowed": False,
                "network": "-",
                "reason": f"Filter API status {response.status_code}"
            }

        data = response.json()
        network = find_first_audit_network_name(data) or "-"

        return {
            "allowed": network == "MODERN_PARK",
            "network": network,
            "reason": f"auditNetworkName={network}"
        }

    except Exception as e:

        return {
            "allowed": False,
            "network": "-",
            "reason": f"Filter API error: {e}"
        }

def get_waybills_by_package_number(package_number):

    package_number = str(package_number or "").strip()

    if not package_number:

        return {
            "success": False,
            "waybill_ids": [],
            "message": "Kode kosong"
        }

    try:

        response = req_session.post(
            PACKAGE_WAYBILLS_URL,
            headers=load_headers(),
            json=[package_number],
            timeout=20
        )

        if response.status_code == 401:

            return {
                "success": False,
                "waybill_ids": [],
                "message": "Token expired"
            }

        if response.status_code != 200:

            return {
                "success": False,
                "waybill_ids": [],
                "message": f"waybillIdsByPackageNumber status {response.status_code}"
            }

        data = response.json()
        rows = data.get("data") or []

        for row in rows:

            if str(row.get("packageNumber") or "").strip() == package_number:

                waybill_ids = [
                    str(item).strip()
                    for item in (row.get("waybillIds") or [])
                    if str(item).strip()
                ]

                return {
                    "success": True,
                    "waybill_ids": waybill_ids,
                    "message": data.get("msg", "OK")
                }

        return {
            "success": True,
            "waybill_ids": [],
            "message": "Kode tidak punya resi"
        }

    except Exception as e:

        return {
            "success": False,
            "waybill_ids": [],
            "message": str(e)
        }


def collect_return_examine_records(value):

    records = []

    if isinstance(value, dict):

        resi = (
            value.get("waybillNo") or
            value.get("waybillId") or
            value.get("waybillCode")
        )
        examine_time = value.get("examineTime")
        network = str(
            value.get("examineNetworkName") or
            value.get("auditNetworkName") or
            ""
        ).strip()

        if resi and examine_time and network == "MODERN_PARK":
            records.append({
                "resi": str(resi).strip(),
                "examine_time": str(examine_time).strip()
            })

        for child in value.values():
            records.extend(collect_return_examine_records(child))

    elif isinstance(value, list):

        for child in value:
            records.extend(collect_return_examine_records(child))

    return records


def check_package_return_quota(waybill_ids, import_date=None):

    selected_date = normalize_import_date(import_date)
    waybill_ids = [
        str(item).strip()
        for item in (waybill_ids or [])
        if str(item).strip()
    ]

    if not waybill_ids:

        return {
            "allowed": False,
            "return_count": 0,
            "reason": "Tidak ada resi dari kode"
        }

    payload = {
        "current": 1,
        "size": 20,
        "waybillId": "",
        "waybillIds": waybill_ids,
        "applyNetworkCode": "",
        "applyNetworkCodes": [],
        "countryId": "1",
        "dateType": 1,
        "startTime": f"{selected_date} 00:00:00",
        "endTime": f"{selected_date} 23:59:59",
        "exportType": 3,
        "expressTypeCodes": [],
        "orderSourceCodes": [],
        "selectTime": "",
        "status": "",
        "totalCount": len(waybill_ids),
        "type": 3
    }

    try:

        return_records = []
        max_pages = max(1, (len(waybill_ids) + 19) // 20)

        for current_page in range(1, max_pages + 1):

            payload["current"] = current_page
            response = req_session.post(
                RETURN_DATE_FOR_PAGE_URL,
                headers=load_headers(),
                json=payload,
                timeout=30
            )

            if response.status_code == 401:

                return {
                    "allowed": False,
                    "return_count": 0,
                    "reason": "Token expired"
                }

            if response.status_code != 200:

                return {
                    "allowed": False,
                    "return_count": 0,
                    "reason": f"returnDateForPage status {response.status_code}"
                }

            data = response.json()
            page_records = (data.get("data") or {}).get("records") or []
            return_records.extend(
                collect_return_examine_records(page_records)
            )

            if len(page_records) < payload["size"]:
                break

        examine_times = {}

        for record in return_records:
            key = record["resi"].upper()
            current = examine_times.get(key)
            candidate = record["examine_time"]
            if not current or candidate > current:
                examine_times[key] = candidate

        return_waybills = [
            item
            for item in waybill_ids
            if item.upper() in examine_times
        ]
        matched_return_count = len(return_waybills)

        return {
            "allowed": matched_return_count >= 5,
            "return_count": matched_return_count,
            "reason": f"{matched_return_count} retur MODERN_PARK",
            "return_waybills": return_waybills,
            "examine_times": examine_times
        }

    except Exception as e:

        return {
            "allowed": False,
            "return_count": 0,
            "reason": f"Filter kode error: {e}"
        }

def build_sensitive_headers(headers):

    sensitive_headers = dict(headers or {})

    sensitive_headers.update({
        "accept": "application/json, text/plain, */*",
        "content-type": "application/json;charset=UTF-8",
        "origin": "https://jms.jntexpress.id",
        "referer": "https://jms.jntexpress.id/",
        "routeName": "sendWaybillSite",
        "routeNameList": "%E7%BD%91%E7%82%B9%E7%BB%8F%E8%90%A5%3E%E8%BF%90%E5%8D%95%E7%AE%A1%E7%90%86%3E%E5%AF%84%E4%BB%B6%E8%BF%90%E5%8D%95%E7%AE%A1%E7%90%86"
    })

    return sensitive_headers


def normalize_mapping_key(value):
    text = str(value or "").strip().upper()
    return "" if text in ("", "-", "NAN", "NONE") else text


def normalize_mapping_rows(raw_mapping):
    if isinstance(raw_mapping, dict) and isinstance(raw_mapping.get("rows"), list):
        source_rows = raw_mapping.get("rows", [])
    elif isinstance(raw_mapping, list):
        source_rows = raw_mapping
    elif isinstance(raw_mapping, dict):
        # Migrasi otomatis format lama: {SPRINTER: IMPLANT}
        source_rows = [
            {"sprinter": sprinter, "seller": "", "implant": implant}
            for sprinter, implant in raw_mapping.items()
        ]
    else:
        source_rows = []

    normalized = {}
    for row in source_rows:
        if not isinstance(row, dict):
            continue

        sprinter = normalize_mapping_key(row.get("sprinter"))
        seller = normalize_mapping_key(row.get("seller"))
        implant = str(row.get("implant") or "-").strip() or "-"

        if not sprinter and not seller:
            continue

        normalized[(sprinter, seller)] = {
            "sprinter": sprinter,
            "seller": seller,
            "implant": implant
        }

    return list(normalized.values())


def build_mapping_indexes(rows):
    by_sprinter = {}
    by_seller = {}

    for row in rows:
        implant = str(row.get("implant") or "-").strip() or "-"
        sprinter = normalize_mapping_key(row.get("sprinter"))
        seller = normalize_mapping_key(row.get("seller"))

        if sprinter:
            by_sprinter[sprinter] = implant
        if seller:
            by_seller[seller] = implant

    return {
        "sprinter": by_sprinter,
        "seller": by_seller
    }


def load_mapping():
    global MAPPING_CACHE, MAPPING_INDEX_CACHE
    global MAPPING_CACHE_MTIME, MAPPING_CACHE_CHECKED_AT

    with MAPPING_LOCK:
        now = time.monotonic()
        if MAPPING_CACHE is not None and now - MAPPING_CACHE_CHECKED_AT < 1.0:
            return list(MAPPING_CACHE)

        current_mtime = (
            os.path.getmtime(MAPPING_FILE)
            if os.path.exists(MAPPING_FILE)
            else None
        )

        if MAPPING_CACHE is None or current_mtime != MAPPING_CACHE_MTIME:
            raw_mapping = {}
            if os.path.exists(MAPPING_FILE):
                with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
                    raw_mapping = json.load(f)

            MAPPING_CACHE = normalize_mapping_rows(raw_mapping)
            MAPPING_INDEX_CACHE = build_mapping_indexes(MAPPING_CACHE)
            MAPPING_CACHE_MTIME = current_mtime

        MAPPING_CACHE_CHECKED_AT = now
        return list(MAPPING_CACHE)


def save_mapping(mapping_rows):
    global MAPPING_CACHE, MAPPING_INDEX_CACHE
    global MAPPING_CACHE_MTIME, MAPPING_CACHE_CHECKED_AT

    rows = normalize_mapping_rows(mapping_rows)
    payload = {"version": 2, "rows": rows}

    with MAPPING_LOCK:
        temporary_file = f"{MAPPING_FILE}.{os.getpid()}.tmp"
        with open(temporary_file, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
            f.flush()
            os.fsync(f.fileno())
        os.replace(temporary_file, MAPPING_FILE)

        MAPPING_CACHE = rows
        MAPPING_INDEX_CACHE = build_mapping_indexes(rows)
        MAPPING_CACHE_MTIME = os.path.getmtime(MAPPING_FILE)
        MAPPING_CACHE_CHECKED_AT = time.monotonic()

    return rows


def resolve_implant_with_source(sprinter=None, seller=None):
    global MAPPING_INDEX_CACHE

    load_mapping()
    with MAPPING_LOCK:
        indexes = MAPPING_INDEX_CACHE or {"sprinter": {}, "seller": {}}
        indexes = {
            "sprinter": dict(indexes.get("sprinter", {})),
            "seller": dict(indexes.get("seller", {}))
        }
    seller_key = normalize_mapping_key(seller)
    sprinter_key = normalize_mapping_key(sprinter)

    # Seller diprioritaskan agar aturan seller khusus bisa override sprinter umum.
    if seller_key and seller_key in indexes["seller"]:
        seller_implant = indexes["seller"][seller_key]
        if seller_implant != "-":
            return seller_implant, "seller"
    if sprinter_key and sprinter_key in indexes["sprinter"]:
        sprinter_implant = indexes["sprinter"][sprinter_key]
        if sprinter_implant != "-":
            return sprinter_implant, "sprinter"
    return "-", ""


def resolve_implant(sprinter=None, seller=None):
    implant, _ = resolve_implant_with_source(
        sprinter=sprinter,
        seller=seller
    )
    return implant


def load_scan_settings():
    if not os.path.exists(SCAN_SETTINGS_FILE):
        return DEFAULT_SCAN_SETTINGS

    try:
        with open(SCAN_SETTINGS_FILE, 'r', encoding='utf-8') as f:
            saved = json.load(f)
    except:
        saved = {}

    settings = json.loads(json.dumps(DEFAULT_SCAN_SETTINGS))

    for group in ["badges", "fields"]:
        settings[group].update(saved.get(group, {}))

    return settings

def save_scan_settings(settings):
    with open(SCAN_SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, indent=2, ensure_ascii=False)
    

def _scan_value_filled(value):

    text = str(value or "").strip()

    return text not in ("", "-")


def _split_scan_badges(value):

    return [
        item.strip()
        for item in str(value or "").split(",")
        if item.strip()
    ]


def _merge_scan_badges(old_value, new_value):

    merged = []

    for badge in _split_scan_badges(old_value) + _split_scan_badges(new_value):

        if badge not in merged:

            merged.append(badge)

    return ",".join(merged)


def _same_day_bounds(scan_time):

    try:

        base = datetime.strptime(str(scan_time), "%Y-%m-%d %H:%M:%S")

    except Exception:

        base = datetime.now()

    start = base.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start + timedelta(days=1)

    return (
        start.strftime("%Y-%m-%d %H:%M:%S"),
        end.strftime("%Y-%m-%d %H:%M:%S")
    )


def _scan_snapshot(db_cursor, scan_id):
    db_cursor.execute("SELECT * FROM scans WHERE id=?", (scan_id,))
    row = db_cursor.fetchone()
    if not row:
        return None
    columns = [item[0] for item in db_cursor.description]
    return dict(zip(columns, row))


def get_manual_flag_types(resi):
    """Read manual flags safely; callers may run this alongside J&T requests."""
    with DB_LOCK:
        rows = cursor.execute("""
            SELECT type
            FROM manual_flags
            WHERE resi = ?
        """, (resi,)).fetchall()
    return [row[0] for row in rows]


def _record_bulk_import_change(db_cursor, batch_id, scan_id, action, old_data=None):
    if not batch_id:
        return
    db_cursor.execute("""
        INSERT OR IGNORE INTO bulk_import_changes
        (batch_id, scan_id, action, old_data)
        VALUES (?, ?, ?, ?)
    """, (
        batch_id,
        scan_id,
        action,
        json.dumps(old_data, ensure_ascii=False) if old_data is not None else None
    ))


def save_sqlite(data, bulk_import_id=None, commit=True):

    try:

        with DB_LOCK:

            resi = data[0]
            waktu = data[9]
            kode = data[12]
            badges = data[13]
            scan_tracer_station = data[14] if len(data) > 14 else ""
            scan_by = data[15] if len(data) > 15 else ""
            is_delivery_scan = _scan_value_filled(scan_tracer_station)
            day_start, day_end = _same_day_bounds(waktu)

            # ANTI DOUBLE SCAN < 5 DETIK
            cursor.execute("""
            SELECT waktu
            FROM scans
            WHERE resi=?
            ORDER BY id DESC
            LIMIT 1
            """, (resi,))

            last_scan = cursor.fetchone()

            if last_scan:

                try:

                    last_time = datetime.strptime(
                        last_scan[0],
                        "%Y-%m-%d %H:%M:%S"
                    )

                    now_time = datetime.now()
                    diff = (now_time - last_time).total_seconds()

                    if diff < 5:

                        print(f"FAST RESCAN: {resi}")

                except Exception as e:

                    print("ANTI DUPLICATE ERROR:", e)

            cursor.execute("""
            SELECT
                id,
                kode,
                badges
            FROM scans
            WHERE resi=?
            AND waktu >= ?
            AND waktu < ?
            ORDER BY
                CASE WHEN COALESCE(NULLIF(TRIM(kode), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                waktu DESC,
                CASE WHEN COALESCE(NULLIF(TRIM(badges), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                id DESC
            LIMIT 1
            """, (resi, day_start, day_end))

            existing = cursor.fetchone()

            if existing:

                existing_id, existing_kode, existing_badges = existing
                if bulk_import_id:
                    _record_bulk_import_change(
                        cursor,
                        bulk_import_id,
                        existing_id,
                        "UPDATE",
                        _scan_snapshot(cursor, existing_id)
                    )
                merged_badges = _merge_scan_badges(existing_badges, badges)
                incoming_kode = str(kode or "").strip()
                final_kode = existing_kode if _scan_value_filled(existing_kode) else incoming_kode
                has_latest_kode = _scan_value_filled(kode) and not _scan_value_filled(existing_kode)
                has_new_badge = merged_badges != str(existing_badges or "").strip()
                delivery_code = final_kode if _scan_value_filled(final_kode) else ""

                if has_latest_kode or has_new_badge or is_delivery_scan:

                    print("SQLITE UPDATE ENRICHED:", data)

                    cursor.execute("""
                    UPDATE scans
                    SET
                        dp_out=?,
                        seller=?,
                        collect_staff=?,
                        last_station=?,
                        last_status=?,
                        harga=?,
                        barang=?,
                        waktu_scan=?,
                        waktu=?,
                        status=?,
                        spot=?,
                        kode=?,
                        badges=?,
                        scan_tracer_station=CASE
                            WHEN ? != '' THEN ?
                            ELSE scan_tracer_station
                        END,
                        scan_by=CASE
                            WHEN ? != '' THEN ?
                            ELSE scan_by
                        END,
                        scan_delivery_time=CASE
                            WHEN ? != '' THEN ?
                            ELSE scan_delivery_time
                        END,
                        scan_delivery_code=CASE
                            WHEN ? != ''
                             AND COALESCE(NULLIF(TRIM(scan_delivery_code), ''), '-') = '-'
                            THEN ?
                            ELSE scan_delivery_code
                        END
                    WHERE id=?
                    """, (
                        data[1], data[2], data[3], data[4], data[5], data[6],
                        data[7], data[8], waktu, data[10], data[11],
                        final_kode, merged_badges,
                        scan_tracer_station if is_delivery_scan else "",
                        scan_tracer_station if is_delivery_scan else "",
                        scan_by if is_delivery_scan else "",
                        scan_by if is_delivery_scan else "",
                        waktu if is_delivery_scan else "",
                        waktu if is_delivery_scan else "",
                        delivery_code if is_delivery_scan else "",
                        delivery_code if is_delivery_scan else "",
                        existing_id
                    ))

                else:

                    print("SQLITE UPDATE TIME ONLY:", resi)

                    cursor.execute("""
                    UPDATE scans
                    SET
                        waktu=?,
                        waktu_scan=?
                    WHERE id=?
                    """, (waktu, data[8], existing_id))

                if commit:
                    conn.commit()
                print("SQLITE SUCCESS")
                return True

            print("SQLITE INSERT:", data)

            cursor.execute("""
            INSERT INTO scans
            (resi, dp_out, seller, collect_staff,
             last_station, last_status, harga,
             barang, waktu_scan, waktu,
             status, spot, kode, badges,
             scan_tracer_station, scan_by,
             scan_delivery_time, scan_delivery_code,
             bulk_import_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                data[0], data[1], data[2], data[3],
                data[4], data[5], data[6], data[7],
                data[8], data[9], data[10], data[11],
                data[12], data[13],
                scan_tracer_station if is_delivery_scan else "",
                scan_by if is_delivery_scan else "",
                waktu if is_delivery_scan else "",
                data[12] if _scan_value_filled(data[12]) else "",
                bulk_import_id
            ))

            if bulk_import_id:
                _record_bulk_import_change(
                    cursor,
                    bulk_import_id,
                    cursor.lastrowid,
                    "INSERT"
                )

            if commit:
                conn.commit()
            print("SQLITE SUCCESS")
            return True

    except Exception as e:

        print("SQLITE ERROR:", e)
        return False


def _write_scan_outbox(data, bulk_import_id, refresh_args):
    """Durably queue a normal scan before responding to the scanner client."""
    SCAN_OUTBOX_DIR.mkdir(parents=True, exist_ok=True)
    job_name = f"{int(time.time() * 1000)}_{secrets.token_hex(8)}.json"
    destination = SCAN_OUTBOX_DIR / job_name
    temporary = SCAN_OUTBOX_DIR / f".{job_name}.tmp"
    payload = {
        "data": data,
        "bulk_import_id": bulk_import_id,
        "refresh_args": list(refresh_args) if refresh_args else None,
    }
    with temporary.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, separators=(",", ":"))
        file.flush()
        os.fsync(file.fileno())
    os.replace(temporary, destination)
    return destination


def _persist_queued_scan(outbox_file):
    """Send one durable scan job to VPS; keep its file when a retry is needed."""
    try:
        with Path(outbox_file).open("r", encoding="utf-8") as file:
            payload = json.load(file)
        saved = save_sqlite(
            payload["data"],
            bulk_import_id=payload.get("bulk_import_id")
        )
        if not saved:
            print(f"SCAN VPS TERTUNDA: {Path(outbox_file).name}")
            return False
        Path(outbox_file).unlink(missing_ok=True)
        refresh_args = payload.get("refresh_args")
        if refresh_args:
            EXECUTOR.submit(refresh_single_live_scan_tracking, *refresh_args)
        return True
    except Exception as error:
        print(f"SCAN VPS TERTUNDA: {Path(outbox_file).name} | {error}")
        return False


def queue_scan_save(data, bulk_import_id=None, refresh_args=None):
    """Queue a scanner save without risking data loss when VPS is unavailable."""
    try:
        outbox_file = _write_scan_outbox(data, bulk_import_id, refresh_args)
        SCAN_SAVE_EXECUTOR.submit(_persist_queued_scan, outbox_file)
        return True
    except Exception as error:
        # A full/read-only disk must never silently drop a scan. Fall back to
        # the old synchronous write in that exceptional case.
        print(f"SCAN OUTBOX ERROR, SIMPAN LANGSUNG: {error}")
        return save_sqlite(data, bulk_import_id=bulk_import_id)


def replay_pending_scan_outbox():
    """Retry scans that were queued before a power outage or VPS interruption."""
    if not SCAN_OUTBOX_DIR.is_dir():
        return
    for outbox_file in sorted(SCAN_OUTBOX_DIR.glob("*.json")):
        SCAN_SAVE_EXECUTOR.submit(_persist_queued_scan, outbox_file)
# === SAVE LOG ===
def save_excel(data, status, implant):
    today_folder = os.path.join(OUTPUT_DIR, datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(today_folder, exist_ok=True)
    if status == "INCOMING":
        filename = "INCOMING.xlsx"

    elif status == "PROBLEM":
        filename = "PROBLEM.xlsx"

    else:
        filename = f"OUTGOING/{implant}.xlsx"
    file_path = os.path.join(today_folder, filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    columns = [
        "RESI",
        "DP OUT",
        "SELLER",
        "SPRINTER",
        "LAST STATION",
        "LAST STATUS",
        "HARGA",
        "WAKTU SCAN",
        "BARANG",
        "WAKTU",
        "STATUS",
        "SPOT"
    ]

    df = pd.DataFrame([data], columns=columns)

    if os.path.exists(file_path):
        df_existing = pd.read_excel(file_path)
        df = pd.concat([df_existing, df], ignore_index=True)

    df.to_excel(file_path, index=False)

def save_error(msg):
    log_folder = os.path.join(OUTPUT_DIR, datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(log_folder, exist_ok=True)
    with open(os.path.join(log_folder, 'error.log'), 'a', encoding='utf-8') as f:
        f.write(f"[{datetime.now()}] {msg}\n")
        
        
def excel_writer():
    pass
 

# =========================
# REFRESH BADGES SCAN
# =========================

def refresh_scan_badges(resi):

    try:

        rows = cursor.execute("""

        SELECT type

        FROM manual_flags

        WHERE resi=?

        """,(resi,)).fetchall()

        badges = []

        manual_types = [

            x[0]

            for x in rows

        ]

        # COMPLAINT
        if "COMPLAINT" in manual_types:

            badges.append(
                "COMPLAINT"
            )

        # CLAIM INTERNAL
        if (

            "CLAIM_INTERNAL" in manual_types

            or

            "CLAIM INTERNAL" in manual_types

        ):

            badges.append(
                "CLAIM INTERNAL"
            )

        badge_text = ",".join(
            badges
        )

        # UPDATE SEMUA SCAN LAMA
        cursor.execute("""

        UPDATE scans

        SET badges =

            CASE

                WHEN badges IS NULL
                THEN ?

                ELSE
                    TRIM(

                        REPLACE(

                            REPLACE(

                                REPLACE(

                                    badges,

                                    'COMPLAINT',

                                    ''

                                ),

                                'CLAIM INTERNAL',

                                ''

                            ),

                            ',,',

                            ','

                        ),

                        ','

                    )

                    ||

                    CASE

                        WHEN ? != ''
                        THEN ',' || ?

                        ELSE ''

                    END

            END

        WHERE resi=?

        """,(

            badge_text,

            badge_text,

            badge_text,

            resi

        ))

        conn.commit()

    except Exception as e:

        print(
            "REFRESH BADGE ERROR:",
            e
        )

@app.route('/admin')
def admin():
    return render_template('admin.html')
    
@app.route('/admin_panel')
def admin_panel():
    
    if not check_login():
        return redirect('/')
    
    return render_template('admin_panel.html')
    
@app.route('/account_management')
def account_management():

    if not check_login():
        return redirect('/')

    if not has_permission(
        "ACCOUNT_MANAGEMENT"
    ):
        return redirect('/')

    cursor.execute("""
    SELECT
        id,
        fullname,
        username,
        password,
        level,
        permissions,
        must_change_credentials,
        attendance_bypass,
        iphone_user,
        disable_location_lock
    FROM users
    WHERE UPPER(level) != ?
    AND LOWER(username) != ?
    ORDER BY id DESC
    """,(
        "SUPERMAN",
        "superman"
    ))

    users = cursor.fetchall()

    total_user = len(users)

    superman_count = 0

    manager_count = len([
        x for x in users
        if x[4] == "MANAGER"
    ])

    processing_count = len([
        x for x in users
        if x[4] == "PROCESSING"
    ])

    driver_count = len([
        x for x in users
        if x[4] == "DRIVER"
    ])

    return render_template(

        'account_management.html',

        users=users,

        total_user=total_user,

        superman_count=superman_count,

        manager_count=manager_count,

        processing_count=processing_count,

        driver_count=driver_count,
        attendance_app_version=get_attendance_app_version(),
        attendance_locations=get_attendance_locations()

    )
    
@app.route('/manual_list')
def manual_list():

    search = request.args.get(
        "search",
        ""
    ).strip()

    filter_type = request.args.get(
        "type",
        "ALL"
    )

    query = """

    SELECT
        id,
        resi,
        type,
        seller,
        nominal,
        tanggal,
        photo,
        created_at,
        keterangan

    FROM manual_flags

    WHERE 1=1

    """

    params = []

    # SEARCH
    if search:

        query += """

        AND (

            resi LIKE ?
            OR seller LIKE ?

        )

        """

        params.extend([

            f"%{search}%",
            f"%{search}%"

        ])

    # FILTER
    if filter_type != "ALL":

        query += """

        AND type=?

        """

        params.append(filter_type)

    page = int(
        request.args.get(
            "page",
            1
        )
    )

    limit = 15

    offset = (page - 1) * limit

    count_query = query.replace(

        """

        SELECT
            id,
            resi,
            type,
            seller,
            nominal,
            tanggal,
            photo,
            created_at,
            keterangan

        """,

        "SELECT COUNT(*)"

    )

    with DB_LOCK:

        total = cursor.execute(
            count_query,
            params
        ).fetchone()[0]

    query += """

    ORDER BY id DESC

    LIMIT ? OFFSET ?

    """

    params.extend([

        limit,
        offset

    ])

    with DB_LOCK:

        rows = cursor.execute(
            query,
            params
        ).fetchall()

    data = []

    for row in rows:

        data.append({

            "id":row[0],
            "resi":row[1],
            "type":row[2],
            "seller":row[3],
            "nominal":row[4],
            "tanggal":row[5],
            "photo":row[6],
            "created_at":row[7],
            "keterangan":row[8]

        })

    return jsonify({

        "data":data,

        "pagination":{

            "page":page,

            "limit":limit,

            "total":total,

            "total_pages":
                (total + limit - 1)
                // limit

        }

    })
    
@app.route('/manual_add', methods=['POST'])
def manual_add():

    try:

        type_data = request.form.get("type")

        resi_text = request.form.get(
            "resi",
            ""
        )

        seller = request.form.get(
            "seller"
        )

        nominal = request.form.get(
            "nominal"
        )

        tanggal = request.form.get(
            "tanggal"
        )

        keterangan = request.form.get(
            "keterangan"
        )

        photo = request.files.get(
            "photo"
        )

        photo_name = ""

        if photo and photo.filename:

            upload_dir = os.path.join(
                BASE_DIR,
                "static",
                "manual_upload"
            )

            os.makedirs(
                upload_dir,
                exist_ok=True
            )

            filename = (
                f"{int(time.time())}_{photo.filename}"
                .replace(" ","")
                .replace("\n","")
                .replace("\r","")
            )

            save_path = os.path.join(
                upload_dir,
                filename
            )

            photo.save(save_path)

            photo_name = filename

        # MULTI RESI
        resi_list = [

            r.strip()

            for r in
            resi_text.splitlines()

            if r.strip()

        ]

        with DB_LOCK:

            for resi in resi_list:

                cursor.execute("""

                INSERT INTO manual_flags (

                    resi,
                    type,
                    seller,
                    nominal,
                    tanggal,
                    photo,
                    created_at,
                    keterangan

                )

                VALUES (?,?,?,?,?,?,?,?)

                """,(

                    resi,
                    type_data,
                    seller,
                    nominal,
                    tanggal,
                    photo_name,

                    datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                    keterangan

                ))

            conn.commit()

            # ====================
            # REFRESH BADGE SCAN
            # ====================

            for resi in resi_list:

                refresh_scan_badges(
                    resi
                )

        return jsonify({

            "success":True,
            "total_insert":
                len(resi_list)

        })

    except Exception as e:

        return jsonify({

            "success":False,
            "error":str(e)

        })
        
@app.route('/manual_delete', methods=['POST'])
def manual_delete():

    try:

        ids = request.json.get("ids", [])

        with DB_LOCK:

            for id_data in ids:

                # ====================
                # AMBIL RESI DULU
                # ====================

                row = cursor.execute("""

                SELECT resi

                FROM manual_flags

                WHERE id=?

                """,(id_data,)).fetchone()

                resi = ""

                if row:

                    resi = row[0]

                # ====================
                # DELETE
                # ====================

                cursor.execute("""

                DELETE FROM manual_flags

                WHERE id=?

                """,(id_data,))

                # ====================
                # REFRESH BADGE
                # ====================

                if resi:

                    refresh_scan_badges(
                        resi
                    )

            conn.commit()

        return jsonify({

            "success":True

        })

    except Exception as e:

        return jsonify({

            "success":False,
            "error":str(e)

        })

@app.route('/manual_update', methods=['POST'])
def manual_update():

    try:

        id_data = request.form.get("id")

        type_data = request.form.get("type")
        resi = request.form.get("resi")
        seller = request.form.get("seller")
        nominal = request.form.get("nominal")
        tanggal = request.form.get("tanggal")
        keterangan = request.form.get("keterangan")

        photo = request.files.get("photo")

        old_photo = cursor.execute("""

        SELECT photo
        FROM manual_flags
        WHERE id=?

        """,(id_data,)).fetchone()

        photo_name = old_photo[0] if old_photo else ""

        if photo and photo.filename:

            upload_dir = os.path.join(
                BASE_DIR,
                "static",
                "manual_upload"
            )

            os.makedirs(upload_dir, exist_ok=True)

            filename = (
                f"{int(time.time())}_{photo.filename}"
                .replace(" ","")
                .replace("\n","")
                .replace("\r","")
            )

            save_path = os.path.join(
                upload_dir,
                filename
            )

            photo.save(save_path)

            photo_name = filename

        with DB_LOCK:

            cursor.execute("""

            UPDATE manual_flags

            SET

                resi=?,
                type=?,
                seller=?,
                nominal=?,
                tanggal=?,
                photo=?,
                keterangan=?,
                updated_at=?

            WHERE id=?

            """,(

                resi,
                type_data,
                seller,
                nominal,
                tanggal,
                photo_name,
                keterangan,

                datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),

                id_data

            ))

            conn.commit()

            # ====================
            # REFRESH BADGE SCAN
            # ====================

            refresh_scan_badges(
                resi
            )

        return jsonify({

            "success":True

        })

    except Exception as e:

        return jsonify({

            "success":False,
            "error":str(e)

        })
  

# =========================
# ATTENDANCE CHECK
# =========================


def already_attendance(username):

    today = datetime.now().strftime(
        "%Y-%m-%d"
    )

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:

        row = conn_local.execute("""

        SELECT id

        FROM attendance

        WHERE username=?
        AND tanggal=?

        LIMIT 1

        """,(

            username,
            today

        )).fetchone()

    finally:

        conn_local.close()

    return row is not None

def get_today_attendance(username):

    today = datetime.now().strftime(
        "%Y-%m-%d"
    )

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:

        row = conn_local.execute("""

        SELECT
            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            clock_out,
            clock_out_photo,
            clock_out_latitude,
            clock_out_longitude,
            clock_out_address,
            shift_id

        FROM attendance

        WHERE username=?
        AND tanggal=?

        ORDER BY id DESC

        LIMIT 1

        """,(

            username,
            today

        )).fetchone()

    except sqlite3.OperationalError:

        row = conn_local.execute("""

        SELECT
            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            '' AS clock_out,
            '' AS clock_out_photo,
            '' AS clock_out_latitude,
            '' AS clock_out_longitude,
            '' AS clock_out_address,
            '' AS shift_id

        FROM attendance

        WHERE username=?
        AND tanggal=?

        ORDER BY id DESC

        LIMIT 1

        """,(

            username,
            today

        )).fetchone()

    conn_local.close()

    if not row:

        return None

    photo = row[4] or ""
    clock_in = row[3] or ""
    clock_out = row[8] or ""
    clock_out_photo = row[9] or ""

    return attach_attendance_shift_meta({
        "username": row[0] or "",
        "fullname": row[1] or "",
        "tanggal": row[2] or "",
        "jam": clock_in,
        "clock_in": clock_in,
        "clock_out": clock_out,
        "photo": photo,
        "photo_url": (
            "/static/uploads/attendance/" + photo
            if photo
            else ""
        ),
        "latitude": row[5] or "",
        "longitude": row[6] or "",
        "address": row[7] or "",
        "clock_out_photo": clock_out_photo,
        "clock_out_photo_url": (
            "/static/uploads/attendance/" + clock_out_photo
            if clock_out_photo
            else ""
        ),
        "clock_out_latitude": row[10] or "",
        "clock_out_longitude": row[11] or "",
        "clock_out_address": row[12] or "",
        "status": "Berhasil",
        "shift_id": row[13] if len(row) > 13 else ""
    })

def attendance_row_to_dict(row):

    if not row:
        return None

    photo = row[5] or ""
    clock_in = row[4] or ""
    clock_out = row[9] or ""
    clock_out_photo = row[10] or ""

    return {
        "id": row[0],
        "username": row[1] or "",
        "fullname": row[2] or "",
        "tanggal": row[3] or "",
        "jam": clock_in,
        "clock_in": clock_in,
        "clock_out": clock_out,
        "photo": photo,
        "photo_url": (
            "/static/uploads/attendance/" + photo
            if photo
            else ""
        ),
        "latitude": row[6] or "",
        "longitude": row[7] or "",
        "address": row[8] or "",
        "clock_out_photo": clock_out_photo,
        "clock_out_photo_url": (
            "/static/uploads/attendance/" + clock_out_photo
            if clock_out_photo
            else ""
        ),
        "clock_out_latitude": row[11] or "",
        "clock_out_longitude": row[12] or "",
        "clock_out_address": row[13] or "",
        "status": "Berhasil",
        "shift_id": row[14] if len(row) > 14 else "",
        "face_score": int(row[15] or 0) if len(row) > 15 else 0,
        "clock_out_face_score": int(row[16] or 0) if len(row) > 16 else 0,
        "device_info": row[17] or "" if len(row) > 17 else "",
        "clock_out_device_info": row[18] or "" if len(row) > 18 else "",
        "ip_address": row[19] or "" if len(row) > 19 else "",
        "clock_out_ip_address": row[20] or "" if len(row) > 20 else "",
        "app_version": row[21] or "" if len(row) > 21 else ""
    }



def get_attendance_base_cutoff_pair(username, settings=None):

    settings = settings or get_attendance_settings()
    default_in = settings.get("clock_in_cutoff") or "07:00"
    default_out = settings.get("clock_out_cutoff") or "12:00"

    raw_username = str(username or "").strip()
    user_cutoffs = settings.get("user_cutoffs") or {}
    user_value = None

    if isinstance(user_cutoffs, dict) and raw_username:
        user_value = (
            user_cutoffs.get(raw_username) or
            user_cutoffs.get(raw_username.lower()) or
            user_cutoffs.get(raw_username.upper())
        )
        if user_value is None:
            normalized_username = raw_username.strip().lower()
            for key, value in user_cutoffs.items():
                if str(key or "").strip().lower() == normalized_username:
                    user_value = value
                    break

    if isinstance(user_value, dict):
        return {
            "in": (
                user_value.get("in") or
                user_value.get("clock_in") or
                user_value.get("clock_in_cutoff") or
                default_in
            ),
            "out": (
                user_value.get("out") or
                user_value.get("clock_out") or
                user_value.get("clock_out_cutoff") or
                default_out
            )
        }

    if isinstance(user_value, str) and user_value.strip():
        return {
            "in": user_value.strip(),
            "out": default_out
        }

    return {
        "in": default_in,
        "out": default_out
    }


def get_attendance_shift_display(username):

    shift = get_attendance_shift(username)
    base = get_attendance_base_cutoff_pair(username)

    if not shift:
        shift = {
            "username": username or "",
            "total_shift": 1,
            "waktu_shift": ""
        }

    shift = dict(shift)
    shift["shift1_clock_in"] = shift.get("shift1_clock_in") or base["in"]
    shift["shift1_clock_out"] = shift.get("shift1_clock_out") or base["out"]
    shift["shift2_clock_in"] = shift.get("shift2_clock_in") or base["in"]
    shift["shift2_clock_out"] = shift.get("shift2_clock_out") or base["out"]

    return shift


def get_attendance_shift_meta(username, shift_id=None):

    settings = get_attendance_settings()
    shift = get_attendance_shift(username)
    base = get_attendance_base_cutoff_pair(username, settings)

    raw_shift_id = str(shift_id or "").strip()
    if raw_shift_id not in {"1", "2"}:
        raw_shift_id = "1"

    try:
        total_shift = int(shift.get("total_shift") or 1)
    except Exception:
        total_shift = 1

    if total_shift < 2:
        start = base["in"]
        end = base["out"]
        label = "Jam Kerja"
    elif raw_shift_id == "2":
        start = shift.get("shift2_clock_in") or base["in"]
        end = shift.get("shift2_clock_out") or base["out"]
        label = "Shift 2"
    else:
        start = shift.get("shift1_clock_in") or base["in"]
        end = shift.get("shift1_clock_out") or base["out"]
        label = "Shift 1"

    return {
        "id": raw_shift_id,
        "label": label,
        "start": start,
        "end": end,
        "range": start + " - " + end
    }

def attach_attendance_shift_meta(attendance):

    if not attendance:
        return attendance

    meta = get_attendance_shift_meta(
        attendance.get("username", ""),
        attendance.get("shift_id", "")
    )

    attendance["shift_id"] = meta["id"]
    attendance["shift_label"] = meta["label"]
    attendance["shift_start"] = meta["start"]
    attendance["shift_end"] = meta["end"]
    attendance["shift_range"] = meta["range"]

    return attendance

def get_attendance_by_id(attendance_id):

    if not attendance_id:
        return None

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        row = conn_local.execute("""

        SELECT
            id,
            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            clock_out,
            clock_out_photo,
            clock_out_latitude,
            clock_out_longitude,
            clock_out_address,
            shift_id,
            COALESCE(face_score, 0),
            COALESCE(clock_out_face_score, 0),
            COALESCE(device_info, ''),
            COALESCE(clock_out_device_info, ''),
            COALESCE(ip_address, ''),
            COALESCE(clock_out_ip_address, ''),
            COALESCE(app_version, '')

        FROM attendance
        WHERE id=?
        LIMIT 1

        """,(
            attendance_id,
        )).fetchone()
    except sqlite3.OperationalError:
        row = None
    finally:
        conn_local.close()

    return attach_attendance_shift_meta(attendance_row_to_dict(row))


def get_pending_clock_out_attendance(username):

    if not username:
        return None

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        row = conn_local.execute("""

        SELECT
            id,
            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            clock_out,
            clock_out_photo,
            clock_out_latitude,
            clock_out_longitude,
            clock_out_address,
            shift_id,
            COALESCE(face_score, 0),
            COALESCE(clock_out_face_score, 0),
            COALESCE(device_info, ''),
            COALESCE(clock_out_device_info, ''),
            COALESCE(ip_address, ''),
            COALESCE(clock_out_ip_address, ''),
            COALESCE(app_version, '')

        FROM attendance
        WHERE username=?
        AND (
            clock_out IS NULL
            OR TRIM(clock_out)=''
        )
        ORDER BY tanggal DESC, id DESC
        LIMIT 1

        """,(
            username,
        )).fetchone()
    except sqlite3.OperationalError:
        row = conn_local.execute("""

        SELECT
            id,
            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            '' AS clock_out,
            '' AS clock_out_photo,
            '' AS clock_out_latitude,
            '' AS clock_out_longitude,
            '' AS clock_out_address,
            '' AS shift_id,
            0 AS face_score,
            0 AS clock_out_face_score,
            '' AS device_info,
            '' AS clock_out_device_info,
            '' AS ip_address,
            '' AS clock_out_ip_address,
            '' AS app_version

        FROM attendance
        WHERE username=?
        ORDER BY tanggal DESC, id DESC
        LIMIT 1

        """,(
            username,
        )).fetchone()
    finally:
        conn_local.close()

    return attach_attendance_shift_meta(attendance_row_to_dict(row))

def get_latest_attendance(username):

    if not username:
        return None

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        row = conn_local.execute("""

        SELECT
            id,
            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            clock_out,
            clock_out_photo,
            clock_out_latitude,
            clock_out_longitude,
            clock_out_address,
            shift_id,
            COALESCE(face_score, 0),
            COALESCE(clock_out_face_score, 0),
            COALESCE(device_info, ''),
            COALESCE(clock_out_device_info, ''),
            COALESCE(ip_address, ''),
            COALESCE(clock_out_ip_address, ''),
            COALESCE(app_version, '')

        FROM attendance
        WHERE username=?
        ORDER BY tanggal DESC, id DESC
        LIMIT 1

        """,(
            username,
        )).fetchone()
    except sqlite3.OperationalError:
        row = None
    finally:
        conn_local.close()

    return attach_attendance_shift_meta(attendance_row_to_dict(row))
def format_attendance_date(value):

    if not value:

        return "-"

    try:

        parsed = datetime.strptime(
            value,
            "%Y-%m-%d"
        )

    except Exception:

        return value

    return parsed.strftime(
        "%d %b %Y"
    )

def format_attendance_day(value):

    days = [
        "Senin",
        "Selasa",
        "Rabu",
        "Kamis",
        "Jumat",
        "Sabtu",
        "Minggu"
    ]

    try:

        parsed = datetime.strptime(
            value,
            "%Y-%m-%d"
        )

    except Exception:

        return "-"

    return days[
        parsed.weekday()
    ]

def normalize_attendance_history_month(month_value=None):

    today = datetime.now()
    current_month = today.replace(
        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    raw_month = str(month_value or "").strip()

    try:
        month_start = datetime.strptime(
            raw_month,
            "%Y-%m"
        )
    except Exception:
        month_start = current_month

    month_start = month_start.replace(
        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    if month_start > current_month:
        month_start = current_month

    return month_start.strftime("%Y-%m"), month_start


def get_attendance_history(username, month=None):
    selected_month, month_start = normalize_attendance_history_month(month)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    month_end = next_month - timedelta(days=1)
    end_date = min(month_end, today)
    if end_date < month_start:
        end_date = month_start
    days = (end_date - month_start).days + 1

    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
    try:
        try:
            rows = conn_local.execute("""
                SELECT tanggal, jam, address, latitude, longitude, photo,
                       clock_out, shift_id, clock_out_address,
                       clock_out_latitude, clock_out_longitude, clock_out_photo,
                       COALESCE(face_score, 0), COALESCE(clock_out_face_score, 0),
                       device_info, clock_out_device_info, ip_address,
                       clock_out_ip_address, app_version
                FROM attendance
                WHERE username=? AND tanggal BETWEEN ? AND ?
                ORDER BY tanggal DESC, id DESC
            """, (
                username,
                month_start.strftime("%Y-%m-%d"),
                end_date.strftime("%Y-%m-%d")
            )).fetchall()
        except sqlite3.OperationalError:
            rows = conn_local.execute("""
                SELECT tanggal, jam, address, latitude, longitude, photo,
                       clock_out, shift_id, clock_out_address,
                       clock_out_latitude, clock_out_longitude, clock_out_photo,
                       0, 0, '', '', '', '', ''
                FROM attendance
                WHERE username=? AND tanggal BETWEEN ? AND ?
                ORDER BY tanggal DESC, id DESC
            """, (
                username,
                month_start.strftime("%Y-%m-%d"),
                end_date.strftime("%Y-%m-%d")
            )).fetchall()
    finally:
        conn_local.close()

    by_date = {}
    for row in rows:
        if row[0] not in by_date:
            by_date[row[0]] = row

    leave_data = get_attendance_leave_history(username, selected_month)
    leave_by_date = {row["date_key"]: row for row in leave_data["rows"]}
    shift_meta_cache = {}

    def history_shift_meta(shift_id):
        cache_key = str(shift_id or "1")
        if cache_key not in shift_meta_cache:
            shift_meta_cache[cache_key] = get_attendance_shift_meta(username, cache_key)
        return shift_meta_cache[cache_key]

    default_shift = history_shift_meta("1")
    history = []
    present_count = 0
    late_count = 0
    absent_count = 0
    leave_count = 0

    for offset in range(days):
        current = end_date - timedelta(days=offset)
        date_key = current.strftime("%Y-%m-%d")
        row = by_date.get(date_key)
        leave_row = leave_by_date.get(date_key)

        if row:
            shift_meta = history_shift_meta(row[7] or "1")
            status = "On Time"
            status_class = "present"
            try:
                clock_in_minutes = int(str(row[1] or "00:00")[:2]) * 60 + int(str(row[1] or "00:00")[3:5])
                cutoff_minutes = int(shift_meta["start"][:2]) * 60 + int(shift_meta["start"][3:5])
                if clock_in_minutes > cutoff_minutes:
                    status = "Late"
                    status_class = "late"
                    late_count += 1
            except Exception:
                pass
            present_count += 1
            photo = row[5] or ""
            clock_out_photo = row[11] or ""
            coordinates = (str(row[3]) + ", " + str(row[4])) if row[3] and row[4] else ""
            clock_out_coordinates = (str(row[9]) + ", " + str(row[10])) if row[9] and row[10] else ""
            history.append({
                "date_key": date_key,
                "tanggal": format_attendance_date(date_key),
                "hari": format_attendance_day(date_key),
                "jam": row[1] or "-",
                "clock_out": row[6] or "-",
                "zona": "WIB",
                "lokasi": row[2] or "-",
                "address": row[2] or "-",
                "koordinat": coordinates,
                "clock_out_lokasi": row[8] or "-",
                "clock_out_koordinat": clock_out_coordinates,
                "status": status,
                "status_class": status_class,
                "metode": "GPS",
                "shift": shift_meta["label"],
                "shift_range": shift_meta["range"],
                "photo_url": "/static/uploads/attendance/" + photo if photo else "",
                "clock_out_photo_url": "/static/uploads/attendance/" + clock_out_photo if clock_out_photo else "",
                "face_score": int(row[12] or 0),
                "clock_out_face_score": int(row[13] or 0),
                "device_info": row[14] or "-",
                "clock_out_device_info": row[15] or "-",
                "ip_address": row[16] or "-",
                "clock_out_ip_address": row[17] or "-",
                "app_version": row[18] or "-"
            })
        elif leave_row:
            leave_count += 1
            history.append({
                "date_key": date_key,
                "tanggal": format_attendance_date(date_key),
                "hari": format_attendance_day(date_key),
                "jam": "-",
                "clock_out": "-",
                "zona": "-",
                "lokasi": "-",
                "koordinat": "",
                "status": leave_row["type"].title(),
                "status_class": "leave",
                "metode": "-",
                "shift": default_shift["label"],
                "shift_range": default_shift["range"]
            })
        else:
            absent_count += 1
            history.append({
                "date_key": date_key,
                "tanggal": format_attendance_date(date_key),
                "hari": format_attendance_day(date_key),
                "jam": "-",
                "clock_out": "-",
                "zona": "-",
                "lokasi": "-",
                "koordinat": "",
                "status": "Tidak Hadir",
                "status_class": "absent",
                "metode": "-",
                "shift": default_shift["label"],
                "shift_range": default_shift["range"]
            })

    on_time_count = max(0, present_count - late_count)
    measured_days = present_count + absent_count
    attendance_rate = int(round((present_count * 100.0) / measured_days)) if measured_days else 0
    return {
        "rows": history,
        "range": format_attendance_date(month_start.strftime("%Y-%m-%d")) + " - " + format_attendance_date(end_date.strftime("%Y-%m-%d")),
        "month": selected_month,
        "count": len(history),
        "present_count": present_count,
        "on_time_count": on_time_count,
        "late_count": late_count,
        "absent_count": absent_count,
        "leave_count": leave_count,
        "attendance_rate": attendance_rate,
        "today": history[0] if history and history[0].get("date_key") == today.strftime("%Y-%m-%d") else {}
    }


def export_attendance_history_csv(username, month=None):

    history = get_attendance_history(
        username,
        month=month
    )

    buffer = StringIO()

    buffer.write(
        "Tanggal,Hari,Jam,Zona,Lokasi,Koordinat,Status,Metode\n"
    )

    for row in history["rows"]:

        values = [
            row["tanggal"],
            row["hari"],
            row["jam"],
            row["zona"],
            row["lokasi"],
            row["koordinat"],
            row["status"],
            row["metode"]
        ]

        escaped = [
            '"' + str(value).replace('"', '""') + '"'
            for value in values
        ]

        buffer.write(
            ",".join(escaped) + "\n"
        )

    output = BytesIO(
        buffer.getvalue().encode(
            "utf-8-sig"
        )
    )

    output.seek(0)

    return output

def render_attendance_success_page(attendance, notice=None):

    pending_clock_out = has_pending_clock_out(
        attendance
    )

    redirect_url = (
        ""
        if pending_clock_out
        else "/logout"
    )

    return render_template(
        'attendance_success.html',
        attendance=attendance or {},
        attendance_only=True,
        redirect_url=redirect_url,
        redirect_text="logout ke main menu",
        auto_redirect=not pending_clock_out,
        redirect_seconds=5,
        notice=notice or session.pop("attendance_notice", "")
    )

def is_superman_level(level):

    return str(
        level or ""
    ).upper() == "SUPERMAN"

def truthy_flag(value):

    return str(
        value or ""
    ).strip().lower() in {
        "1",
        "true",
        "yes",
        "on"
    }

def requires_attendance_for_level(level, attendance_bypass=False):

    return (
        not is_superman_level(
            level
        ) and
        not truthy_flag(
            attendance_bypass
        )
    )

def clear_pending_login_session():

    session.pop(
        "pending_username",
        None
    )

    session.pop(
        "pending_fullname",
        None
    )

    session.pop(
        "pending_level",
        None
    )

    session.pop(
        "pending_permissions",
        None
    )

    session.pop(
        "pending_iphone_user",
        None
    )

    session.pop(
        "pending_attendance_bypass",
        None
    )
    

def get_user_attendance_bypass_status(username):

    if not username:
        return False

    try:
        with DB_LOCK:
            columns = [
                row[1]
                for row in cursor.execute(
                    "PRAGMA table_info(users)"
                ).fetchall()
            ]

            select_columns = "level, permissions"
            if "attendance_bypass" in columns:
                select_columns += ", attendance_bypass"

            row = cursor.execute(f"""
                SELECT {select_columns}
                FROM users
                WHERE username=?
                LIMIT 1
            """, (
                username,
            )).fetchone()

        if not row:
            return False

        level = row[0]
        permissions = row[1] or ""
        attendance_bypass = row[2] if len(row) > 2 else 0

        return (
            is_superman_level(level) or
            truthy_flag(attendance_bypass) or
            has_attendance_bypass_permission(permissions)
        )

    except Exception:
        return False


def promote_pending_login_if_attendance_bypass():

    if session.get("logged_in"):
        return True

    username = session.get("pending_username")
    if not username:
        return False

    if not (
        truthy_flag(
            session.get("pending_attendance_bypass", 0)
        ) or
        get_user_attendance_bypass_status(username)
    ):
        return False

    session["logged_in"] = True
    session["username"] = username
    session["fullname"] = session.get(
        "pending_fullname",
        username
    )
    session["level"] = session.get(
        "pending_level",
        "ADMIN"
    )
    session["permissions"] = session.get(
        "pending_permissions",
        []
    )
    session["attendance_bypass"] = 1
    clear_pending_login_session()
    return True


def is_mobile_request():

    ua = request.headers.get(
        "User-Agent",
        ""
    ).lower()

    mobile_keywords = [

        "android",
        "iphone",
        "ipad",
        "mobile"

    ]

    return any(
        x in ua
        for x in mobile_keywords
    )


def get_user_iphone_user_status(username):

    if not username:
        return False

    try:
        with DB_LOCK:
            row = cursor.execute("""
                SELECT iphone_user
                FROM users
                WHERE username=?
                LIMIT 1
            """, (username,)).fetchone()

        return bool(row and truthy_flag(row[0]))

    except Exception:
        return False


def get_user_disable_location_lock_status(username):

    if not username:
        return False

    try:
        with DB_LOCK:
            row = cursor.execute("""
                SELECT disable_location_lock
                FROM users
                WHERE username=?
                LIMIT 1
            """, (username,)).fetchone()

        return bool(row and truthy_flag(row[0]))

    except Exception:
        return False


def get_last_attendance_photo(username):

    if not username:
        return ""

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        row = conn_local.execute("""

            SELECT
                COALESCE(NULLIF(TRIM(clock_out_photo), ''), NULLIF(TRIM(photo), '')) AS last_photo

            FROM attendance

            WHERE username=?
            AND COALESCE(NULLIF(TRIM(clock_out_photo), ''), NULLIF(TRIM(photo), '')) IS NOT NULL

            ORDER BY tanggal DESC, id DESC

            LIMIT 1

        """,(
            username,
        )).fetchone()

        return row[0] if row and row[0] else ""

    except Exception:
        return ""

    finally:
        conn_local.close()

def get_last_attendance_photo_url(username):

    photo = get_last_attendance_photo(
        username
    )

    if photo:
        return "/static/uploads/attendance/" + photo

    return "/static/uploads/attendance/nophoto.png"

def row_to_attendance_shift(row):

    if not row:
        return {}

    return {
        "username": row[0],
        "total_shift": int(row[1] or 1),
        "waktu_shift": row[2] or "",
        "shift1_clock_in": row[3] or "",
        "shift1_clock_out": row[4] or "",
        "shift2_clock_in": row[5] or "",
        "shift2_clock_out": row[6] or ""
    }

def get_attendance_shift(username):

    if not username:
        return {}

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        row = conn_local.execute("""

            SELECT
                username,
                total_shift,
                waktu_shift,
                shift1_clock_in,
                shift1_clock_out,
                shift2_clock_in,
                shift2_clock_out

            FROM attendance_shifts

            WHERE LOWER(username)=LOWER(?)

            LIMIT 1

        """,(
            username,
        )).fetchone()

        return row_to_attendance_shift(row)

    except Exception:
        return {}

    finally:
        conn_local.close()

def get_all_attendance_shifts():

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        rows = conn_local.execute("""

            SELECT
                username,
                total_shift,
                waktu_shift,
                shift1_clock_in,
                shift1_clock_out,
                shift2_clock_in,
                shift2_clock_out

            FROM attendance_shifts

        """).fetchall()

        return {
            str(row[0] or "").lower(): row_to_attendance_shift(row)
            for row in rows
            if str(row[0] or "").strip()
        }

    except Exception:
        return {}

    finally:
        conn_local.close()
# === ROUTES ===

@app.route('/login', methods=['POST'])
def login():
    
    session.clear()

    data = request.json

    username = data.get("username")
    password = data.get("password")

    with DB_LOCK:

        conn_login = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:

            cur_login = conn_login.cursor()

            cur_login.execute("""

            SELECT *

            FROM users

            WHERE username=?
            AND password=?

            """, (

                username,
                password

            ))

            user = cur_login.fetchone()

        finally:

            conn_login.close()

    if user:

        if (
            len(user) > 6 and
            int(user[6] or 0) == 1
        ):

            session["pending_credentials_user_id"] = user[0]

            return jsonify({

                "success":True,

                "force_change_credentials":True,

                "need_attendance":False,

                "redirect":"/"

            })

        session["pending_username"] = user[2]

        session["pending_fullname"] = user[1]

        session["pending_level"] = user[4]
        
        session["pending_permissions"] = (

            user[5] or ""

        ).split(",")

        attendance_bypass = (
            user[7]
            if len(user) > 7
            else 0
        )
        attendance_bypass = (
            truthy_flag(
                attendance_bypass
            ) or
            has_attendance_bypass_permission(
                user[5] or ""
            )
        )

        session["pending_attendance_bypass"] = int(
            truthy_flag(
                attendance_bypass
            )
        )

        today_future = AUTH_EXECUTOR.submit(get_today_attendance, user[2])
        pending_future = AUTH_EXECUTOR.submit(get_pending_clock_out_attendance, user[2])
        today_attendance = today_future.result()
        pending_clock_out = pending_future.result()
        attendance_only = is_attendance_only_permissions(
            user[5] or ""
        )

        need_attendance = (
            requires_attendance_for_level(
                user[4],
                attendance_bypass
            ) and
            today_attendance is None and
            pending_clock_out is None
        )
        
        if not need_attendance:

            session["logged_in"] = True

            session["username"] = user[2]

            session["fullname"] = user[1]

            session["level"] = user[4]
            
            session["permissions"] = (

                user[5] or ""

            ).split(",")

            session["attendance_bypass"] = int(
                truthy_flag(
                    attendance_bypass
                )
            )

            if pending_clock_out:

                session["last_attendance_id"] = pending_clock_out.get("id")

                if (
                    attendance_only and
                    pending_clock_out.get("tanggal") != current_attendance_date()
                ):

                    session["attendance_notice"] = "Belum clock out di sesi sebelumnya. Clock out dulu untuk clock in tanggal hari ini."

            if attendance_only:

                mark_attendance_only_session_date()

            clear_pending_login_session()

        user_agent = request.headers.get(
            "User-Agent",
            ""
        ).lower()

        is_mobile = any(x in user_agent for x in [

            "android",
            "iphone",
            "ipad",
            "mobile"

        ])

        can_open_attendance_dashboard = (
            is_superman_level(user[4]) or
            "DASHBOARD_ABSEN" in normalize_permissions(user[5] or "")
        )

        redirect_url = "/"

        if (
            not need_attendance and
            (attendance_only or not can_open_attendance_dashboard) and
            (today_attendance or pending_clock_out)
        ):

            redirect_url = "/attendance_success"

        if need_attendance:

            if is_mobile:

                if get_user_iphone_user_status(user[2]):

                    redirect_url = "/attendance_mobile"

                else:

                    attendance_qr_token = create_attendance_qr_token("clock_in")
                    redirect_url = url_for(
                        'attendance_app_launcher',
                        token=attendance_qr_token
                    )

            else:

                redirect_url = "/attendance"
        return jsonify({

            "success":True,

            "need_attendance":need_attendance,

            "redirect":redirect_url,

            "attendance_only":
                attendance_only

        })

    return jsonify({
        "success":False
    })

@app.route('/api/app_login', methods=['POST'])
def api_app_login():

    session.clear()

    data = request.get_json(silent=True) or {}
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", "")).strip()

    if not username or not password:

        return jsonify({
            "success": False,
            "error": "Username dan password wajib diisi"
        }), 400

    with DB_LOCK:

        conn_login = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:

            user = conn_login.execute("""

                SELECT id, fullname, username, password, level, permissions,
                       must_change_credentials, attendance_bypass, iphone_user,
                       disable_location_lock
                FROM users
                WHERE username=?
                AND password=?

            """, (
                username,
                password
            )).fetchone()

        finally:

            conn_login.close()

    if not user:

        return jsonify({
            "success": False,
            "error": "Username atau password salah"
        }), 401

    if (
        len(user) > 6 and
        int(user[6] or 0) == 1
    ):

        return jsonify({
            "success": False,
            "force_change_credentials": True,
            "error": "Akun wajib ganti username/password dari web dulu"
        }), 403

    permissions_text = user[5] or ""
    attendance_bypass = (
        user[7]
        if len(user) > 7
        else 0
    )
    attendance_bypass = (
        truthy_flag(attendance_bypass) or
        has_attendance_bypass_permission(permissions_text)
    )
    permissions = permissions_text.split(",")
    attendance_only = is_attendance_only_permissions(permissions_text)

    login_started = time.perf_counter()
    today_future = AUTH_EXECUTOR.submit(get_today_attendance, user[2])
    pending_future = AUTH_EXECUTOR.submit(get_pending_clock_out_attendance, user[2])
    profile_future = AUTH_EXECUTOR.submit(get_app_profile_data, user[2])
    locations_future = AUTH_EXECUTOR.submit(get_attendance_locations)
    shift_future = AUTH_EXECUTOR.submit(get_attendance_shift_display, user[2])
    today_attendance = today_future.result()
    pending_clock_out = pending_future.result()

    need_attendance = (
        requires_attendance_for_level(
            user[4],
            attendance_bypass
        ) and
        today_attendance is None and
        pending_clock_out is None
    )

    session["pending_username"] = user[2]
    session["pending_fullname"] = user[1]
    session["pending_level"] = user[4]
    session["pending_permissions"] = permissions
    iphone_user = user[8] if len(user) > 8 else 0
    disable_location_lock = user[9] if len(user) > 9 else 0
    session["pending_iphone_user"] = int(truthy_flag(iphone_user))
    session["iphone_user"] = int(truthy_flag(iphone_user))
    session["disable_location_lock"] = int(truthy_flag(disable_location_lock))
    session["pending_attendance_bypass"] = int(truthy_flag(attendance_bypass))

    mode = "clock_in"
    active_attendance = {}

    if pending_clock_out:

        session["logged_in"] = True
        session["username"] = user[2]
        session["fullname"] = user[1]
        session["level"] = user[4]
        session["permissions"] = permissions
        session["iphone_user"] = int(truthy_flag(iphone_user))
        session["attendance_bypass"] = int(truthy_flag(attendance_bypass))
        session["last_attendance_id"] = pending_clock_out.get("id")
        mode = "clock_out"
        active_attendance = pending_clock_out

    elif not need_attendance:

        session["logged_in"] = True
        session["username"] = user[2]
        session["fullname"] = user[1]
        session["level"] = user[4]
        session["permissions"] = permissions
        session["iphone_user"] = int(truthy_flag(iphone_user))
        session["attendance_bypass"] = int(truthy_flag(attendance_bypass))

        if attendance_only:

            mark_attendance_only_session_date()

    web_url = url_for(
        'attendance_mobile',
        _external=True
    )

    app_profile = profile_future.result()
    attendance_locations = locations_future.result()
    attendance_shift = shift_future.result()

    response = jsonify({
        "success": True,
        "username": user[2],
        "fullname": user[1],
        "level": user[4] or "",
        "display_name": app_profile.get("display_name", ""),
        "department": app_profile.get("department", ""),
        "profile_photo_url": app_profile.get("profile_photo_url", ""),
        "mode": mode,
        "token": "",
        "sync_token": "",
        "web_url": web_url,
        "last_photo_url": "",
        "total_shift": int(attendance_shift.get("total_shift") or 1),
        "shift": attendance_shift,
        "need_attendance": need_attendance,
        "attendance_only": attendance_only,
        "disable_location_lock": bool(truthy_flag(disable_location_lock)),
        "active_attendance": active_attendance,
        "attendance_locations": attendance_locations
    })
    response.headers["Server-Timing"] = "app_login;dur={:.1f}".format(
        (time.perf_counter() - login_started) * 1000.0
    )
    return response

def get_attendance_app_version():
    try:
        build_file = os.path.join(
            BASE_DIR,
            "android_attendance",
            "app",
            "build.gradle"
        )
        with open(build_file, "r", encoding="utf-8") as handle:
            match = re.search(r'versionName\s+["\']([^"\']+)["\']', handle.read())
            if match:
                return match.group(1).strip()
    except Exception:
        pass
    return ""


def get_attendance_history_legacy(username, month=None):

    selected_month, month_start = normalize_attendance_history_month(
        month
    )

    today = datetime.now().replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    if month_start.month == 12:
        next_month = month_start.replace(
            year=month_start.year + 1,
            month=1
        )
    else:
        next_month = month_start.replace(
            month=month_start.month + 1
        )

    month_end = next_month - timedelta(
        days=1
    )

    end_date = min(
        month_end,
        today
    )

    if end_date < month_start:
        end_date = month_start

    days = (
        end_date - month_start
    ).days + 1

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    try:
        rows = conn_local.execute("""

        SELECT
            tanggal,
            jam,
            address,
            latitude,
            longitude,
            photo,
            clock_out,
            shift_id

        FROM attendance

        WHERE username=?
        AND tanggal BETWEEN ? AND ?

        ORDER BY tanggal DESC

        """,(

            username,
            month_start.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d")

        )).fetchall()
    except sqlite3.OperationalError:
        rows = conn_local.execute("""

        SELECT
            tanggal,
            jam,
            address,
            latitude,
            longitude,
            photo,
            '' AS clock_out,
            '' AS shift_id

        FROM attendance

        WHERE username=?
        AND tanggal BETWEEN ? AND ?

        ORDER BY tanggal DESC

        """,(

            username,
            month_start.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d")

        )).fetchall()

    conn_local.close()

    by_date = {
        row[0]: row
        for row in rows
    }

    history = []
    present_count = 0
    late_count = 0

    for offset in range(days):

        current = end_date - timedelta(
            days=offset
        )

        date_key = current.strftime(
            "%Y-%m-%d"
        )

        row = by_date.get(
            date_key
        )

        if row:

            location = row[2] or "-"
            coordinates = ""
            shift_meta = get_attendance_shift_meta(username, row[7] if len(row) > 7 else "")
            status = "Hadir"
            status_class = "present"

            try:
                clock_in_minutes = int(str(row[1] or "00:00")[:2]) * 60 + int(str(row[1] or "00:00")[3:5])
                cutoff_minutes = int(shift_meta["start"][:2]) * 60 + int(shift_meta["start"][3:5])
                if clock_in_minutes > cutoff_minutes:
                    status = "Terlambat"
                    status_class = "late"
                    late_count += 1
            except Exception:
                pass

            present_count += 1

            if row[3] and row[4]:

                coordinates = (
                    str(row[3]) +
                    ", " +
                    str(row[4])
                )

            history.append({
                "date_key": date_key,
                "tanggal": format_attendance_date(
                    date_key
                ),
                "hari": format_attendance_day(
                    date_key
                ),
                "jam": row[1] or "-",
                "clock_out": row[6] or "-",
                "zona": "WIB",
                "lokasi": location,
                "koordinat": coordinates,
                "status": status,
                "status_class": status_class,
                "metode": "GPS",
                "shift": shift_meta["label"],
                "shift_range": shift_meta["range"]
            })

        else:

            history.append({
                "date_key": date_key,
                "tanggal": format_attendance_date(
                    date_key
                ),
                "hari": format_attendance_day(
                    date_key
                ),
                "jam": "-",
                "clock_out": "-",
                "zona": "-",
                "lokasi": "-",
                "koordinat": "",
                "status": "Tidak Hadir",
                "status_class": "absent",
                "metode": "-",
                "shift": "-",
                "shift_range": "-"
            })

    return {
        "rows": history,
        "range": (
            format_attendance_date(
                month_start.strftime("%Y-%m-%d")
            ) +
            " - " +
            format_attendance_date(
                end_date.strftime("%Y-%m-%d")
            )
        ),
        "month": selected_month,
        "count": len(history),
        "present_count": present_count,
        "late_count": late_count,
        "absent_count": max(0, len(history) - present_count)
    }


def get_attendance_leave_history(username, month=None):
    selected_month, month_start = normalize_attendance_history_month(month)
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    month_end = next_month - timedelta(days=1)

    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
    try:
        try:
            rows = conn_local.execute("""
                SELECT tanggal, type, keterangan, photo,
                       COALESCE(NULLIF(TRIM(status), ''), 'Tercatat')
                FROM attendance_leave
                WHERE username=? AND tanggal BETWEEN ? AND ?
                ORDER BY tanggal DESC, id DESC
            """, (
                username,
                month_start.strftime("%Y-%m-%d"),
                month_end.strftime("%Y-%m-%d")
            )).fetchall()
        except sqlite3.OperationalError:
            rows = conn_local.execute("""
                SELECT tanggal, type, keterangan, photo, 'Tercatat'
                FROM attendance_leave
                WHERE username=? AND tanggal BETWEEN ? AND ?
                ORDER BY tanggal DESC, id DESC
            """, (
                username,
                month_start.strftime("%Y-%m-%d"),
                month_end.strftime("%Y-%m-%d")
            )).fetchall()
    finally:
        conn_local.close()

    result = []
    for row in rows:
        photo = row[3] or ""
        result.append({
            "date_key": row[0],
            "tanggal": format_attendance_date(row[0]),
            "hari": format_attendance_day(row[0]),
            "type": row[1] or "IZIN",
            "keterangan": row[2] or "-",
            "photo_url": (
                "/static/uploads/attendance_leave/" + photo
                if photo else ""
            ),
            "status": row[4] or "Tercatat"
        })
    return {
        "month": selected_month,
        "count": len(result),
        "rows": result
    }


def get_app_profile_data(username):
    if not username:
        return {}

    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
    try:
        row = conn_local.execute("""
            SELECT fullname, username, level,
                   COALESCE(app_display_name, ''),
                   COALESCE(department, ''),
                   COALESCE(profile_photo, '')
            FROM users
            WHERE username=?
            LIMIT 1
        """, (username,)).fetchone()
    except sqlite3.OperationalError:
        row = conn_local.execute("""
            SELECT fullname, username, level, '', '', ''
            FROM users
            WHERE username=?
            LIMIT 1
        """, (username,)).fetchone()
    finally:
        conn_local.close()

    if not row:
        return {}

    profile_photo = row[5] or ""
    return {
        "fullname": row[0] or row[1] or "",
        "username": row[1] or "",
        "level": row[2] or "",
        "display_name": row[3] or "",
        "department": row[4] or "",
        "profile_photo_url": (
            "/static/profile_uploads/" + profile_photo
            if profile_photo else ""
        )
    }


@app.route('/api/app_health', methods=['GET'])
def api_app_health():
    return jsonify({
        "success": True,
        "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })


@app.route('/api/app_profile', methods=['GET', 'POST'])
def api_app_profile():
    current_username = session.get("username") or session.get("pending_username")
    if not current_username:
        return jsonify({"success": False, "error": "Session absensi tidak ditemukan"}), 401

    if request.method == 'GET':
        profile = get_app_profile_data(current_username)
        return jsonify({"success": True, "profile": profile})

    data = request.form if request.form else (request.get_json(silent=True) or {})
    display_name = str(data.get("display_name", "")).strip()[:80]
    new_username = str(data.get("username", current_username)).strip().lower()
    password = str(data.get("password", "")).strip()
    department = str(data.get("department", "")).strip()[:100]

    if not re.fullmatch(r"[a-z0-9._-]{3,40}", new_username):
        return jsonify({
            "success": False,
            "error": "Username 3-40 karakter dan hanya boleh huruf, angka, titik, garis bawah, atau minus"
        }), 400
    if password and len(password) < 4:
        return jsonify({"success": False, "error": "Password minimal 4 karakter"}), 400

    photo_name = ""
    photo = request.files.get("photo")
    if photo and photo.filename:
        extension = os.path.splitext(secure_filename(photo.filename))[1].lower()
        if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
            return jsonify({"success": False, "error": "Foto harus JPG, PNG, atau WEBP"}), 400
        photo_name = "{}_{}{}".format(
            int(time.time() * 1000),
            secrets.token_hex(6),
            extension
        )
        photo.save(os.path.join(PROFILE_UPLOAD_FOLDER, photo_name))

    with DB_LOCK:
        conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
        try:
            existing = conn_local.execute("""
                SELECT id, fullname, username, level, COALESCE(profile_photo, '')
                FROM users
                WHERE username=?
                LIMIT 1
            """, (current_username,)).fetchone()
            if not existing:
                return jsonify({"success": False, "error": "Akun tidak ditemukan"}), 404
            if is_reserved_superman_account(existing[2], existing[3]) and new_username != current_username:
                return jsonify({"success": False, "error": "Username akun utama tidak dapat diubah"}), 400
            duplicate = conn_local.execute("""
                SELECT 1 FROM users WHERE username=? AND id<>? LIMIT 1
            """, (new_username, existing[0])).fetchone()
            if duplicate:
                return jsonify({"success": False, "error": "Username sudah digunakan"}), 409

            saved_photo = photo_name or existing[4] or ""
            if password:
                conn_local.execute("""
                    UPDATE users
                    SET username=?, password=?, app_display_name=?, department=?, profile_photo=?
                    WHERE id=?
                """, (
                    new_username, password, display_name, department, saved_photo, existing[0]
                ))
            else:
                conn_local.execute("""
                    UPDATE users
                    SET username=?, app_display_name=?, department=?, profile_photo=?
                    WHERE id=?
                """, (
                    new_username, display_name, department, saved_photo, existing[0]
                ))

            if new_username != current_username:
                for table_name in (
                    "attendance",
                    "attendance_leave",
                    "attendance_offline_sync",
                    "attendance_shifts"
                ):
                    conn_local.execute(
                        "UPDATE {} SET username=? WHERE username=?".format(table_name),
                        (new_username, current_username)
                    )
                conn_local.execute(
                    "UPDATE account_messages SET sender_username=? WHERE sender_username=?",
                    (new_username, current_username)
                )
                conn_local.execute(
                    "UPDATE account_messages SET recipient_username=? WHERE recipient_username=?",
                    (new_username, current_username)
                )
            conn_local.commit()
        finally:
            conn_local.close()

    if session.get("username") == current_username:
        session["username"] = new_username
    if session.get("pending_username") == current_username:
        session["pending_username"] = new_username

    return jsonify({
        "success": True,
        "message": "Profil aplikasi diperbarui",
        "profile": get_app_profile_data(new_username)
    })


@app.route('/api/app_messages', methods=['GET'])
def api_app_messages():
    recipient_username = session.get("username") or session.get("pending_username")
    if not recipient_username:
        return jsonify({"success": False, "error": "Session absensi tidak ditemukan"}), 401

    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
    try:
        rows = conn_local.execute("""
            SELECT id, sender_username, sender_fullname, message, image_path,
                   created_at, read_at,
                   COALESCE(message_type, 'message'),
                   COALESCE(action_url, ''),
                   COALESCE(action_label, ''),
                   COALESCE(app_version, '')
            FROM account_messages
            WHERE recipient_username=?
            ORDER BY id DESC
            LIMIT 1
        """, (recipient_username,)).fetchall()
    finally:
        conn_local.close()

    messages = []
    unread_count = 0
    for row in rows:
        unread = not bool(row[6])
        if unread:
            unread_count += 1
        messages.append({
            "id": row[0],
            "sender_username": row[1] or "",
            "sender_fullname": row[2] or row[1] or "Admin",
            "text": row[3] or "",
            "image_path": row[4] or "",
            "created_at": row[5] or "",
            "unread": unread,
            "type": row[7] or "message",
            "action_url": row[8] or "",
            "action_label": row[9] or "",
            "version": row[10] or ""
        })
    return jsonify({
        "success": True,
        "unread_count": unread_count,
        "messages": messages
    })


@app.route('/api/app_messages/ack', methods=['POST'])
def api_app_message_ack():
    recipient_username = session.get("username") or session.get("pending_username")
    if not recipient_username:
        return jsonify({"success": False, "error": "Session absensi tidak ditemukan"}), 401
    data = request.get_json(silent=True) or request.form
    message_id = data.get("id")
    if not message_id:
        return jsonify({"success": False, "error": "ID pesan wajib"}), 400
    with DB_LOCK:
        conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
        try:
            conn_local.execute("""
                UPDATE account_messages
                SET read_at=?
                WHERE id=? AND recipient_username=?
            """, (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                message_id,
                recipient_username
            ))
            conn_local.commit()
        finally:
            conn_local.close()
    return jsonify({"success": True})


@app.route('/api/app_messages/delete', methods=['POST'])
def api_app_message_delete():
    recipient_username = session.get("username") or session.get("pending_username")
    if not recipient_username:
        return jsonify({"success": False, "error": "Session absensi tidak ditemukan"}), 401
    data = request.get_json(silent=True) or request.form
    try:
        message_id = int(data.get("id", 0) or 0)
    except Exception:
        message_id = 0
    if message_id <= 0:
        return jsonify({"success": False, "error": "ID pesan wajib"}), 400
    with DB_LOCK:
        conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
        try:
            conn_local.execute("""
                DELETE FROM account_messages
                WHERE recipient_username=? AND id<=?
            """, (recipient_username, message_id))
            conn_local.commit()
        finally:
            conn_local.close()
    return jsonify({"success": True})


@app.route('/api/attendance_leave', methods=['GET', 'POST'])
def api_attendance_leave():
    username = session.get("username") or session.get("pending_username")
    fullname = session.get("fullname") or session.get("pending_fullname") or username
    if not username:
        return jsonify({"success": False, "error": "Session absensi tidak ditemukan"}), 401

    selected_month, _month_start = normalize_attendance_history_month(
        request.args.get("month", "")
    )

    if request.method == 'GET':
        response = jsonify({
            "success": True,
            "history": get_attendance_leave_history(username, selected_month)
        })
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return response

    data = request.get_json(silent=True) or request.form
    tanggal = str(data.get("tanggal", "")).strip()
    leave_type = str(data.get("type", "")).strip().upper()
    keterangan = str(data.get("keterangan", "")).strip()
    if leave_type not in {"IZIN", "SAKIT"}:
        return jsonify({"success": False, "error": "Jenis izin tidak valid"}), 400
    try:
        parsed_date = datetime.strptime(tanggal, "%Y-%m-%d")
    except Exception:
        return jsonify({"success": False, "error": "Tanggal harus berformat YYYY-MM-DD"}), 400
    if parsed_date.date() < datetime.now().date():
        return jsonify({"success": False, "error": "Tanggal izin tidak boleh sebelum hari ini"}), 400
    if not keterangan:
        return jsonify({"success": False, "error": "Keterangan wajib diisi"}), 400

    photo_name = ""
    photo = request.files.get("photo")
    if photo and photo.filename:
        extension = os.path.splitext(secure_filename(photo.filename))[1].lower()
        if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
            return jsonify({"success": False, "error": "Lampiran harus JPG, PNG, atau WEBP"}), 400
        upload_dir = os.path.join(BASE_DIR, "static", "uploads", "attendance_leave")
        os.makedirs(upload_dir, exist_ok=True)
        photo_name = "{}_{}_{}{}".format(
            datetime.now().strftime("%Y%m%d_%H%M%S"),
            secure_filename(username),
            secrets.token_hex(4),
            extension
        )
        photo.save(os.path.join(upload_dir, photo_name))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with DB_LOCK:
        conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
        try:
            conn_local.execute("""
                INSERT INTO attendance_leave (
                    username, fullname, tanggal, type, keterangan,
                    photo, status, created_at, updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?)
                ON CONFLICT(username, tanggal) DO UPDATE SET
                    fullname=excluded.fullname,
                    type=excluded.type,
                    keterangan=excluded.keterangan,
                    photo=CASE
                        WHEN excluded.photo<>'' THEN excluded.photo
                        ELSE attendance_leave.photo
                    END,
                    status=excluded.status,
                    updated_at=excluded.updated_at
            """, (
                username, fullname, tanggal, leave_type, keterangan,
                photo_name, "Tercatat", now, now
            ))
            conn_local.commit()
        finally:
            conn_local.close()

    return jsonify({
        "success": True,
        "message": "Pengajuan izin tersimpan",
        "history": get_attendance_leave_history(username, tanggal[:7])
    })


@app.route("/send_account_notification", methods=["POST"])
def send_account_notification():
    data = request.get_json(silent=True) or request.form
    notification_type = str(data.get("type", "message")).strip().lower()
    message = str(data.get("message", "")).strip()
    app_version = str(data.get("version", "")).strip()[:30]
    if notification_type not in {"update", "message"}:
        return jsonify({"success": False, "error": "Jenis notifikasi tidak valid"}), 400

    action_url = ""
    action_label = ""
    if notification_type == "update":
        app_version = app_version or get_attendance_app_version()
        if not app_version:
            return jsonify({"success": False, "error": "Versi update wajib diisi"}), 400
        if not re.fullmatch(r"[A-Za-z0-9._-]{1,30}", app_version):
            return jsonify({"success": False, "error": "Format versi update tidak valid"}), 400
        if message:
            if app_version.lower() not in message.lower():
                message = "Versi {}\n{}".format(app_version, message)
        else:
            message = "Update aplikasi absensi versi {} sudah tersedia. Ketuk untuk mengunduh APK terbaru.".format(app_version)
        action_url = ATTENDANCE_APK_DOWNLOAD_URL
        action_label = "Download Update"
    elif not message:
        return jsonify({"success": False, "error": "Isi pesan wajib diisi"}), 400

    sender_username = session.get("username", "")
    sender_fullname = session.get("fullname", sender_username) or "Admin"
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with DB_LOCK:
        cursor.execute("DELETE FROM account_messages")
        cursor.execute("""
            INSERT INTO account_messages (
                sender_username,
                sender_fullname,
                recipient_username,
                recipient_fullname,
                message,
                image_path,
                message_type,
                action_url,
                action_label,
                app_version,
                created_at,
                read_at
            )
            SELECT ?, ?, username, COALESCE(NULLIF(fullname, ''), username),
                   ?, '', ?, ?, ?, ?, ?, NULL
            FROM users
        """, (
            sender_username,
            sender_fullname,
            message,
            notification_type,
            action_url,
            action_label,
            app_version,
            created_at
        ))
        sent_count = cursor.rowcount
        conn.commit()

    return jsonify({
        "success": True,
        "sent_count": max(0, sent_count),
        "version": app_version,
        "message": "Notifikasi dikirim ke semua akun"
    })

@app.route('/force_change_credentials', methods=['POST'])
def force_change_credentials():

    try:

        user_id = session.get(
            "pending_credentials_user_id"
        )

        if not user_id:

            return jsonify({
                "success":False,
                "error":"Sesi ganti akun tidak ditemukan"
            }), 400

        data = request.json or {}

        username = str(
            data.get(
                "username",
                ""
            )
        ).strip().lower()

        password = str(
            data.get(
                "password",
                ""
            )
        ).strip()

        password_confirm = str(
            data.get(
                "password_confirm",
                ""
            )
        ).strip()

        if not username or not password:

            return jsonify({
                "success":False,
                "error":"Username dan password wajib diisi"
            }), 400

        if not re.fullmatch(
            r"[a-z0-9._-]+",
            username
        ):

            return jsonify({
                "success":False,
                "error":"Username hanya boleh huruf, angka, titik, underscore, atau strip"
            }), 400

        if password != password_confirm:

            return jsonify({
                "success":False,
                "error":"Konfirmasi password tidak sama"
            }), 400

        existing = cursor.execute("""

            SELECT id
            FROM users
            WHERE username=?
            AND id<>?

        """,(username,user_id)).fetchone()

        if existing:

            return jsonify({
                "success":False,
                "error":"Username sudah digunakan"
            }), 400

        target = cursor.execute("""

            SELECT must_change_credentials
            FROM users
            WHERE id=?

        """,(user_id,)).fetchone()

        if not target:

            session.clear()

            return jsonify({
                "success":False,
                "error":"Akun tidak ditemukan"
            }), 404

        if int(target[0] or 0) != 1:

            session.clear()

            return jsonify({
                "success":False,
                "error":"Data akun ini sudah pernah diubah"
            }), 400

        cursor.execute("""

            UPDATE users
            SET
                username=?,
                password=?,
                must_change_credentials=0
            WHERE id=?

        """,(
            username,
            password,
            user_id
        ))

        conn.commit()

        session.clear()

        return jsonify({
            "success":True,
            "redirect":"/"
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        }), 500

@app.route('/api/session')
def api_session():

    return jsonify({

        "logged_in":
            session.get("logged_in", False),

        "username":
            session.get("username", "")

    })

@app.route('/logout')
def logout():

    session.clear()

    return redirect('/')

@app.route('/')
def home():

    if (
        check_login() and
        is_attendance_only_session()
    ):

        attendance = (
            get_pending_clock_out_attendance(
                session.get("username")
            ) or
            get_today_attendance(
                session.get("username")
            )
        )

        if attendance:

            return render_attendance_success_page(
                attendance
            )

        session.clear()

    return render_template(
        'menu.html'
    )


def require_superman_page():
    return bool(check_login() and is_superman_session())


def get_server_cluster_snapshot():
    now_dt = datetime.now()
    conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=12))
    try:
        state = conn_local.execute("""
            SELECT active_node_id, lease_until, generation, updated_at
            FROM server_cluster_state
            WHERE cluster_id=?
            LIMIT 1
        """, (SERVER_CLUSTER_ID,)).fetchone()
        rows = conn_local.execute("""
            SELECT node_id, node_name, node_kind, priority, hostname, local_ip,
                   app_url, app_version, flask_ok, db_tunnel_ok,
                   cloudflared_ok, last_mode, last_seen, started_at
            FROM server_nodes
            ORDER BY priority DESC, node_name ASC
        """).fetchall()
    finally:
        conn_local.close()

    active_node_id = str(state[0] or "") if state else ""
    lease_until = str(state[1] or "") if state else ""
    lease_dt = parse_cluster_datetime(lease_until)
    lease_active = bool(lease_dt and lease_dt > now_dt)
    nodes = []
    for row in rows:
        last_seen = str(row[12] or "")
        last_seen_dt = parse_cluster_datetime(last_seen)
        age_seconds = int((now_dt - last_seen_dt).total_seconds()) if last_seen_dt else 999999
        online = bool(
            age_seconds <= SERVER_HEARTBEAT_TIMEOUT_SECONDS and
            int(row[8] or 0) == 1 and
            int(row[9] or 0) == 1
        )
        is_writer = bool(
            online and lease_active and str(row[0]) == active_node_id
        )
        nodes.append({
            "node_id": row[0] or "",
            "node_name": row[1] or row[0] or "Server",
            "node_kind": row[2] or "backup",
            "priority": int(row[3] or 0),
            "hostname": row[4] or "-",
            "local_ip": row[5] or "-",
            "app_url": row[6] or "",
            "app_version": row[7] or "-",
            "flask_ok": bool(row[8]),
            "db_tunnel_ok": bool(row[9]),
            "cloudflared_ok": bool(row[10]),
            "reported_mode": row[11] or "read_only",
            "last_seen": last_seen or "-",
            "started_at": row[13] or "-",
            "age_seconds": age_seconds,
            "online": online,
            "writer": is_writer,
            "access_mode": "write" if is_writer else "read_only",
            "state": "active" if is_writer else ("standby" if online else "offline")
        })

    return {
        "success": True,
        "cluster_id": SERVER_CLUSTER_ID,
        "active_node_id": active_node_id if lease_active else "",
        "lease_until": lease_until,
        "generation": int(state[2] or 0) if state else 0,
        "updated_at": str(state[3] or "") if state else "",
        "nodes": nodes
    }


@app.route('/server_management')
def server_management():
    if not require_superman_page():
        return redirect('/')
    return render_template('server_management.html')


@app.route('/api/server_nodes')
def api_server_nodes():
    if not require_superman_page():
        return jsonify({"success": False, "error": "Khusus akun SUPERMAN"}), 403
    try:
        return jsonify(get_server_cluster_snapshot())
    except Exception as error:
        return jsonify({"success": False, "error": str(error)}), 500


@app.route('/server_installer')
def server_installer_download():
    if not require_superman_page():
        return redirect('/')
    installer_path = os.path.join(
        BASE_DIR, "server_packages", "Tracer-Server-Backup.zip"
    )
    if not os.path.isfile(installer_path):
        return "Paket installer server belum tersedia.", 404
    return send_file(
        installer_path,
        as_attachment=True,
        download_name="Tracer-Server-Backup.zip",
        mimetype="application/zip",
        conditional=False,
        etag=False,
        max_age=0
    )
    


@app.route('/attendance_qr/<token>.png')
def attendance_qr_image(token):

    cleanup_attendance_qr_tokens()

    payload = ATTENDANCE_QR_TOKENS.get(
        token
    )

    if not payload:

        return redirect('/')

    qr_url = build_attendance_app_links(token)["launcher_url"]

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=9,
        border=2
    )
    qr.add_data(qr_url)
    qr.make(fit=True)

    image = qr.make_image(
        fill_color="black",
        back_color="white"
    )

    buffer = BytesIO()
    image.save(
        buffer,
        format="PNG"
    )
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='image/png'
    )


@app.route('/attendance_app/<token>')
def attendance_app_launcher(token):

    cleanup_attendance_qr_tokens()

    payload = ATTENDANCE_QR_TOKENS.get(
        token
    )

    if not payload:

        return redirect('/')

    links = build_attendance_app_links(
        token
    )
    username = payload.get(
        "username",
        ""
    )

    if (
        truthy_flag(payload.get("iphone_user", 0)) or
        get_user_iphone_user_status(username)
    ):

        return redirect(links["web_url"])

    return render_template(
        'attendance_app_launcher.html',
        app_url=links["app_url"],
        install_url=links["install_url"],
        web_url=links["web_url"],
        user_fullname=payload.get("fullname", ""),
        user_username=username,
        attendance_mode=payload.get("mode", "clock_in")
    )


@app.route('/attendance_apk')
def attendance_apk_download():

    apk_path = os.path.join(
        BASE_DIR,
        "static",
        "apk",
        "attendance.apk"
    )

    if not os.path.exists(apk_path):

        return (
            "File APK belum tersedia. Letakkan file attendance.apk di "
            "static/apk/attendance.apk.",
            404
        )

    app_version = get_attendance_app_version() or "latest"
    safe_version = re.sub(r"[^A-Za-z0-9._-]", "-", app_version)
    response = send_file(
        apk_path,
        as_attachment=True,
        download_name="Tracer-{}.apk".format(safe_version),
        mimetype="application/vnd.android.package-archive",
        conditional=False,
        etag=False,
        max_age=0
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["X-APK-Version"] = app_version
    return response


@app.route('/attendance_mobile_qr/<token>')
def attendance_mobile_qr(token):

    cleanup_attendance_qr_tokens()

    payload = ATTENDANCE_QR_TOKENS.get(
        token
    )

    if not payload:

        return redirect('/')

    username = payload.get(
        "username",
        ""
    )

    if not username:

        return redirect('/')

    mode = payload.get(
        "mode",
        "clock_in"
    )

    session.clear()

    if mode == "clock_out":

        session["logged_in"] = True
        session["username"] = username
        session["fullname"] = payload.get(
            "fullname",
            username
        )
        session["level"] = payload.get(
            "level",
            ""
        )
        session["permissions"] = payload.get(
            "permissions",
            []
        )
        session["attendance_bypass"] = int(
            truthy_flag(
                payload.get("attendance_bypass", 0)
            )
        )

        if is_attendance_only_session():

            mark_attendance_only_session_date()

        return redirect('/attendance_clock_out')

    if already_attendance(
        username
    ):

        session["logged_in"] = True
        session["username"] = username
        session["fullname"] = payload.get(
            "fullname",
            username
        )
        session["level"] = payload.get(
            "level",
            ""
        )
        session["permissions"] = payload.get(
            "permissions",
            []
        )
        session["attendance_bypass"] = int(
            truthy_flag(
                payload.get("attendance_bypass", 0)
            )
        )

        if is_attendance_only_session():

            mark_attendance_only_session_date()

        return redirect('/attendance_success')

    session["pending_username"] = username
    session["pending_fullname"] = payload.get(
        "fullname",
        username
    )
    session["pending_level"] = payload.get(
        "level",
        ""
    )
    session["pending_permissions"] = payload.get(
        "permissions",
        []
    )
    session["pending_attendance_bypass"] = int(
        truthy_flag(
            payload.get("attendance_bypass", 0)
        )
    )

    return redirect('/attendance_mobile')


@app.route('/attendance')
def attendance():

    if promote_pending_login_if_attendance_bypass():

        return redirect('/')

    if not session.get(
        "pending_username"
    ):

        return redirect('/')

    if not requires_attendance_for_level(
        session.get(
            "pending_level"
        ),
        session.get(
            "pending_attendance_bypass",
            0
        )
    ):

        clear_pending_login_session()

        return redirect('/')

    attendance_qr_token = create_attendance_qr_token("clock_in")
    attendance_qr_url = build_attendance_app_links(attendance_qr_token)["launcher_url"]

    return render_template(
        'attendance.html',
        user_fullname=session.get(
            "pending_fullname",
            session.get(
                "pending_username",
                ""
            )
        ),
        user_username=session.get(
            "pending_username",
            ""
        ),
        attendance_qr_url=attendance_qr_url,
        attendance_qr_image_url=url_for(
            'attendance_qr_image',
            token=attendance_qr_token
        ),
        attendance_mode="clock_in"
    )
    
@app.route('/attendance_mobile')
def attendance_mobile():

    if promote_pending_login_if_attendance_bypass():

        return redirect('/')

    if not session.get(
        "pending_username"
    ):

        return redirect('/')

    if not requires_attendance_for_level(
        session.get(
            "pending_level"
        ),
        session.get(
            "pending_attendance_bypass",
            0
        )
    ):

        clear_pending_login_session()

        return redirect('/')

    if (
        request.args.get("source") != "app" and
        not truthy_flag(session.get("pending_iphone_user", 0)) and
        not get_user_iphone_user_status(session.get("pending_username", ""))
    ):

        attendance_qr_token = create_attendance_qr_token("clock_in")
        return redirect(
            url_for(
                'attendance_app_launcher',
                token=attendance_qr_token
            )
        )

    if get_user_iphone_user_status(session.get("pending_username", "")):

        session["pending_iphone_user"] = 1

    return render_template(
        'attendance_mobile.html',
        user_fullname=session.get(
            "pending_fullname",
            session.get(
                "pending_username",
                ""
            )
        ),
        user_username=session.get(
            "pending_username",
            ""
        ),
        attendance_shift=get_attendance_shift_display(session.get("pending_username", "")),
        attendance_locations=get_attendance_locations(),
        disable_location_lock=get_user_disable_location_lock_status(
            session.get("pending_username", "")
        ),
        last_photo_url=get_last_attendance_photo_url(session.get("pending_username", ""))
    )

@app.route('/attendance_clock_out')
def attendance_clock_out_page():

    if not check_login():

        return redirect('/')

    attendance = get_pending_clock_out_attendance(
        session.get("username")
    )

    if not attendance:

        return redirect('/')

    if attendance.get("clock_out"):

        return redirect('/attendance_success')

    if (
        is_mobile_request() and
        request.args.get("source") != "app" and
        not truthy_flag(session.get("iphone_user", 0))
    ):

        attendance_qr_token = create_attendance_qr_token("clock_out")
        return redirect(
            url_for(
                'attendance_app_launcher',
                token=attendance_qr_token
            )
        )

    if is_mobile_request():

        return render_template(
            'attendance_mobile.html',
            user_fullname=session.get(
                "fullname",
                session.get(
                    "username",
                    ""
                )
            ),
            user_username=session.get(
                "username",
                ""
            ),
            attendance_mode="clock_out",
            attendance_shift=get_attendance_shift_display(session.get("username", "")),
            attendance_locations=get_attendance_locations(),
            disable_location_lock=get_user_disable_location_lock_status(
                session.get("username", "")
            ),
            last_photo_url=get_last_attendance_photo_url(session.get("username", ""))
        )

    attendance_qr_token = create_attendance_qr_token("clock_out")
    attendance_qr_url = build_attendance_app_links(attendance_qr_token)["launcher_url"]

    return render_template(
        'attendance.html',
        user_fullname=session.get(
            "fullname",
            session.get(
                "username",
                ""
            )
        ),
        user_username=session.get(
            "username",
            ""
        ),
        attendance_qr_url=attendance_qr_url,
        attendance_qr_image_url=url_for(
            'attendance_qr_image',
            token=attendance_qr_token
        ),
        attendance_mode="clock_out",
        attendance_shift=get_attendance_shift_display(session.get("username", "")),
        last_photo_url=get_last_attendance_photo_url(session.get("username", ""))
    )

@app.route('/attendance_success')
def attendance_success():

    if not check_login():

        return redirect('/')

    attendance = (
        get_attendance_by_id(
            session.get("last_attendance_id")
        ) or
        get_pending_clock_out_attendance(
            session.get("username")
        ) or
        get_today_attendance(
            session.get("username")
        )
    )

    if not attendance:

        return redirect('/')

    if is_attendance_only_session():

        return render_attendance_success_page(
            attendance
        )

    return render_template(
        'attendance_success.html',
        attendance=attendance,
        attendance_only=False,
        redirect_url="/",
        redirect_text="masuk menu",
        auto_redirect=True,
        redirect_seconds=10
    )

@app.route('/attendance_history')
def attendance_history():

    username = session.get("username") or session.get("pending_username")

    if not username:

        return redirect('/')

    fullname = (
        session.get("fullname") or
        session.get("pending_fullname") or
        username
    )

    selected_month = request.args.get(
        "month",
        ""
    )

    selected_month, _month_start = normalize_attendance_history_month(
        selected_month
    )

    return render_template(
        'attendance_history.html',
        history=get_attendance_history(
            username,
            month=selected_month
        ),
        username=username,
        fullname=fullname,
        selected_month=selected_month
    )

@app.route('/api/attendance_history', methods=['GET'])
def api_attendance_history():

    username = (
        session.get("username") or
        session.get("pending_username")
    )

    if not username:

        return jsonify({
            "success": False,
            "error": "Session absensi tidak ditemukan"
        }), 401

    fullname = (
        session.get("fullname") or
        session.get("pending_fullname") or
        username
    )

    selected_month = request.args.get(
        "month",
        ""
    )
    selected_month, _month_start = normalize_attendance_history_month(
        selected_month
    )

    history = get_attendance_history(
        username,
        month=selected_month
    )
    response = jsonify({
        "success": True,
        "username": username,
        "fullname": fullname,
        "history": history,
        "leave_history": get_attendance_leave_history(username, selected_month)
    })
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response

@app.route('/attendance_history_export')
def attendance_history_export():

    if not check_login():

        return redirect('/')

    username = session.get(
        "username",
        "user"
    )

    selected_month = request.args.get(
        "month",
        ""
    )

    selected_month, _month_start = normalize_attendance_history_month(
        selected_month
    )

    return send_file(
        export_attendance_history_csv(
            username,
            month=selected_month
        ),
        as_attachment=True,
        download_name=(
            "riwayat_absensi_" +
            username +
            "_" +
            selected_month +
            ".csv"
        ),
        mimetype="text/csv"
    )
    
@app.route('/api/check-attendance')
def check_attendance():

    username = session.get("username")

    level = session.get(
        "level"
    )

    if not username:

        username = session.get("pending_username")

        level = session.get(
            "pending_level"
        )
        attendance_bypass = session.get(
            "pending_attendance_bypass",
            0
        )

    else:

        attendance_bypass = session.get(
            "attendance_bypass",
            0
        )

    if not username:

        return jsonify({

            "success":False

        })

    if not requires_attendance_for_level(
        level,
        attendance_bypass
    ):

        return jsonify({
            "success":True,
            "attended":True,
            "attendance_only":
                is_attendance_only_session()
        })

    check_mode = request.args.get(
        "mode",
        "clock_in"
    )

    if check_mode == "clock_out":
        attendance = (
            get_pending_clock_out_attendance(username) or
            get_latest_attendance(username)
        )
        attended = bool(
            attendance and
            attendance.get("clock_out")
        )
    else:
        attendance = get_today_attendance(
            username
        )
        attended = attendance is not None

    if attended and not session.get("logged_in"):

        attendance_only = is_pending_attendance_only_session()

        session["logged_in"] = True

        session["username"] = session.get(
            "pending_username"
        )

        session["fullname"] = session.get(
            "pending_fullname"
        )

        session["level"] = session.get(
            "pending_level"
        )

        session["permissions"] = session.get(
            "pending_permissions",
            []
        )

        session["attendance_bypass"] = int(
            truthy_flag(
                session.get(
                    "pending_attendance_bypass",
                    0
                )
            )
        )

        
        if attendance_only:

            mark_attendance_only_session_date()

        session.pop(
            "pending_username",
            None
        )

        session.pop(
            "pending_fullname",
            None
        )

        session.pop(
            "pending_level",
            None
        )

        session.pop(
            "pending_permissions",
            None
        )

        session.pop(
            "pending_iphone_user",
            None
        )

        session.pop(
            "pending_attendance_bypass",
            None
        )

    attendance_only = is_attendance_only_session()

    return jsonify({

        "success":True,

        "attended":
            attended,

        "attendance_only":
            attendance_only,

        "attendance":
                attendance,

            "redirect":(
                "/attendance_success"
                if attendance_only
                else "/attendance_dashboard"
            )

        })
    
def get_request_client_ip():
    """Return the original device IP when Flask is behind Cloudflare/Nginx."""
    for header_name in ("CF-Connecting-IP", "True-Client-IP"):
        value = str(request.headers.get(header_name, "")).strip()
        if value:
            return value[:80]
    forwarded = str(request.headers.get("X-Forwarded-For", ""))
    if forwarded:
        value = forwarded.split(",", 1)[0].strip()
        if value:
            return value[:80]
    return str(request.remote_addr or "").strip()[:80]


@app.route('/api/attendance', methods=['POST'])
def submit_attendance():

    if not session.get("pending_username"):

        return jsonify({
            "success":False,
            "message":"Session expired"
        })

    if not requires_attendance_for_level(
        session.get(
            "pending_level"
        ),
        session.get(
            "pending_attendance_bypass",
            0
        )
    ):

        clear_pending_login_session()

        return jsonify({
            "success":False,
            "message":"Akun ini tidak perlu absen"
        }), 403

    username = session.get(
        "pending_username"
    )

    fullname = session.get(
        "pending_fullname"
    )

    if not username:

        return jsonify({

            "success":False,

            "message":"Session expired"

        })

    if already_attendance(username):

        return jsonify({

            "success":False,
            "message":"Sudah absen hari ini"

        })

    try:

        photo = request.files.get("photo")

        latitude = request.form.get(
            "latitude",
            ""
        )

        longitude = request.form.get(
            "longitude",
            ""
        )

        address = request.form.get(
            "address",
            ""
        )

        

        shift_id = request.form.get(
            "shift_id",
            ""
        )
        try:
            face_score = max(0, min(100, int(request.form.get("face_score", "0") or 0)))
        except Exception:
            face_score = 0
        device_info = str(request.form.get("device_info", "")).strip()[:160]
        app_version = str(request.form.get("app_version", "")).strip()[:40]
        client_ip = get_request_client_ip()
        folder = os.path.join(
            BASE_DIR,
            "static",
            "uploads",
            "attendance"
        )

        os.makedirs(
            folder,
            exist_ok=True
        )

        filename = (
            datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            + "_"
            + username
            + ".jpg"
        )

        save_path = os.path.join(
            folder,
            filename
        )

        photo.save(save_path)

        tanggal = datetime.now().strftime(
            "%Y-%m-%d"
        )

        jam = datetime.now().strftime(
            "%H:%M:%S"
        )

        cursor.execute("""

        INSERT INTO attendance (

            username,
            fullname,
            tanggal,
            jam,
            photo,
            latitude,
            longitude,
            address,
            face_score,
            device_info,
            ip_address,
            app_version,
            created_at,
            shift_id

        )

        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)

        """,(

            username,
            fullname,
            tanggal,
            jam,
            filename,
            latitude,
            longitude,
            address,
            face_score,
            device_info,
            client_ip,
            app_version,

            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            shift_id

        ))

        attendance_id = cursor.lastrowid

        conn.commit()

        attendance_only = is_pending_attendance_only_session()

        attendance = {
            "id": attendance_id,
            "username": username,
            "fullname": fullname,
            "tanggal": tanggal,
            "jam": jam,
            "photo": filename,
            "photo_url": "/static/uploads/attendance/" + filename,
            "latitude": latitude,
            "longitude": longitude,
            "address": address,
            "face_score": face_score,
            "device_info": device_info,
            "ip_address": client_ip,
            "app_version": app_version,
            "status": "Berhasil"
        }
        attendance = attach_attendance_shift_meta(attendance)
        
        session["logged_in"] = True

        session["username"] = session[
            "pending_username"
        ]

        session["fullname"] = session[
            "pending_fullname"
        ]

        session["level"] = session[
            "pending_level"
        ]
        
        session["permissions"] = session.get(
            "pending_permissions",
            []
        )

        
        
        session["last_attendance_id"] = attendance_id
        if attendance_only:

            mark_attendance_only_session_date()

        session.pop(
            "pending_username",
            None
        )

        session.pop(
            "pending_fullname",
            None
        )

        session.pop(
            "pending_level",
            None
        )

        session.pop(
            "pending_permissions",
            None
        )

        session.pop(
            "pending_iphone_user",
            None
        )
       

        return jsonify({

            "success":True,

            "username":session["username"],

            "attendance_only":
                attendance_only,

            "attendance":
                attendance,

            "redirect":(
                "/attendance_success"
                if attendance_only
                else "/attendance_dashboard"
            )

        })

    except Exception as e:

        return jsonify({

            "success":False,
            "message":str(e)

        })

@app.route('/api/attendance_sync', methods=['POST'])
def attendance_sync():

    event_id = str(request.form.get("event_id", "")).strip()
    username = str(request.form.get("username", "")).strip()
    fullname = str(request.form.get("fullname", "")).strip()
    mode = str(request.form.get("mode", "clock_in")).strip()
    shift_id = str(request.form.get("shift_id", "")).strip()
    captured_at = str(request.form.get("captured_at", "")).strip()
    latitude = str(request.form.get("latitude", "")).strip()
    longitude = str(request.form.get("longitude", "")).strip()
    address = str(request.form.get("address", "")).strip()
    try:
        face_score = max(0, min(100, int(request.form.get("face_score", "0") or 0)))
    except Exception:
        face_score = 0
    device_info = str(request.form.get("device_info", "")).strip()[:160]
    app_version = str(request.form.get("app_version", "")).strip()[:40]
    client_ip = get_request_client_ip()

    if not event_id:
        return jsonify({
            "success": False,
            "message": "event_id wajib diisi"
        }), 400

    if not username:
        return jsonify({
            "success": False,
            "message": "username wajib diisi"
        }), 400

    if mode not in ("clock_in", "clock_out"):
        return jsonify({
            "success": False,
            "message": "mode tidak valid"
        }), 400

    try:
        captured_dt = datetime.strptime(
            captured_at,
            "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        captured_dt = datetime.now()

    tanggal = captured_dt.strftime("%Y-%m-%d")
    jam = captured_dt.strftime("%H:%M:%S")
    safe_event_id = secure_filename(event_id) or datetime.now().strftime("%Y%m%d%H%M%S")
    already_applied = False

    with DB_LOCK:

        synced = cursor.execute("""

            SELECT attendance_id
            FROM attendance_offline_sync
            WHERE event_id=?
            LIMIT 1

        """, (
            event_id,
        )).fetchone()

        if synced:
            attendance = get_attendance_by_id(synced[0])
            return jsonify({
                "success": True,
                "already_synced": True,
                "attendance": attendance
            })

        user = cursor.execute("""

            SELECT username, fullname
            FROM users
            WHERE LOWER(username)=LOWER(?)
            LIMIT 1

        """, (
            username,
        )).fetchone()

        if not user and fullname:
            fullname_matches = cursor.execute("""

                SELECT username, fullname
                FROM users
                WHERE LOWER(fullname)=LOWER(?)
                ORDER BY id DESC
                LIMIT 2

            """, (
                fullname,
            )).fetchall()
            if len(fullname_matches) == 1:
                user = fullname_matches[0]

        if not user:
            app.logger.warning(
                "attendance_sync user_not_found event=%s username=%s fullname=%s mode=%s",
                event_id,
                username,
                fullname,
                mode
            )
            return jsonify({
                "success": False,
                "message": "User tidak ditemukan"
            }), 404

        username = user[0] or username
        if not fullname:
            fullname = user[1] or username

    photo = request.files.get("photo")
    filename = ""
    if photo:
        folder = os.path.join(
            BASE_DIR,
            "static",
            "uploads",
            "attendance"
        )
        os.makedirs(folder, exist_ok=True)
        filename = (
            captured_dt.strftime("%Y%m%d_%H%M%S")
            + "_offline_"
            + ("clockout_" if mode == "clock_out" else "")
            + secure_filename(username)
            + "_"
            + safe_event_id
            + ".jpg"
        )
        photo.save(os.path.join(folder, filename))

    with DB_LOCK:

        if mode == "clock_in":

            row = cursor.execute("""

                SELECT id, jam, created_at, photo
                FROM attendance
                WHERE username=?
                AND tanggal=?
                ORDER BY id DESC
                LIMIT 1

            """, (
                username,
                tanggal
            )).fetchone()

            if row:
                attendance_id = row[0]
                captured_text = captured_dt.strftime("%Y-%m-%d %H:%M:%S")
                existing_text = str(row[2] or "").strip()
                if not existing_text and row[1]:
                    existing_text = tanggal + " " + str(row[1])
                if not existing_text or captured_text < existing_text:
                    cursor.execute("""

                        UPDATE attendance
                        SET jam=?,
                            photo=CASE WHEN ?<>'' THEN ? ELSE photo END,
                            latitude=?,
                            longitude=?,
                            address=?,
                            face_score=?,
                            device_info=?,
                            ip_address=?,
                            app_version=COALESCE(NULLIF(?, ''), app_version),
                            created_at=?,
                            shift_id=COALESCE(NULLIF(?, ''), shift_id)
                        WHERE id=?

                    """, (
                        jam,
                        filename,
                        filename,
                        latitude,
                        longitude,
                        address,
                        face_score,
                        device_info,
                        client_ip,
                        app_version,
                        captured_text,
                        shift_id,
                        attendance_id
                    ))
                elif filename and not str(row[3] or "").strip():
                    cursor.execute(
                        "UPDATE attendance SET photo=? WHERE id=?",
                        (filename, attendance_id)
                    )
                elif filename:
                    unused_photo = os.path.join(
                        BASE_DIR,
                        "static",
                        "uploads",
                        "attendance",
                        filename
                    )
                    try:
                        os.remove(unused_photo)
                    except OSError:
                        pass
            else:
                cursor.execute("""

                    INSERT INTO attendance (

                        username,
                        fullname,
                        tanggal,
                        jam,
                        photo,
                        latitude,
                        longitude,
                        address,
                        face_score,
                        device_info,
                        ip_address,
                        app_version,
                        created_at,
                        shift_id

                    )

                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)

                """, (
                    username,
                    fullname,
                    tanggal,
                    jam,
                    filename,
                    latitude,
                    longitude,
                    address,
                    face_score,
                    device_info,
                    client_ip,
                    app_version,
                    captured_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    shift_id
                ))

                attendance_id = cursor.lastrowid

        else:

            row = cursor.execute("""

                SELECT id, clock_out
                FROM attendance
                WHERE username=?
                AND (
                    clock_out IS NULL
                    OR TRIM(clock_out)=''
                )
                ORDER BY tanggal DESC, id DESC
                LIMIT 1

            """, (
                username,
            )).fetchone()

            if not row:
                completed = cursor.execute("""

                    SELECT id
                    FROM attendance
                    WHERE username=?
                    AND tanggal=?
                    AND clock_out IS NOT NULL
                    AND TRIM(clock_out)<>''
                    ORDER BY id DESC
                    LIMIT 1

                """, (
                    username,
                    tanggal
                )).fetchone()
                if completed:
                    attendance_id = completed[0]
                    already_applied = True
                    if filename:
                        unused_photo = os.path.join(
                            BASE_DIR,
                            "static",
                            "uploads",
                            "attendance",
                            filename
                        )
                        try:
                            os.remove(unused_photo)
                        except OSError:
                            pass
                else:
                    app.logger.warning(
                        "attendance_sync clock_in_missing event=%s username=%s captured_at=%s",
                        event_id,
                        username,
                        captured_at
                    )
                    return jsonify({
                        "success": False,
                        "message": "Belum ada clock in yang belum clock out"
                    }), 409
            else:
                attendance_id = row[0]
                cursor.execute("""

                    UPDATE attendance
                    SET clock_out=?,
                        clock_out_at=?,
                        clock_out_photo=?,
                        clock_out_latitude=?,
                        clock_out_longitude=?,
                        clock_out_address=?,
                        clock_out_face_score=?,
                        clock_out_device_info=?,
                        clock_out_ip_address=?,
                        app_version=COALESCE(NULLIF(?, ''), app_version),
                        shift_id=COALESCE(NULLIF(shift_id, ''), NULLIF(?, ''), '1')
                    WHERE id=?

                """, (
                    jam,
                    captured_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    filename,
                    latitude,
                    longitude,
                    address,
                    face_score,
                    device_info,
                    client_ip,
                    app_version,
                    shift_id,
                    attendance_id
                ))

        cursor.execute("""

            INSERT OR REPLACE INTO attendance_offline_sync (
                event_id,
                username,
                mode,
                attendance_id,
                synced_at
            )
            VALUES (?,?,?,?,?)

        """, (
            event_id,
            username,
            mode,
            attendance_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

        conn.commit()

    attendance = get_attendance_by_id(attendance_id)

    return jsonify({
        "success": True,
        "already_synced": False,
        "already_applied": already_applied,
        "attendance": attendance
    })


@app.route('/api/attendance_clock_out', methods=['POST'])
def attendance_clock_out():

    if not check_login():

        return jsonify({
            "success":False,
            "message":"Session expired"
        }), 401

    username = session.get("username")

    if not username:

        return jsonify({
            "success":False,
            "message":"Session expired"
        }), 401

    today = datetime.now().strftime("%Y-%m-%d")
    clock_out = datetime.now().strftime("%H:%M:%S")
    clock_out_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    latitude = request.form.get("latitude", "")
    longitude = request.form.get("longitude", "")
    address = request.form.get("address", "")
    
    shift_id = request.form.get("shift_id", "")
    try:
        face_score = max(0, min(100, int(request.form.get("face_score", "0") or 0)))
    except Exception:
        face_score = 0
    device_info = str(request.form.get("device_info", "")).strip()[:160]
    app_version = str(request.form.get("app_version", "")).strip()[:40]
    client_ip = get_request_client_ip()
    filename = ""

    try:
        photo = request.files.get("photo")
        if photo:
            folder = os.path.join(
                BASE_DIR,
                "static",
                "uploads",
                "attendance"
            )
            os.makedirs(folder, exist_ok=True)
            filename = (
                datetime.now().strftime("%Y%m%d_%H%M%S")
                + "_clockout_"
                + username
                + ".jpg"
            )
            photo.save(os.path.join(folder, filename))

        with DB_LOCK:

            row = cursor.execute("""

                SELECT id, clock_out
                FROM attendance
                WHERE username=?
                AND (
                    clock_out IS NULL
                    OR TRIM(clock_out)=''
                )
                ORDER BY tanggal DESC, id DESC
                LIMIT 1

            """,(
                username,
            )).fetchone()

            if not row:

                return jsonify({
                    "success":False,
                    "message":"Belum ada clock in yang belum clock out"
                }), 404

            if row[1]:

                attendance = get_attendance_by_id(row[0])

                return jsonify({
                    "success":True,
                    "already_clock_out":True,
                    "attendance":attendance,
                    "attendance_only":is_attendance_only_session()
                })

            cursor.execute("""

                UPDATE attendance
                SET clock_out=?,
                    clock_out_at=?,
                    clock_out_photo=?,
                    clock_out_latitude=?,
                    clock_out_longitude=?,
                    clock_out_address=?,
                    clock_out_face_score=?,
                    clock_out_device_info=?,
                    clock_out_ip_address=?,
                    app_version=COALESCE(NULLIF(?, ''), app_version),
                    shift_id=COALESCE(NULLIF(shift_id, ''), NULLIF(?, ''), '1')
                WHERE id=?

            """,(
                clock_out,
                clock_out_at,
                filename,
                latitude,
                longitude,
                address,
                face_score,
                device_info,
                client_ip,
                app_version,
                shift_id,
                row[0]
            ))

            conn.commit()

            session["last_attendance_id"] = row[0]

    except sqlite3.OperationalError:

        with DB_LOCK:
            row = cursor.execute("""

                SELECT id
                FROM attendance
                WHERE username=?
                AND (
                    clock_out IS NULL
                    OR TRIM(clock_out)=''
                )
                ORDER BY tanggal DESC, id DESC
                LIMIT 1

            """,(
                username,
            )).fetchone()

            if not row:
                return jsonify({
                    "success":False,
                    "message":"Belum ada clock in yang belum clock out"
                }), 404

            cursor.execute("""

                UPDATE attendance
                SET clock_out=?,
                    clock_out_at=?
                WHERE id=?

            """,(
                clock_out,
                clock_out_at,
                row[0]
            ))
            conn.commit()
            session["last_attendance_id"] = row[0]

    except Exception as e:

        return jsonify({
            "success":False,
            "message":str(e)
        }), 500

    attendance = get_attendance_by_id(
        session.get("last_attendance_id")
    )

    return jsonify({
        "success":True,
        "already_clock_out":False,
        "attendance":attendance,
        "attendance_only":is_attendance_only_session()
    })

@app.route('/scan')
def scan():

    if (
        not check_login() and
        not promote_pending_login_if_attendance_bypass()
    ):
        return redirect('/')

    if not has_permission(
        "SCAN_SORTIR"
    ):
        return redirect('/')

    return render_template(
        'index.html'
    )
    
@app.route('/dashboard')
def dashboard():

    if not check_login():
        return redirect('/')

    if not has_permission(
        "DASHBOARD_RETUR"
    ):
        return redirect('/')

    user_agent = (
        request.headers.get("User-Agent") or ""
    ).lower()
    mobile_markers = (
        "android",
        "iphone",
        "ipod",
        "windows phone",
        "mobile"
    )
    if (
        request.args.get("desktop") != "1" and
        any(marker in user_agent for marker in mobile_markers)
    ):
        return redirect('/summary')

    return render_template(
        'dashboard.html'
    )
    
@app.route('/summary')
def summary():

    if not check_login():
        return redirect('/')

    if not has_permission(
        "DASHBOARD_RETUR"
    ):
        return redirect('/')

    return render_template(
        'summary.html'
    )
    



@app.route('/summary_data')
def summary_data():

    if not check_login():
        return jsonify({"ok": False, "error": "LOGIN_REQUIRED"}), 401

    if not has_permission(
        "DASHBOARD_RETUR"
    ):
        return jsonify({"ok": False, "error": "ACCESS_DENIED"}), 403

    mode = str(request.args.get("mode") or "harian").strip().lower()
    selected = str(request.args.get("selected") or "").strip()
    if mode not in ("harian", "bulanan", "akumulasi"):
        mode = "harian"

    def non_empty(column):
        return f"COALESCE(NULLIF(NULLIF(TRIM({column}), ''), '-'), '') != ''"

    def valid_waktu():
        return "(waktu >= '2000-01-01' AND waktu < '2100-01-01')"

    def next_day(value):
        return (
            datetime.strptime(value, "%Y-%m-%d") +
            timedelta(days=1)
        ).strftime("%Y-%m-%d")

    def next_month(value):
        base = datetime.strptime(value[:7], "%Y-%m")
        return (
            base.replace(day=28) +
            timedelta(days=4)
        ).replace(day=1).strftime("%Y-%m-%d")

    latest_row = conn.execute(
        f"""
        SELECT
            MAX(substr(waktu,1,10)) AS latest_date,
            MAX(substr(waktu,1,7)) AS latest_month
        FROM scans
        WHERE {valid_waktu()}
        """
    ).fetchone()
    latest_date = (latest_row[0] if latest_row else "") or datetime.now().strftime("%Y-%m-%d")
    latest_month = (latest_row[1] if latest_row else "") or latest_date[:7]

    if mode == "harian":
        active_period = selected[:10] if selected[:10].count("-") == 2 else latest_date
        active_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
        active_params = [active_period, next_day(active_period)]
        active_dt = datetime.strptime(active_period, "%Y-%m-%d")
        series_start_dt = active_dt.replace(day=1)
        series_end_dt = max(active_dt, series_start_dt + timedelta(days=6))
        series_start = series_start_dt.strftime("%Y-%m-%d")
        series_end = next_day(series_end_dt.strftime("%Y-%m-%d"))
        series_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
        series_params = [series_start, series_end]
        bucket_expr = "substr(waktu,1,10)"
        label = active_period
        kode_label = "Total Karung Harian"
        compare_label = "Dibanding tanggal sebelumnya"
        prev_row = conn.execute(
            f"""
            SELECT MAX(substr(waktu,1,10))
            FROM scans
            WHERE {valid_waktu()} AND substr(waktu,1,10) < ?
            """,
            [active_period]
        ).fetchone()
        previous_period = prev_row[0] if prev_row and prev_row[0] else ""

    elif mode == "bulanan":
        active_period = selected[:7] if selected[:7].count("-") == 1 else latest_month
        active_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
        active_params = [f"{active_period}-01", next_month(active_period)]
        month_base = datetime.strptime(active_period[:7], "%Y-%m")
        series_start = month_base.replace(month=1, day=1).strftime("%Y-%m-%d")
        series_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
        series_params = [series_start, next_month(active_period)]
        bucket_expr = "substr(waktu,1,7)"
        label = active_period
        kode_label = "Total Karung Bulanan"
        compare_label = "Dibanding bulan sebelumnya"
        prev_row = conn.execute(
            f"""
            SELECT MAX(substr(waktu,1,7))
            FROM scans
            WHERE {valid_waktu()} AND substr(waktu,1,7) < ?
            """,
            [active_period]
        ).fetchone()
        previous_period = prev_row[0] if prev_row and prev_row[0] else ""

    else:
        active_period = latest_month
        active_where = f"{valid_waktu()} AND waktu < ?"
        active_params = [next_month(active_period)]
        month_base = datetime.strptime(active_period[:7], "%Y-%m")
        series_start = month_base.replace(month=1, day=1).strftime("%Y-%m-%d")
        series_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
        series_params = [series_start, next_month(active_period)]
        bucket_expr = "substr(waktu,1,7)"
        label = f"Akumulasi sampai {active_period}"
        kode_label = "Total Karung Akumulasi"
        compare_label = "Total sampai bulan terbaru"
        previous_period = ""

    def aggregate(where_sql, params):
        return conn.execute(
            f"""
            SELECT
                COUNT(DISTINCT CASE
                    WHEN {non_empty('scan_pack_time')} AND {non_empty('scan_pack_code')}
                    THEN scan_pack_code END) AS total_kode,
                COUNT(DISTINCT CASE
                    WHEN {non_empty('scan_pack_time')} THEN resi END) AS total_resi,
                COUNT(DISTINCT CASE
                    WHEN NOT ({non_empty('scan_pack_time')})
                     AND NOT ({non_empty('scan_pack_code')})
                     AND NOT ({non_empty('kode')})
                     AND UPPER(TRIM(COALESCE(status, ''))) NOT IN ('INCOMING', 'ABNORMAL')
                    THEN resi END) AS resi_tanpa_kode,
                COUNT(DISTINCT CASE WHEN {non_empty('seller')} THEN seller END) AS total_seller,
                COUNT(DISTINCT CASE WHEN {non_empty('scan_pack_time')} THEN resi END) AS scan_pack,
                COUNT(DISTINCT CASE WHEN {non_empty('scan_delivery_time')} THEN resi END) AS scan_delivery,
                COUNT(DISTINCT CASE WHEN {non_empty('received_at')} THEN resi END) AS terima_seller,
                COUNT(DISTINCT CASE WHEN badges LIKE '%HIGH VALUE%' THEN resi END) AS selisih_hv,
                COUNT(DISTINCT CASE WHEN badges LIKE '%COMPLAINT%' THEN resi END) AS total_complaint,
                COUNT(DISTINCT CASE WHEN badges LIKE '%AUTOCLAIM%' THEN resi END) AS total_autoclaim,
                COUNT(DISTINCT COALESCE(NULLIF(TRIM(scan_tracer_station), ''), NULLIF(TRIM(collect_staff), ''))) AS total_implant
            FROM scans
            WHERE {where_sql}
            """,
            params
        ).fetchone()

    agg = aggregate(active_where, active_params)
    total_kode = int(agg[0] or 0)
    total_resi = int(agg[1] or 0)
    resi_tanpa_kode = int(agg[2] or 0)
    total_seller = int(agg[3] or 0)
    scan_pack_total = int(agg[4] or 0)
    scan_delivery_total = int(agg[5] or 0)
    terima_seller_total = int(agg[6] or 0)
    selisih_hv = int(agg[7] or 0)
    total_complaint = int(agg[8] or 0)
    total_autoclaim = int(agg[9] or 0)
    total_implant = int(agg[10] or 0)
    resi_abnormal_total = int(conn.execute(f"""
        SELECT COUNT(DISTINCT resi)
        FROM scans
        WHERE {active_where}
          AND UPPER(TRIM(COALESCE(status, ''))) IN ('INCOMING', 'ABNORMAL')
        """, active_params).fetchone()[0] or 0)
    status_terima_total = int(conn.execute(f"""
        SELECT COUNT(DISTINCT resi)
        FROM scans
        WHERE {active_where}
          AND {non_empty('received_at')}
        """, active_params).fetchone()[0] or 0)
    status_proses_total = int(conn.execute(f"""
        SELECT COUNT(DISTINCT resi)
        FROM scans
        WHERE {active_where}
          AND NOT ({non_empty('received_at')})
        """, active_params).fetchone()[0] or 0)
    proses_total = max(status_proses_total, 0)
    progress_base = status_proses_total + status_terima_total
    progress = round((status_terima_total / progress_base) * 100, 2) if progress_base else 0.0

    delivery_warehouse_summary = []

    previous_kode = 0
    if previous_period:
        if mode == "harian":
            prev_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
            prev_params = [previous_period, next_day(previous_period)]
        else:
            prev_where = f"{valid_waktu()} AND waktu >= ? AND waktu < ?"
            prev_params = [f"{previous_period}-01", next_month(previous_period)]
        previous_kode = int(aggregate(prev_where, prev_params)[0] or 0)
    kode_delta = round(((total_kode - previous_kode) / previous_kode) * 100, 1) if previous_kode else 0

    series_rows = conn.execute(
        f"""
        SELECT
            {bucket_expr} AS bucket,
            COUNT(DISTINCT CASE WHEN {non_empty('scan_pack_time')} THEN resi END) AS scan_pack,
            COUNT(DISTINCT CASE WHEN {non_empty('scan_delivery_time')} THEN resi END) AS scan_delivery,
            COUNT(DISTINCT CASE WHEN {non_empty('received_at')} THEN resi END) AS terima_seller,
            COUNT(DISTINCT CASE
                WHEN {non_empty('scan_pack_time')} AND {non_empty('scan_pack_code')}
                THEN scan_pack_code END) AS total_kode
        FROM scans
        WHERE {series_where}
        GROUP BY bucket
        HAVING bucket IS NOT NULL AND bucket != ''
        ORDER BY bucket ASC
        """,
        series_params
    ).fetchall()

    labels = []
    scan_pack = []
    scan_delivery = []
    terima_seller = []
    mini = []
    running_pack = running_delivery = running_seller = running_kode = 0
    month_names = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
        "05": "Mei", "06": "Jun", "07": "Jul", "08": "Agu",
        "09": "Sep", "10": "Okt", "11": "Nov", "12": "Des"
    }
    series_map = {str(row[0] or ""): row for row in series_rows}
    buckets = []
    if mode == "harian":
        cur = datetime.strptime(series_start, "%Y-%m-%d")
        end = datetime.strptime(series_params[1], "%Y-%m-%d")
        while cur < end:
            buckets.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
    else:
        cur = datetime.strptime(series_start[:7], "%Y-%m")
        end = datetime.strptime(active_period[:7], "%Y-%m")
        while cur <= end:
            buckets.append(cur.strftime("%Y-%m"))
            year = cur.year + (1 if cur.month == 12 else 0)
            month = 1 if cur.month == 12 else cur.month + 1
            cur = cur.replace(year=year, month=month)

    for bucket in buckets:
        row = series_map.get(bucket)
        if mode == "harian":
            try:
                labels.append(datetime.strptime(bucket[:10], "%Y-%m-%d").strftime("%d %b"))
            except Exception:
                labels.append(bucket)
        else:
            labels.append(month_names.get(bucket[-2:], bucket))

        p = int(row[1] or 0) if row else 0
        d = int(row[2] or 0) if row else 0
        s = int(row[3] or 0) if row else 0
        k = int(row[4] or 0) if row else 0
        if mode == "akumulasi":
            running_pack += p
            running_delivery += d
            running_seller += s
            running_kode += k
            scan_pack.append(running_pack)
            scan_delivery.append(running_delivery)
            terima_seller.append(running_seller)
            mini.append(running_kode)
        else:
            scan_pack.append(p)
            scan_delivery.append(d)
            terima_seller.append(s)
            mini.append(k)

    all_seller_rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(TRIM(seller), ''), 'Tanpa Seller') AS name,
               COUNT(DISTINCT resi) AS total
        FROM scans
        WHERE {active_where}
        GROUP BY name
        ORDER BY total DESC, name ASC
        """,
        active_params
    ).fetchall()
    top_seller_rows = all_seller_rows[:3]
    implant_sources = conn.execute(
        f"""
        SELECT collect_staff, seller, COUNT(*) AS total
        FROM scans
        WHERE {active_where}
        GROUP BY collect_staff, seller
        """,
        active_params
    ).fetchall()
    implant_totals = {}
    for sprinter, seller, total in implant_sources:
        implant = resolve_implant(sprinter=sprinter, seller=seller)
        if implant == "-":
            continue
        implant_totals[implant] = implant_totals.get(implant, 0) + int(total or 0)
    all_implant_rows = sorted(
        implant_totals.items(),
        key=lambda item: (-item[1], item[0])
    )
    top_implant_rows = all_implant_rows[:3]
    total_implant = len(all_implant_rows)

    return jsonify({
        "ok": True,
        "mode": mode,
        "activePeriod": active_period,
        "latestDate": latest_date,
        "latestMonth": latest_month,
        "label": label,
        "kodeLabel": kode_label,
        "compareLabel": compare_label,
        "labels": labels,
        "scanPack": scan_pack,
        "scanDelivery": scan_delivery,
        "terimaSeller": terima_seller,
        "mini": mini,
        "kodeTotal": total_kode,
        "kodeDelta": kode_delta,
        "deliveryWarehouses": delivery_warehouse_summary,
        "topSeller": [[row[0], int(row[1] or 0)] for row in top_seller_rows],
        "topImplant": [[row[0], int(row[1] or 0)] for row in top_implant_rows],
        "allSeller": [[row[0], int(row[1] or 0)] for row in all_seller_rows],
        "allImplant": [[row[0], int(row[1] or 0)] for row in all_implant_rows],
        "indicators": {
            "totalKode": total_kode,
            "totalResi": total_resi,
            "resiTanpaKode": resi_tanpa_kode,
            "resiAbnormal": resi_abnormal_total,
            "totalImplant": total_implant,
            "totalSeller": total_seller,
            "proses": proses_total,
            "sudahTerima": status_terima_total,
            "scanPack": scan_pack_total,
            "scanDelivery": scan_delivery_total,
            "terimaSeller": terima_seller_total,
            "selisihHv": selisih_hv,
            "selisihDelivery": max(scan_pack_total - scan_delivery_total, 0),
            "totalComplaint": total_complaint,
            "totalAutoclaim": total_autoclaim,
            "progress": f"{progress:.2f}%"
        }
    })


@app.route('/dashboard_data')
def dashboard_data():

    date = request.args.get("date")
    filter_type = request.args.get("filter","ALL")
    mode = request.args.get("mode", "full")
    summary_only = mode == "summary"
    period = str(request.args.get("period") or "day").strip().lower()
    requested_sort = str(request.args.get("sort_by") or "waktu").strip().lower()
    sort_direction = (
        "ASC" if str(request.args.get("sort_dir") or "desc").lower() == "asc"
        else "DESC"
    )
    sort_columns = {
        "waktu": "waktu",
        "kode": "kode",
        "resi": "resi",
        "seller": "seller",
        "sprinter": "collect_staff",
        "status": "status",
        "station": "scan_tracer_station",
        "badges": "badges",
        "sla": "waktu_scan",
        "keterangan": "received_at"
    }
    sort_column = sort_columns.get(requested_sort, "waktu")

    if not date:
        date = datetime.now().strftime("%Y-%m-%d")

    if period == "month":
        month_value = str(date or "").strip()
        try:
            range_start_dt = datetime.strptime(month_value[:7], "%Y-%m")
        except ValueError:
            range_start_dt = datetime.now().replace(
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0
            )
        range_end_dt = (
            range_start_dt.replace(day=28) +
            timedelta(days=4)
        ).replace(day=1)
        date = range_start_dt.strftime("%Y-%m")
    else:
        try:
            range_start_dt = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            range_start_dt = datetime.now().replace(
                hour=0,
                minute=0,
                second=0,
                microsecond=0
            )
            date = range_start_dt.strftime("%Y-%m-%d")
        range_end_dt = range_start_dt + timedelta(days=1)

    day_start = range_start_dt.strftime("%Y-%m-%d 00:00:00")
    day_end = range_end_dt.strftime("%Y-%m-%d 00:00:00")

    def implant_for(sprinter, seller=None):
        return resolve_implant(sprinter=sprinter, seller=seller)

    def date_where(alias=""):

        column = f"{alias}.waktu" if alias else "waktu"
        return f"{column} >= ? AND {column} < ?"

    def date_params():

        return [day_start, day_end]

    def dashboard_scans_cte():

        return f"""
            WITH ranked_scans AS (
                SELECT
                    *,
                    ROW_NUMBER() OVER (
                        PARTITION BY TRIM(resi)
                        ORDER BY
                            CASE WHEN COALESCE(NULLIF(TRIM(kode), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                            waktu DESC,
                            CASE WHEN COALESCE(NULLIF(TRIM(badges), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                            id DESC
                    ) as scan_rank
                FROM scans
                WHERE {date_where()}
            ),
            dashboard_scans AS (
                SELECT *
                FROM ranked_scans
                WHERE scan_rank=1
            )
        """

    def dashboard_station_sql():
        return """
            CASE
                WHEN COALESCE(NULLIF(TRIM(scan_tracer_station), ''), '-') != '-'
                THEN TRIM(scan_tracer_station)
                WHEN (
                    COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                    OR COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') != '-'
                )
                THEN 'Gudang Retur 188'
                ELSE 'Gudang Modern Park (Drop Off)'
            END
        """

    def delivery_station_filter_parts(value):
        filter_value = str(value or "").strip().upper()
        prefix = "SCAN_DELIVERY_STATION_"
        if not filter_value.startswith(prefix):
            return None

        station_filters = {
            "RETUR": "Gudang Retur 188",
            "MODERN": "Gudang Modern Park (Drop Off)",
            "TANAH": "Gudang Tanah Tinggi"
        }
        rest = filter_value[len(prefix):]
        for key, station_name in station_filters.items():
            if rest == key:
                return station_name, ""
            key_prefix = f"{key}_"
            if rest.startswith(key_prefix):
                return station_name, rest[len(key_prefix):]
        return None

    def add_filter(query, params):

        station_filter = delivery_station_filter_parts(filter_type)
        if station_filter:
            station_name, station_suffix = station_filter
            query += """
            AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
            """
            query += f" AND ({dashboard_station_sql()}) = ? "
            params.append(station_name)
            if station_suffix == "CODE":
                query += " AND COALESCE(NULLIF(NULLIF(TRIM(scan_delivery_code), ''), '-'), '-') != '-' "
            elif station_suffix == "HV":
                query += " AND badges LIKE '%HIGH VALUE%' "
            elif station_suffix == "ABNORMAL":
                query += " AND status='INCOMING' "

        elif filter_type == "INCOMING":

            query += " AND UPPER(TRIM(COALESCE(status, ''))) IN ('INCOMING', 'ABNORMAL') "

        elif filter_type == "OUTGOING":

            query += " AND status='OUTGOING' "

        elif filter_type == "HIGH_VALUE":

            query += " AND badges LIKE '%HIGH VALUE%' "

        elif filter_type == "PROBLEM":

            query += """
            AND (
                badges LIKE '%COMPLAINT%'
                OR badges LIKE '%AUTOCLAIM%'
                OR badges LIKE '%CLAIM INTERNAL%'
            )
            """

        elif filter_type == "SCAN_PACK":
            query += """
            AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
            """

        elif filter_type == "SCAN_DELIVERY":
            query += """
            AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
            """

        elif filter_type == "SCAN_PACK_SELLER":
            query += """
            AND COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
            """

        elif filter_type in (
            "SCAN_PACK_CODE", "SCAN_PACK_HV", "SCAN_PACK_ABNORMAL"
        ):
            query += """
            AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
            """
            if filter_type == "SCAN_PACK_CODE":
                query += " AND COALESCE(NULLIF(NULLIF(TRIM(scan_pack_code), ''), '-'), '-') != '-' "
            elif filter_type == "SCAN_PACK_HV":
                query += " AND badges LIKE '%HIGH VALUE%' "
            else:
                query += " AND status='INCOMING' "

        elif filter_type in (
            "SCAN_DELIVERY_CODE", "SCAN_DELIVERY_HV",
            "SCAN_DELIVERY_ABNORMAL"
        ):
            query += """
            AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
            """
            if filter_type == "SCAN_DELIVERY_CODE":
                query += " AND COALESCE(NULLIF(NULLIF(TRIM(scan_delivery_code), ''), '-'), '-') != '-' "
            elif filter_type == "SCAN_DELIVERY_HV":
                query += " AND badges LIKE '%HIGH VALUE%' "
            else:
                query += " AND status='INCOMING' "

        elif filter_type in (
            "SCAN_PACK_SELLER_CODE", "SCAN_PACK_SELLER_HV",
            "SCAN_PACK_SELLER_ABNORMAL"
        ):
            query += """
            AND COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
            """
            if filter_type == "SCAN_PACK_SELLER_CODE":
                query += " AND COALESCE(NULLIF(NULLIF(TRIM(kode), ''), '-'), '-') != '-' "
            elif filter_type == "SCAN_PACK_SELLER_HV":
                query += " AND badges LIKE '%HIGH VALUE%' "
            else:
                query += " AND status='INCOMING' "

        elif filter_type == "MISSING_DELIVERY":
            query += """
            AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
            AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') = '-'
            """

        elif filter_type == "MISSING_PACK_SELLER":
            query += """
            AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
            AND COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') = '-'
            """

        elif filter_type == "MISSING_HIGH_VALUE":
            query += """
            AND badges LIKE '%HIGH VALUE%'
            AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
            AND COALESCE(NULLIF(TRIM(received_at), ''), '-') = '-'
            """

        elif filter_type == "NO_CODE":
            query += """
            AND COALESCE(NULLIF(NULLIF(TRIM(kode), ''), '-'), '-') = '-'
            """

        elif filter_type == "NO_PACK":
            query += """
            AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') = '-'
            AND COALESCE(NULLIF(NULLIF(TRIM(scan_pack_code), ''), '-'), '-') = '-'
            AND COALESCE(NULLIF(NULLIF(TRIM(kode), ''), '-'), '-') = '-'
            AND UPPER(TRIM(COALESCE(status, ''))) NOT IN ('INCOMING', 'ABNORMAL')
            """

        return query, params

    with DB_LOCK:

        page = max(
            int(request.args.get("page", 1) or 1),
            1
        )
        limit = min(
            max(int(request.args.get("limit", 50) or 50), 10),
            100
        )
        offset = (page - 1) * limit

        live_query = f"""
            {dashboard_scans_cte()}
            SELECT
                waktu,
                kode,
                resi,
                seller,
                collect_staff,
                CASE
                    WHEN COALESCE(NULLIF(TRIM(scan_tracer_station), ''), '-') != '-'
                    THEN TRIM(scan_tracer_station)
                    WHEN (
                        COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                        OR COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') != '-'
                    )
                    THEN 'Gudang Retur 188'
                    ELSE 'Gudang Modern Park (Drop Off)'
                END AS station,
                COALESCE(NULLIF(TRIM(scan_by), ''), '-') AS scan_by,
                status,
                spot,
                harga,
                waktu_scan,
                badges,
                received_at,
                received_photo,
                COUNT(*) OVER() AS total_filtered
            FROM dashboard_scans
            WHERE 1=1
        """

        live_params = date_params()
        live_query, live_params = add_filter(live_query, live_params)

        total_live_query = f"""
            {dashboard_scans_cte()}
            SELECT COUNT(*)
            FROM dashboard_scans
            WHERE 1=1
        """
        total_live_params = date_params()
        total_live_query, total_live_params = add_filter(
            total_live_query,
            total_live_params
        )

        live_query += f"""
            ORDER BY
                CASE WHEN {sort_column} IS NULL OR TRIM(CAST({sort_column} AS TEXT))='' THEN 1 ELSE 0 END,
                {sort_column} {sort_direction},
                id DESC
            LIMIT ? OFFSET ?
        """
        # Tabel live selalu diminta lagi oleh loadLiveTable(). Jangan jalankan
        # CTE berat ini dua kali saat request dashboard penuh/summary.
        if mode != "live":
            live = []
            total_live = 0
        else:
            live_params.extend([limit, offset])

            live = cursor.execute(
                live_query,
                live_params
            ).fetchall()

            total_live = live[0][14] if live else 0

        summary = {
            "total_scan":0,
            "problem":0,
            "high_value":0,
            "incoming":0,
            "outgoing":0,
            "problem_total":0,
            "complaint_total":0,
            "autoclaim_total":0,
            "internal_total":0,
            "scan_pack_codes":0,
            "scan_pack_resi":0,
            "scan_delivery_codes":0,
            "scan_delivery_resi":0,
            "scan_pack_seller_codes":0,
            "scan_pack_seller_resi":0,
            "scan_pack_hv":0,
            "scan_delivery_hv":0,
            "scan_pack_seller_hv":0,
            "scan_pack_abnormal":0,
            "scan_delivery_abnormal":0,
            "scan_pack_seller_abnormal":0,
            "scan_pack_seller_total_seller":0,
            "total_without_code":0,
            "total_implant":0,
            "total_seller":0,
        }
        sellers = []
        sprinters = []
        implant_sources = []
        kodes = []
        chart = []
        problems = []
        delivery_stations = []

        if mode != "live":

            summary_row = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    COUNT(*) as total_scan,
                    SUM(CASE
                        WHEN UPPER(TRIM(COALESCE(status, ''))) IN ('INCOMING', 'ABNORMAL')
                        THEN 1 ELSE 0 END) as incoming,
                    SUM(CASE WHEN status='OUTGOING' THEN 1 ELSE 0 END) as outgoing,
                    SUM(CASE WHEN badges LIKE '%HIGH VALUE%' THEN 1 ELSE 0 END) as high_value,
                    SUM(CASE WHEN badges LIKE '%COMPLAINT%' THEN 1 ELSE 0 END) as complaint_total,
                    SUM(CASE WHEN badges LIKE '%AUTOCLAIM%' THEN 1 ELSE 0 END) as autoclaim_total,
                    SUM(CASE WHEN badges LIKE '%CLAIM INTERNAL%' THEN 1 ELSE 0 END) as internal_total,
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') = '-'
                         AND COALESCE(NULLIF(NULLIF(TRIM(scan_pack_code), ''), '-'), '-') = '-'
                         AND COALESCE(NULLIF(NULLIF(TRIM(kode), ''), '-'), '-') = '-'
                         AND UPPER(TRIM(COALESCE(status, ''))) NOT IN ('INCOMING', 'ABNORMAL')
                        THEN 1 ELSE 0 END) as total_without_code,
                    COUNT(DISTINCT CASE
                        WHEN COALESCE(NULLIF(TRIM(seller), ''), '-') != '-'
                        THEN TRIM(seller) END) as total_seller
                FROM dashboard_scans
                WHERE 1=1
            """, date_params()).fetchone()

            total_problem = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT COUNT(*)
                FROM dashboard_scans
                WHERE 1=1
                AND (
                    badges LIKE '%COMPLAINT%'
                    OR badges LIKE '%AUTOCLAIM%'
                    OR badges LIKE '%CLAIM INTERNAL%'
                )
            """, date_params()).fetchone()[0]

            summary = {
                "total_scan":summary_row[0] or 0,
                "incoming":summary_row[1] or 0,
                "outgoing":summary_row[2] or 0,
                "high_value":summary_row[3] or 0,
                "problem":total_problem or 0,
                "problem_total":total_problem or 0,
                "complaint_total":summary_row[4] or 0,
                "autoclaim_total":summary_row[5] or 0,
                "internal_total":summary_row[6] or 0,
                "total_without_code":summary_row[7] or 0,
                "total_seller":summary_row[8] or 0,
                "total_implant":0,
            }

            timeline_row = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    COUNT(DISTINCT CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
                        AND COALESCE(NULLIF(NULLIF(TRIM(scan_pack_code), ''), '-'), '-') != '-'
                        THEN TRIM(scan_pack_code) END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
                        THEN 1 ELSE 0 END),
                    COUNT(DISTINCT CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                        AND COALESCE(NULLIF(NULLIF(TRIM(scan_delivery_code), ''), '-'), '-') != '-'
                        THEN TRIM(scan_delivery_code) END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                        THEN 1 ELSE 0 END),
                    COUNT(DISTINCT CASE
                        WHEN COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
                        AND COALESCE(NULLIF(NULLIF(TRIM(kode), ''), '-'), '-') != '-'
                        THEN TRIM(kode) END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
                        THEN 1 ELSE 0 END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
                        AND badges LIKE '%HIGH VALUE%'
                        THEN 1 ELSE 0 END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                        AND badges LIKE '%HIGH VALUE%'
                        THEN 1 ELSE 0 END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
                        AND badges LIKE '%HIGH VALUE%'
                        THEN 1 ELSE 0 END),
                    COUNT(DISTINCT CASE
                        WHEN COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
                        AND COALESCE(NULLIF(TRIM(seller), ''), '-') != '-'
                        THEN TRIM(seller) END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-'
                        AND status='INCOMING'
                        THEN 1 ELSE 0 END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                        AND status='INCOMING'
                        THEN 1 ELSE 0 END),
                    SUM(CASE
                        WHEN COALESCE(NULLIF(TRIM(received_at), ''), '-') != '-'
                        AND status='INCOMING'
                        THEN 1 ELSE 0 END)
                FROM dashboard_scans
            """, date_params()).fetchone()

            summary.update({
                "scan_pack_codes": timeline_row[0] or 0,
                "scan_pack_resi": timeline_row[1] or 0,
                "scan_delivery_codes": timeline_row[2] or 0,
                "scan_delivery_resi": timeline_row[3] or 0,
                "scan_pack_seller_codes": timeline_row[4] or 0,
                "scan_pack_seller_resi": timeline_row[5] or 0,
                "scan_pack_hv": timeline_row[6] or 0,
                "scan_delivery_hv": timeline_row[7] or 0,
                "scan_pack_seller_hv": timeline_row[8] or 0,
                "scan_pack_seller_total_seller": timeline_row[9] or 0,
                "scan_pack_abnormal": timeline_row[10] or 0,
                "scan_delivery_abnormal": timeline_row[11] or 0,
                "scan_pack_seller_abnormal": timeline_row[12] or 0,
            })

            delivery_stations = []

            
            if summary_only:
                implant_sources = cursor.execute(f"""
                    {dashboard_scans_cte()}
                    SELECT
                        collect_staff,
                        seller,
                        COUNT(*) as total
                    FROM dashboard_scans
                    WHERE 1=1
                    GROUP BY collect_staff, seller
                """, date_params()).fetchall()

                implant_totals = {}
                for sprinter, seller, total in implant_sources:
                    implant = implant_for(sprinter, seller)
                    if implant == "-":
                        continue
                    implant_totals[implant] = (
                        implant_totals.get(implant, 0) +
                        total
                    )
                summary["total_implant"] = len(implant_totals)

                return jsonify({
                    "summary":summary,
                    "delivery_stations":delivery_stations,
                    "pagination":{
                        "page":page,
                        "limit":limit,
                        "total":0,
                        "total_pages":0
                    },
                    "live":[],
                    "chart":[],
                    "problems":[],
                    "sellers":[],
                    "sprinters":[],
                    "implants":[],
                    "kodes":[]
                })

            chart = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    strftime('%H', waktu) as jam,
                    COUNT(*)
                FROM dashboard_scans
                WHERE 1=1
                GROUP BY jam
                ORDER BY jam
            """, date_params()).fetchall()

            problems = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    resi,
                    seller,
                    status,
                    waktu
                FROM dashboard_scans
                WHERE 1=1
                AND (
                    status='PROBLEM'
                    OR harga >= 1000000
                )
                ORDER BY waktu DESC, id DESC
                LIMIT 15
            """, date_params()).fetchall()

            sellers = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    seller,
                    COUNT(*) as total
                FROM dashboard_scans
                WHERE 1=1
                GROUP BY seller
                ORDER BY total DESC
                LIMIT 100
            """, date_params()).fetchall()

            sprinters = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    collect_staff,
                    COUNT(*) as total
                FROM dashboard_scans
                WHERE 1=1
                GROUP BY collect_staff
                ORDER BY total DESC
                LIMIT 100
            """, date_params()).fetchall()

            implant_sources = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    collect_staff,
                    seller,
                    COUNT(*) as total
                FROM dashboard_scans
                WHERE 1=1
                GROUP BY collect_staff, seller
            """, date_params()).fetchall()

            kodes = cursor.execute(f"""
                {dashboard_scans_cte()}
                SELECT
                    COALESCE(NULLIF(TRIM(kode), ''), '-') as kode,
                    COUNT(*) as total
                FROM dashboard_scans
                WHERE 1=1
                AND COALESCE(NULLIF(TRIM(kode), ''), '-') != '-'
                GROUP BY COALESCE(NULLIF(TRIM(kode), ''), '-')
                ORDER BY total DESC
                LIMIT 100
            """, date_params()).fetchall()

    implant_totals = {}

    for sprinter, seller, total in implant_sources:

        implant = implant_for(sprinter, seller)

        if implant == "-":

            continue

        implant_totals[implant] = (
            implant_totals.get(implant, 0) +
            total
        )

    implants = sorted(
        implant_totals.items(),
        key=lambda item: item[1],
        reverse=True
    )

    summary["total_implant"] = len(implants)

    return jsonify({

        "summary":summary,

        "delivery_stations":delivery_stations,

        "pagination":{
            "page":page,
            "limit":limit,
            "total":total_live,
            "total_pages":
                (total_live + limit - 1) // limit
        },

        "live":[
            {
                "waktu":x[0],
                "kode":x[1],
                "resi":x[2],
                "seller":x[3],
                "sprinter":x[4],
                "implant":implant_for(x[4], x[3]),
                "station":x[5],
                "scan_by":x[6],
                "status":x[7],
                "spot":x[8],
                "harga":x[9],
                "waktu_scan":x[10],
                "badges":x[11],
                "received_at":x[12],
                "received_photo":x[13]
            }
            for x in live
        ],

        "chart":[
            {
                "jam":x[0],
                "total":x[1]
            }
            for x in chart
        ],

        "problems":[
            {
                "resi":x[0],
                "seller":x[1],
                "status":x[2],
                "waktu":x[3]
            }
            for x in problems
        ],

        "sellers":[
            {
                "seller":x[0],
                "total":x[1]
            }
            for x in sellers
        ],

        "sprinters":[
            {
                "sprinter":x[0],
                "total":x[1]
            }
            for x in sprinters
        ],

        "implants":[
            {
                "implant":x[0],
                "total":x[1]
            }
            for x in implants
        ],

        "kodes":[
            {
                "kode":x[0],
                "total":x[1]
            }
            for x in kodes
        ]

    })
ATTENDANCE_SETTINGS_DEFAULTS = {
    "clock_in_cutoff": "07:00",
    "clock_out_cutoff": "12:00",
    "user_cutoffs": {}
}

DEFAULT_ATTENDANCE_LOCATIONS = [
    {
        "id": "default-office",
        "name": "Kantor Utama",
        "latitude": -6.261473599217103,
        "longitude": 106.58763553681588,
        "radius": 100
    }
]


def normalize_attendance_location(row):
    if not isinstance(row, dict):
        return None
    try:
        latitude = float(row.get("latitude"))
        longitude = float(row.get("longitude"))
        radius = int(float(row.get("radius", 100)))
    except Exception:
        return None
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        return None
    return {
        "id": str(row.get("id") or secrets.token_hex(8))[:40],
        "name": str(row.get("name") or "Lokasi Absensi").strip()[:80],
        "latitude": round(latitude, 8),
        "longitude": round(longitude, 8),
        "radius": max(20, min(2000, radius))
    }


def get_attendance_locations():
    value = ""
    with DB_LOCK:
        conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
        try:
            row = conn_local.execute("""
                SELECT value FROM attendance_settings
                WHERE key='gps_locations'
                LIMIT 1
            """).fetchone()
            value = row[0] if row else ""
        finally:
            conn_local.close()
    try:
        raw_locations = json.loads(value or "[]")
    except Exception:
        raw_locations = []
    locations = [
        location for location in
        (normalize_attendance_location(item) for item in raw_locations)
        if location
    ]
    if locations:
        return locations
    return [dict(item) for item in DEFAULT_ATTENDANCE_LOCATIONS]


def save_attendance_locations(locations):
    normalized = [
        location for location in
        (normalize_attendance_location(item) for item in locations)
        if location
    ]
    if not normalized:
        raise ValueError("Minimal satu titik GPS harus tersedia")
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with DB_LOCK:
        conn_local = configure_sqlite_connection(sqlite3.connect(DB_FILE, timeout=15))
        try:
            conn_local.execute("""
                INSERT INTO attendance_settings (key, value, updated_at)
                VALUES ('gps_locations', ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value=excluded.value,
                    updated_at=excluded.updated_at
            """, (json.dumps(normalized), updated_at))
            conn_local.commit()
        finally:
            conn_local.close()
    return normalized


def get_account_bulk_settings():
    rows = conn.execute("""
        SELECT attendance_bypass, iphone_user, permissions,
               disable_location_lock
        FROM users
    """).fetchall()
    total = len(rows)
    iphone_count = sum(1 for row in rows if truthy_flag(row[1]))
    bypass_count = sum(
        1 for row in rows
        if truthy_flag(row[0]) or has_attendance_bypass_permission(row[2])
    )
    location_lock_disabled_count = sum(
        1 for row in rows if truthy_flag(row[3])
    )
    return {
        "total": total,
        "iphone_count": iphone_count,
        "bypass_count": bypass_count,
        "location_lock_disabled_count": location_lock_disabled_count,
        "all_iphone_user": bool(total and iphone_count == total),
        "all_attendance_bypass": bool(total and bypass_count == total),
        "all_disable_location_lock": bool(
            total and location_lock_disabled_count == total
        ),
    }


@app.route('/api/account_bulk_settings', methods=['GET', 'POST'])
def api_account_bulk_settings():
    if request.method == 'GET':
        return jsonify({"success": True, **get_account_bulk_settings()})

    payload = request.get_json(silent=True) or {}
    iphone_user = 1 if truthy_flag(payload.get("iphone_user", 0)) else 0
    attendance_bypass = 1 if truthy_flag(
        payload.get("attendance_bypass", 0)
    ) else 0
    disable_location_lock = 1 if truthy_flag(
        payload.get("disable_location_lock", 0)
    ) else 0

    with DB_LOCK:
        users = conn.execute("SELECT id, permissions FROM users").fetchall()
        for user_id, permissions in users:
            permission_items = [
                item for item in normalize_permissions(permissions)
                if item != ATTENDANCE_BYPASS_PERMISSION
            ]
            if attendance_bypass:
                permission_items.append(ATTENDANCE_BYPASS_PERMISSION)
            permissions_text = ",".join(
                sanitize_user_permissions(permission_items)
            )
            conn.execute("""
                UPDATE users
                SET iphone_user=?, attendance_bypass=?, permissions=?,
                    disable_location_lock=?
                WHERE id=?
            """, (
                iphone_user,
                attendance_bypass,
                permissions_text,
                disable_location_lock,
                user_id,
            ))
        conn.commit()

    return jsonify({
        "success": True,
        "message": f"Setting massal diterapkan ke {len(users)} akun",
        **get_account_bulk_settings(),
    })


@app.route('/api/attendance_locations', methods=['GET', 'POST'])
def api_attendance_locations():
    if request.method == 'GET':
        return jsonify({"success": True, "locations": get_attendance_locations()})

    data = request.get_json(silent=True) or request.form
    name = str(data.get("name", "")).strip()
    try:
        latitude = float(data.get("latitude"))
        longitude = float(data.get("longitude"))
        radius = int(float(data.get("radius", 100)))
    except Exception:
        return jsonify({"success": False, "error": "Koordinat atau radius tidak valid"}), 400
    if not name:
        return jsonify({"success": False, "error": "Nama lokasi wajib diisi"}), 400
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        return jsonify({"success": False, "error": "Koordinat di luar rentang yang valid"}), 400
    locations = get_attendance_locations()
    if len(locations) >= 50:
        return jsonify({"success": False, "error": "Maksimal 50 titik GPS"}), 400
    locations.append({
        "id": secrets.token_hex(8),
        "name": name,
        "latitude": latitude,
        "longitude": longitude,
        "radius": radius
    })
    locations = save_attendance_locations(locations)
    return jsonify({
        "success": True,
        "message": "Titik GPS berhasil ditambahkan",
        "locations": locations
    })


@app.route('/api/attendance_locations/<location_id>', methods=['DELETE'])
def api_delete_attendance_location(location_id):
    locations = get_attendance_locations()
    remaining = [item for item in locations if str(item.get("id")) != str(location_id)]
    if len(remaining) == len(locations):
        return jsonify({"success": False, "error": "Titik GPS tidak ditemukan"}), 404
    if not remaining:
        return jsonify({"success": False, "error": "Minimal satu titik GPS harus tersedia"}), 400
    remaining = save_attendance_locations(remaining)
    return jsonify({
        "success": True,
        "message": "Titik GPS dihapus",
        "locations": remaining
    })


def normalize_attendance_setting_time(value, fallback):

    raw = str(value or "").strip()

    try:
        parts = raw.split(":")
        if len(parts) != 2:
            return fallback
        hour = int(parts[0])
        minute = int(parts[1])
    except Exception:
        return fallback

    if hour < 0 or hour > 24 or minute < 0 or minute > 59:
        return fallback

    if hour == 24:
        minute = 0

    return f"{hour:02d}:{minute:02d}"


def get_attendance_settings():

    settings = dict(ATTENDANCE_SETTINGS_DEFAULTS)
    rows = []

    with DB_LOCK:
        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:
            rows = conn_local.execute("""
                SELECT key, value
                FROM attendance_settings
                WHERE key IN (?, ?, ?)
            """,(
                "clock_in_cutoff",
                "clock_out_cutoff",
                "user_cutoffs"
            )).fetchall()
        finally:
            conn_local.close()

    settings["has_saved_settings"] = bool(rows)

    for key, value in rows:
        if key in ["clock_in_cutoff", "clock_out_cutoff"]:
            settings[key] = normalize_attendance_setting_time(
                value,
                settings[key]
            )
        elif key == "user_cutoffs":
            try:
                settings[key] = json.loads(value or "{}")
            except Exception:
                settings[key] = {}

    return settings


def cutoff_pair_matches(value, clock_in, clock_out):

    if not isinstance(value, dict):
        return False

    return (
        str(value.get("in") or value.get("clock_in") or "").strip() == clock_in and
        str(value.get("out") or value.get("clock_out") or "").strip() == clock_out
    )


def save_attendance_settings_values(clock_in_cutoff, clock_out_cutoff, user_cutoffs=None, allow_mass_cutoff_overwrite=False):

    existing_settings = get_attendance_settings()
    existing_user_cutoffs = existing_settings.get("user_cutoffs")
    if not isinstance(existing_user_cutoffs, dict):
        existing_user_cutoffs = {}

    incoming_user_cutoffs = user_cutoffs if isinstance(user_cutoffs, dict) else {}
    normalized_clock_in = normalize_attendance_setting_time(
        clock_in_cutoff,
        existing_settings.get("clock_in_cutoff") or ATTENDANCE_SETTINGS_DEFAULTS["clock_in_cutoff"]
    )
    normalized_clock_out = normalize_attendance_setting_time(
        clock_out_cutoff,
        existing_settings.get("clock_out_cutoff") or ATTENDANCE_SETTINGS_DEFAULTS["clock_out_cutoff"]
    )

    merged_user_cutoffs = dict(existing_user_cutoffs)
    for key, value in incoming_user_cutoffs.items():
        existing_value = existing_user_cutoffs.get(key)
        if (
            not allow_mass_cutoff_overwrite and
            existing_value and
            not cutoff_pair_matches(existing_value, normalized_clock_in, normalized_clock_out) and
            cutoff_pair_matches(value, normalized_clock_in, normalized_clock_out)
        ):
            continue
        merged_user_cutoffs[key] = value

    settings = {
        "clock_in_cutoff": normalized_clock_in,
        "clock_out_cutoff": normalized_clock_out,
        "user_cutoffs": merged_user_cutoffs
    }

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with DB_LOCK:
        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:
            for key, value in settings.items():
                conn_local.execute("""
                    INSERT INTO attendance_settings (key, value, updated_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(key) DO UPDATE SET
                        value=excluded.value,
                        updated_at=excluded.updated_at
                """,(
                    key,
                    json.dumps(value) if key == "user_cutoffs" else value,
                    updated_at
                ))
            conn_local.commit()
        finally:
            conn_local.close()
    settings["has_saved_settings"] = True

    return settings

@app.route('/attendance_dashboard')
def attendance_dashboard():

    if not check_login():
        return redirect('/')

    if not has_permission(
        "DASHBOARD_ABSEN"
    ):
        return redirect('/')

    return render_template(
        'attendance_dashboard.html'
    )


def serve_attendance_media(upload_folder, filename):
    """Serve local legacy media, otherwise use the canonical VPS copy."""
    requested_name = str(filename or "").strip()
    safe_name = os.path.basename(requested_name)

    if (
        not safe_name
        or safe_name != requested_name
        or safe_name in {".", ".."}
        or "\x00" in safe_name
    ):
        return "", 404

    local_folder = os.path.join(
        BASE_DIR,
        "static",
        "uploads",
        upload_folder
    )
    local_path = os.path.join(local_folder, safe_name)

    if os.path.isfile(local_path):
        return send_from_directory(local_folder, safe_name)

    media_origin = str(os.environ.get(
        "ATTENDANCE_MEDIA_ORIGIN",
        "https://attendance-api.asyscntr.com"
    )).strip().rstrip("/")

    return redirect(
        media_origin
        + "/static/uploads/"
        + quote(upload_folder, safe="")
        + "/"
        + quote(safe_name, safe=""),
        code=302
    )


@app.route('/attendance_media/<path:filename>')
def attendance_media(filename):
    return serve_attendance_media("attendance", filename)


@app.route('/attendance_leave_media/<path:filename>')
def attendance_leave_media(filename):
    return serve_attendance_media("attendance_leave", filename)
    
@app.route('/attendance_dashboard_data')
def attendance_dashboard_data():

    include_empty = truthy_flag(request.args.get("include_empty", "0"))
    month = request.args.get(
        "month",
        "5"
    )
    try:
        selected_month = max(1, min(12, int(month)))
    except (TypeError, ValueError):
        selected_month = datetime.now().month
    try:
        selected_year = int(request.args.get("year") or datetime.now().year)
    except (TypeError, ValueError):
        selected_year = datetime.now().year
    range_start_dt = datetime(selected_year, selected_month, 1)
    range_end_dt = (
        range_start_dt.replace(day=28) + timedelta(days=4)
    ).replace(day=1)
    attendance_start = range_start_dt.strftime("%Y-%m-%d")
    attendance_end = range_end_dt.strftime("%Y-%m-%d")

    user_ids = {}
    account_users = {}

    with DB_LOCK:

        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:

            superman_usernames = {
                str(row[0] or "").lower()
                for row in conn_local.execute("""

                    SELECT username

                    FROM users

                    WHERE UPPER(level)=?
                    OR LOWER(username)=?

                """,(
                    "SUPERMAN",
                    "superman"
                )).fetchall()
            }

            try:

                rows = conn_local.execute("""

                SELECT

                    username,
                    fullname,
                    photo,
                    tanggal,
                    jam,
                    latitude,
                    longitude,
                    address,
                    clock_out,
                    clock_out_at,
                    clock_out_photo,
                    clock_out_latitude,
                    clock_out_longitude,
                    clock_out_address,
                    shift_id,
                    COALESCE(device_info, ''),
                    COALESCE(clock_out_device_info, ''),
                    COALESCE(ip_address, ''),
                    COALESCE(clock_out_ip_address, ''),
                    COALESCE(app_version, '')

                FROM attendance

                WHERE tanggal >= ?
                AND tanggal < ?

                ORDER BY id DESC

                """,(attendance_start, attendance_end)).fetchall()

            except sqlite3.OperationalError:

                rows = conn_local.execute("""

                SELECT

                    username,
                    fullname,
                    photo,
                    tanggal,
                    jam,
                    '' AS latitude,
                    '' AS longitude,
                    '' AS address,
                    '' AS clock_out,
                    '' AS clock_out_at,
                    '' AS clock_out_photo,
                    '' AS clock_out_latitude,
                    '' AS clock_out_longitude,
                    '' AS clock_out_address,
                    '' AS shift_id,
                    '' AS device_info,
                    '' AS clock_out_device_info,
                    '' AS ip_address,
                    '' AS clock_out_ip_address,
                    '' AS app_version

                FROM attendance

                WHERE tanggal >= ?
                AND tanggal < ?

                ORDER BY id DESC

                """,(attendance_start, attendance_end)).fetchall()

            account_rows = conn_local.execute("""

                SELECT
                    id,
                    username,
                    fullname

                FROM users

                WHERE UPPER(level) != ?
                AND LOWER(username) != ?

                ORDER BY fullname COLLATE NOCASE, username COLLATE NOCASE

            """,(
                "SUPERMAN",
                "superman"
            )).fetchall()

            attendance_history_usernames = {
                str(row[0] or "").lower()
                for row in conn_local.execute("""

                    SELECT DISTINCT username

                    FROM attendance

                    WHERE username IS NOT NULL
                    AND username != ''

                """).fetchall()
            }

            account_users = {}

            for account_row in account_rows:

                account_key = str(account_row[1] or "").lower()

                if not account_key:

                    continue

                account_users[account_key] = {
                    "id": account_row[0],
                    "username": account_row[1],
                    "fullname": account_row[2] or account_row[1],
                    "has_any_attendance": account_key in attendance_history_usernames
                }

            user_ids = {
                key: value["id"]
                for key, value in account_users.items()
            }

        finally:

            conn_local.close()

    with DB_LOCK:

        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:

            leave_rows = conn_local.execute("""

            SELECT

                username,
                fullname,
                tanggal,
                type,
                keterangan,
                photo

            FROM attendance_leave

            WHERE tanggal >= ?
            AND tanggal < ?

            ORDER BY id DESC

            """,(attendance_start, attendance_end)).fetchall()

        finally:

            conn_local.close()

    users = {}

    for row in rows:

        username = row[0]

        if (
            str(username or "").lower()
            in superman_usernames
        ):

            continue

        fullname = row[1]
        photo = row[2]
        tanggal = row[3]
        jam = row[4]
        latitude = row[5]
        longitude = row[6]
        address = row[7]
        clock_out = row[8] if len(row) > 8 else ''
        clock_out_at = row[9] if len(row) > 9 else ''
        clock_out_photo = row[10] if len(row) > 10 else ''
        clock_out_latitude = row[11] if len(row) > 11 else ''
        clock_out_longitude = row[12] if len(row) > 12 else ''
        clock_out_address = row[13] if len(row) > 13 else ''
        shift_id = row[14] if len(row) > 14 else ''
        device_info = row[15] if len(row) > 15 else ''
        clock_out_device_info = row[16] if len(row) > 16 else ''
        ip_address = row[17] if len(row) > 17 else ''
        clock_out_ip_address = row[18] if len(row) > 18 else ''
        app_version = row[19] if len(row) > 19 else ''

        if username not in users:

            users[username] = {

                "id": user_ids.get(
                    str(username or "").lower(),
                    ""
                ),
                "username": username,
                "fullname": fullname,
                "photo": photo,
                "days": {},
                "hadir": 0,
                "absen": 0

            }
         
        users[username]["photo"] = photo         

        try:

            day = int(
                tanggal.split("-")[2]
            )

            users[username]["days"][
                str(day)
            ] = {

                "time": jam[:5],
                "clock_in": jam[:5],
                "photo": photo,
                "clock_out": (clock_out or "")[:5],
                "clock_out_at": clock_out_at,
                "clock_out_photo": clock_out_photo,
                "clock_out_latitude": clock_out_latitude,
                "clock_out_longitude": clock_out_longitude,
                "clock_out_address": clock_out_address,
                "address": address,
                "latitude": latitude,
                "longitude": longitude,
                "shift_id": shift_id,
                "device_info": device_info,
                "clock_out_device_info": clock_out_device_info,
                "ip_address": ip_address,
                "clock_out_ip_address": clock_out_ip_address,
                "app_version": app_version

            }

        except:
            pass

    for row in leave_rows:

        username = row[0]

        if (
            str(username or "").lower()
            in superman_usernames
        ):

            continue

        fullname = row[1]
        tanggal = row[2]
        leave_type = row[3]
        keterangan = row[4] or ""
        photo = row[5] or ""

        if username not in users:

            users[username] = {

                "id": user_ids.get(
                    str(username or "").lower(),
                    ""
                ),
                "username": username,
                "fullname": fullname,
                "photo": "",
                "days": {},
                "hadir": 0,
                "absen": 0

            }

        try:

            day = int(
                tanggal.split("-")[2]
            )

            existing_record = users[username]["days"].get(
                str(day),
                {}
            )

            existing_record.update({

                "type": leave_type,
                "note": keterangan,
                "proof_photo": photo

            })

            users[username]["days"][
                str(day)
            ] = existing_record

        except:
            pass

    existing_user_keys = {
        str(username or "").lower(): username
        for username in users.keys()
    }

    for account_key, account_meta in account_users.items():

        existing_username = existing_user_keys.get(
            account_key
        )

        if existing_username:

            users[existing_username]["id"] = account_meta.get(
                "id",
                users[existing_username].get("id", "")
            )
            users[existing_username]["fullname"] = users[existing_username].get(
                "fullname"
            ) or account_meta.get("fullname", "")
            users[existing_username]["has_any_attendance"] = bool(
                account_meta.get("has_any_attendance")
            )
            continue

        if not include_empty:
            continue

        account_username = account_meta.get("username")

        if not account_username:

            continue

        users[account_username] = {
            "id": account_meta.get("id", ""),
            "username": account_username,
            "fullname": account_meta.get("fullname", account_username),
            "photo": "",
            "days": {},
            "hadir": 0,
            "absen": 0,
            "has_any_attendance": bool(account_meta.get("has_any_attendance")),
            "is_no_data": not bool(account_meta.get("has_any_attendance"))
        }
    attendance_shift_map = get_all_attendance_shifts()

    for username in users:

        hadir = len(
            users[username]["days"]
        )

        users[username]["hadir"] = hadir

        users[username]["absen"] = 31 - hadir

        users[username]["is_no_data"] = not bool(
            users[username].get("has_any_attendance", hadir > 0)
        )

        users[username]["shift"] = attendance_shift_map.get(
            str(username or "").lower(),
            {}
        )
    return jsonify({

        "data": list(
            users.values()
        ),
        "settings": get_attendance_settings(),
        "shifts": attendance_shift_map

    })


@app.route('/save_attendance_manual', methods=['POST'])
def save_attendance_manual():

    if not check_login():
        return jsonify({
            "success": False,
            "error": "Session expired"
        }), 401

    if not has_permission("MANAGE_ABSEN"):
        return jsonify({
            "success": False,
            "error": "Tidak punya akses"
        }), 403

    try:
        username = request.form.get("username", "").strip()
        fullname = request.form.get("fullname", "").strip() or username
        tanggal = request.form.get("tanggal", "").strip()
        manual_type = request.form.get("type", "").strip().upper()
        manual_time = request.form.get("time", "").strip()
        shift_id = request.form.get("shift_id", "1").strip()
        if shift_id not in {"1", "2"}:
            shift_id = "1"

        if not username or not tanggal or manual_type not in {"CLOCK_IN", "CLOCK_OUT"}:
            return jsonify({
                "success": False,
                "error": "Data manual tidak lengkap"
            }), 400

        try:
            datetime.strptime(tanggal, "%Y-%m-%d")
        except Exception:
            return jsonify({
                "success": False,
                "error": "Tanggal tidak valid"
            }), 400

        manual_time = normalize_attendance_setting_time(
            manual_time,
            datetime.now().strftime("%H:%M")
        )
        manual_clock = manual_time + ":00"
        manual_text = "Data Input manual"
        photo_name = get_last_attendance_photo(username)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with DB_LOCK:
            conn_local = configure_sqlite_connection(
                sqlite3.connect(
                    DB_FILE,
                    timeout=15
                )
            )

            try:
                row = conn_local.execute("""
                    SELECT id
                    FROM attendance
                    WHERE username=?
                    AND tanggal=?
                    ORDER BY id DESC
                    LIMIT 1
                """,(
                    username,
                    tanggal
                )).fetchone()

                if manual_type == "CLOCK_IN":
                    if row:
                        conn_local.execute("""
                            UPDATE attendance
                            SET fullname=?,
                                jam=?,
                                photo=?,
                                latitude=?,
                                longitude=?,
                                address=?,
                                shift_id=?,
                                manual_entry=1
                            WHERE id=?
                        """,(
                            fullname,
                            manual_clock,
                            photo_name,
                            manual_text,
                            manual_text,
                            manual_text,
                            shift_id,
                            row[0]
                        ))
                    else:
                        conn_local.execute("""
                            INSERT INTO attendance (
                                username,
                                fullname,
                                tanggal,
                                jam,
                                photo,
                                latitude,
                                longitude,
                                address,
                                created_at,
                                shift_id,
                                manual_entry
                            )
                            VALUES (?,?,?,?,?,?,?,?,?,?,1)
                        """,(
                            username,
                            fullname,
                            tanggal,
                            manual_clock,
                            photo_name,
                            manual_text,
                            manual_text,
                            manual_text,
                            now,
                            shift_id
                        ))
                else:
                    clock_out_at = tanggal + " " + manual_clock
                    if row:
                        conn_local.execute("""
                            UPDATE attendance
                            SET fullname=?,
                                clock_out=?,
                                clock_out_at=?,
                                clock_out_photo=?,
                                clock_out_latitude=?,
                                clock_out_longitude=?,
                                clock_out_address=?,
                                shift_id=?,
                                manual_entry=1
                            WHERE id=?
                        """,(
                            fullname,
                            manual_clock,
                            clock_out_at,
                            photo_name,
                            manual_text,
                            manual_text,
                            manual_text,
                            shift_id,
                            row[0]
                        ))
                    else:
                        conn_local.execute("""
                            INSERT INTO attendance (
                                username,
                                fullname,
                                tanggal,
                                jam,
                                photo,
                                latitude,
                                longitude,
                                address,
                                clock_out,
                                clock_out_at,
                                clock_out_photo,
                                clock_out_latitude,
                                clock_out_longitude,
                                clock_out_address,
                                created_at,
                                shift_id,
                                manual_entry
                            )
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
                        """,(
                            username,
                            fullname,
                            tanggal,
                            "",
                            photo_name,
                            manual_text,
                            manual_text,
                            manual_text,
                            manual_clock,
                            clock_out_at,
                            photo_name,
                            manual_text,
                            manual_text,
                            manual_text,
                            now,
                            shift_id
                        ))

                conn_local.commit()
            finally:
                conn_local.close()

        return jsonify({
            "success": True
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/save_attendance_shift', methods=['POST'])
def save_attendance_shift():

    if not check_login():
        return jsonify({
            "success": False,
            "error": "Session expired"
        }), 401

    if not has_permission("DASHBOARD_ABSEN"):
        return jsonify({
            "success": False,
            "error": "Tidak punya akses"
        }), 403

    try:
        payload = request.get_json(silent=True) or {}
        usernames = payload.get("usernames") or []
        if isinstance(usernames, str):
            usernames = [usernames]
        usernames = [
            str(username or "").strip()
            for username in usernames
            if str(username or "").strip()
        ]

        if not usernames:
            return jsonify({
                "success": False,
                "error": "Pilih user dulu"
            }), 400

        total_shift = int(payload.get("total_shift") or 1)
        total_shift = max(1, min(total_shift, 2))
        waktu_shift = str(payload.get("waktu_shift") or "").strip()
        shift1_clock_in = normalize_attendance_setting_time(payload.get("shift1_clock_in"), "07:00")
        shift1_clock_out = normalize_attendance_setting_time(payload.get("shift1_clock_out"), "12:00")
        shift2_clock_in = normalize_attendance_setting_time(payload.get("shift2_clock_in"), shift1_clock_in)
        shift2_clock_out = normalize_attendance_setting_time(payload.get("shift2_clock_out"), shift1_clock_out)
        updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with DB_LOCK:
            conn_local = configure_sqlite_connection(
                sqlite3.connect(
                    DB_FILE,
                    timeout=15
                )
            )

            try:
                for username in usernames:
                    conn_local.execute("""
                        INSERT INTO attendance_shifts (
                            username,
                            total_shift,
                            waktu_shift,
                            shift1_clock_in,
                            shift1_clock_out,
                            shift2_clock_in,
                            shift2_clock_out,
                            updated_at
                        )
                        VALUES (?,?,?,?,?,?,?,?)
                        ON CONFLICT(username)
                        DO UPDATE SET
                            total_shift=excluded.total_shift,
                            waktu_shift=excluded.waktu_shift,
                            shift1_clock_in=excluded.shift1_clock_in,
                            shift1_clock_out=excluded.shift1_clock_out,
                            shift2_clock_in=excluded.shift2_clock_in,
                            shift2_clock_out=excluded.shift2_clock_out,
                            updated_at=excluded.updated_at
                    """,(
                        username,
                        total_shift,
                        waktu_shift,
                        shift1_clock_in,
                        shift1_clock_out,
                        shift2_clock_in,
                        shift2_clock_out,
                        updated_at
                    ))

                conn_local.commit()
            finally:
                conn_local.close()

        return jsonify({
            "success": True,
            "shifts": get_all_attendance_shifts()
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
@app.route('/save_attendance_settings', methods=['POST'])
def save_attendance_settings():

    if not check_login():
        return jsonify({
            "success": False,
            "error": "Session expired"
        }), 401

    if not has_permission("DASHBOARD_ABSEN"):
        return jsonify({
            "success": False,
            "error": "Tidak punya akses"
        }), 403

    payload = request.get_json(silent=True) or request.form

    settings = save_attendance_settings_values(
        payload.get(
            "clock_in_cutoff",
            payload.get("clock_in", "07:00")
        ),
        payload.get(
            "clock_out_cutoff",
            payload.get("clock_out", "12:00")
        ),
        payload.get("user_cutoffs", {}),
        str(payload.get("allow_mass_cutoff_overwrite", "")).lower() in {"1", "true", "yes", "on"}
    )

    return jsonify({
        "success": True,
        "settings": settings
    })

@app.route('/save_attendance_leave', methods=['POST'])
def save_attendance_leave():

    if not check_login():

        return jsonify({
            "success":False,
            "error":"Session expired"
        }), 401

    if not has_permission(
        "MANAGE_ABSEN"
    ):

        return jsonify({
            "success":False,
            "error":"Tidak punya akses"
        }), 403

    try:

        username = request.form.get(
            "username",
            ""
        ).strip()

        fullname = request.form.get(
            "fullname",
            ""
        ).strip()

        tanggal = request.form.get(
            "tanggal",
            ""
        ).strip()

        leave_type = request.form.get(
            "type",
            ""
        ).strip().upper()

        keterangan = request.form.get(
            "keterangan",
            ""
        ).strip()

        if leave_type not in [
            "IZIN",
            "SAKIT"
        ]:

            return jsonify({
                "success":False,
                "error":"Jenis tidak valid"
            }), 400

        if not username or not tanggal:

            return jsonify({
                "success":False,
                "error":"User dan tanggal wajib"
            }), 400

        photo_name = ""

        photo = request.files.get(
            "photo"
        )

        if photo and photo.filename:

            upload_dir = os.path.join(
                BASE_DIR,
                "static",
                "uploads",
                "attendance_leave"
            )

            os.makedirs(
                upload_dir,
                exist_ok=True
            )

            safe_filename = secure_filename(
                photo.filename
            )

            photo_name = (
                datetime.now().strftime(
                    "%Y%m%d_%H%M%S"
                )
                + "_"
                + username
                + "_"
                + safe_filename
            )

            photo.save(
                os.path.join(
                    upload_dir,
                    photo_name
                )
            )

        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        try:

            old_row = conn_local.execute("""

            SELECT photo

            FROM attendance_leave

            WHERE username=?
            AND tanggal=?

            LIMIT 1

            """,(
                username,
                tanggal
            )).fetchone()

            if not photo_name and old_row:

                photo_name = old_row[0] or ""

            now = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )

            conn_local.execute("""

            INSERT INTO attendance_leave (

                username,
                fullname,
                tanggal,
                type,
                keterangan,
                photo,
                created_at,
                updated_at

            )

            VALUES (?,?,?,?,?,?,?,?)

            ON CONFLICT(username, tanggal)
            DO UPDATE SET

                fullname=excluded.fullname,
                type=excluded.type,
                keterangan=excluded.keterangan,
                photo=excluded.photo,
                updated_at=excluded.updated_at

            """,(
                username,
                fullname,
                tanggal,
                leave_type,
                keterangan,
                photo_name,
                now,
                now
            ))

            conn_local.commit()

        finally:

            conn_local.close()

        return jsonify({
            "success":True
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        }), 500
    
def append_live_filter_sql(query, filter_type):
    filter_type = str(filter_type or "ALL").strip().upper()

    def station_sql():
        return """
            CASE
                WHEN COALESCE(NULLIF(TRIM(scan_tracer_station), ''), '-') != '-'
                THEN TRIM(scan_tracer_station)
                WHEN (
                    COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                    OR COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') != '-'
                )
                THEN 'Gudang Retur 188'
                ELSE 'Gudang Modern Park (Drop Off)'
            END
        """

    station_filters = {
        "RETUR": "Gudang Retur 188",
        "MODERN": "Gudang Modern Park (Drop Off)",
        "TANAH": "Gudang Tanah Tinggi"
    }
    prefix = "SCAN_DELIVERY_STATION_"
    if filter_type.startswith(prefix):
        rest = filter_type[len(prefix):]
        for key, station_name in station_filters.items():
            suffix = None
            if rest == key:
                suffix = ""
            elif rest.startswith(f"{key}_"):
                suffix = rest[len(key) + 1:]
            if suffix is None:
                continue
            query += " AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-' "
            query += f" AND ({station_sql()}) = '{station_name}' "
            if suffix == "CODE":
                query += " AND COALESCE(NULLIF(NULLIF(TRIM(scan_delivery_code), ''), '-'), '-') != '-' "
            elif suffix == "HV":
                query += " AND badges LIKE '%HIGH VALUE%' "
            elif suffix == "ABNORMAL":
                query += " AND status='INCOMING' "
            return query

    simple_filters = {
        "INCOMING": " AND status='INCOMING' ",
        "OUTGOING": " AND status='OUTGOING' ",
        "HIGH_VALUE": " AND badges LIKE '%HIGH VALUE%' ",
        "SCAN_PACK": " AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-' ",
        "SCAN_DELIVERY": " AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-' ",
        "SCAN_PACK_SELLER": " AND COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') != '-' ",
        "MISSING_DELIVERY": " AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-' AND COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') = '-' ",
        "MISSING_PACK_SELLER": " AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-' AND COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') = '-' ",
        "MISSING_HIGH_VALUE": " AND badges LIKE '%HIGH VALUE%' AND COALESCE(NULLIF(TRIM(scan_pack_time), ''), '-') != '-' AND COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') = '-' ",
        "NO_CODE": " AND COALESCE(NULLIF(NULLIF(TRIM(kode), ''), '-'), '-') = '-' "
    }
    if filter_type in simple_filters:
        return query + simple_filters[filter_type]
    if filter_type == "PROBLEM":
        return query + " AND (badges LIKE '%COMPLAINT%' OR badges LIKE '%AUTOCLAIM%' OR badges LIKE '%CLAIM INTERNAL%') "

    stage_filters = {
        "SCAN_PACK_SELLER": ("scan_pack_seller_time", "scan_pack_seller_code"),
        "SCAN_DELIVERY": ("scan_delivery_time", "scan_delivery_code"),
        "SCAN_PACK": ("scan_pack_time", "scan_pack_code")
    }
    for stage_name, (time_column, code_column) in stage_filters.items():
        prefix = f"{stage_name}_"
        if not filter_type.startswith(prefix):
            continue
        query += f" AND COALESCE(NULLIF(TRIM({time_column}), ''), '-') != '-' "
        suffix = filter_type[len(prefix):]
        if suffix == "CODE":
            query += f" AND COALESCE(NULLIF(NULLIF(TRIM({code_column}), ''), '-'), '-') != '-' "
        elif suffix == "HV":
            query += " AND badges LIKE '%HIGH VALUE%' "
        elif suffix == "ABNORMAL":
            query += " AND status='INCOMING' "
        return query
    return query


@app.route('/search_resi', methods=['POST'])
def search_resi():
    try:
        data = request.get_json(silent=True) or {}
        keyword = str(data.get("keyword") or "").strip()
        search_type = str(data.get("type") or "resi").strip().lower()
        search_scope = str(data.get("scope") or "current_date").strip().lower()
        selected_date = normalize_import_date(data.get("date"))
        active_filter = str(data.get("filter") or "ALL").strip().upper()

        if not keyword:
            return jsonify({
                "success": False,
                "error": "Keyword kosong"
            })

        if search_scope not in ("all_date", "current_date", "filter"):
            search_scope = "current_date"
        if search_scope == "filter" and active_filter == "ALL":
            return jsonify({
                "success": False,
                "error": "Klik salah satu filter card terlebih dahulu"
            }), 400

        search_column = {
            "seller": "seller",
            "kode": "kode",
            "resi": "resi"
        }.get(search_type, "resi")
        keywords = [
            item.strip()
            for item in re.split(r"[\r\n,;]+", keyword)
            if item.strip()
        ]
        if not keywords:
            return jsonify({
                "success": False,
                "error": "Keyword kosong"
            }), 400
        keyword_where = " OR ".join(
            f"{search_column} LIKE ?" for _ in keywords
        )
        if search_scope in ("current_date", "filter"):
            day_start = f"{selected_date} 00:00:00"
            day_end = (
                datetime.strptime(selected_date, "%Y-%m-%d") + timedelta(days=1)
            ).strftime("%Y-%m-%d 00:00:00")
            query = f"""
                WITH ranked_search_scans AS (
                    SELECT *, ROW_NUMBER() OVER (
                        PARTITION BY TRIM(resi)
                        ORDER BY
                            CASE WHEN COALESCE(NULLIF(TRIM(kode), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                            waktu DESC,
                            CASE WHEN COALESCE(NULLIF(TRIM(badges), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                            id DESC
                    ) AS search_rank
                    FROM scans
                    WHERE waktu >= ? AND waktu < ?
                )
                SELECT
                    waktu, kode, resi, seller, collect_staff,
                    CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_tracer_station), ''), '-') != '-'
                        THEN TRIM(scan_tracer_station)
                        WHEN (
                            COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                            OR COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') != '-'
                        )
                        THEN 'Gudang Retur 188'
                    ELSE 'Gudang Modern Park (Drop Off)'
                END AS station,
                    COALESCE(NULLIF(TRIM(scan_by), ''), '-') AS scan_by,
                    status, spot, harga, waktu_scan, badges, received_at, received_photo
                FROM ranked_search_scans
                WHERE search_rank=1 AND ({keyword_where})
            """
            params = [day_start, day_end] + [f"%{item}%" for item in keywords]
        else:
            query = f"""
                SELECT
                    waktu, kode, resi, seller, collect_staff,
                    CASE
                        WHEN COALESCE(NULLIF(TRIM(scan_tracer_station), ''), '-') != '-'
                        THEN TRIM(scan_tracer_station)
                        WHEN (
                            COALESCE(NULLIF(TRIM(scan_delivery_time), ''), '-') != '-'
                            OR COALESCE(NULLIF(TRIM(scan_pack_seller_time), ''), '-') != '-'
                        )
                        THEN 'Gudang Retur 188'
                    ELSE 'Gudang Modern Park (Drop Off)'
                END AS station,
                    COALESCE(NULLIF(TRIM(scan_by), ''), '-') AS scan_by,
                    status, spot, harga, waktu_scan, badges, received_at, received_photo
                FROM scans
                WHERE ({keyword_where})
            """
            params = [f"%{item}%" for item in keywords]

        if search_scope == "filter":
            query = append_live_filter_sql(query, active_filter)

        query += " ORDER BY id DESC "

        with DB_LOCK:
            rows = cursor.execute(query, params).fetchall()

        return jsonify({
            "success": True,
            "scope": search_scope,
            "filter": active_filter if search_scope == "filter" else "ALL",
            "live": [
                {
                    "waktu": x[0],
                    "kode": x[1],
                    "resi": x[2],
                    "seller": x[3],
                    "sprinter": x[4],
                    "implant": resolve_implant(sprinter=x[4], seller=x[3]),
                    "station": x[5],
                    "scan_by": x[6],
                    "status": x[7],
                    "spot": x[8],
                    "harga": x[9],
                    "waktu_scan": x[10],
                    "badges": x[11],
                    "received_at": x[12],
                    "received_photo": x[13]
                }
                for x in rows
            ]
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/export_preview')
def export_preview():

    try:

        tanggal = request.args.get("date")
        
        export_retur = (
            request.args.get("retur")
            == "true"
        )
        


        with DB_LOCK:

            # OUTGOING
            cursor.execute("""

            SELECT COUNT(*)

            FROM scans

            WHERE DATE(waktu)=?
            AND status='OUTGOING'

            """,(tanggal,))

            outgoing = cursor.fetchone()[0]

            # INCOMING
            cursor.execute("""

            SELECT COUNT(*)

            FROM scans

            WHERE DATE(waktu)=?
            AND status='INCOMING'

            """,(tanggal,))

            incoming = cursor.fetchone()[0]

            # PROBLEM
            cursor.execute("""

            SELECT COUNT(*)

            FROM scans

            WHERE DATE(waktu)=?
            AND (

                badges LIKE '%COMPLAINT%'
                OR badges LIKE '%AUTOCLAIM%'
                OR badges LIKE '%CLAIM INTERNAL%'

            )

            """,(tanggal,))

            problem = cursor.fetchone()[0]

            # HIGH VALUE
            cursor.execute("""

            SELECT COUNT(*)

            FROM scans

            WHERE DATE(waktu)=?
            AND badges LIKE '%HIGH VALUE%'

            """,(tanggal,))

            high_value = cursor.fetchone()[0]

            # TOP SELLER
            cursor.execute("""

            SELECT
                seller,
                COUNT(*) total

            FROM scans

            WHERE DATE(waktu)=?
            AND status='OUTGOING'

            GROUP BY seller

            ORDER BY total DESC

            LIMIT 10

            """,(tanggal,))

            sellers = []

            for row in cursor.fetchall():

                sellers.append({

                    "seller": row[0],
                    "total": row[1]

                })
                
            sellers.append({

                "seller": "ABNORMAL",
                "total": incoming

            })

        return jsonify({

            "success": True,

            "summary": {

                "outgoing": outgoing,
                "incoming": incoming,
                "problem": problem,
                "high_value": high_value

            },

            "sellers": sellers

        })

    except Exception as e:

        return jsonify({

            "success": False,
            "error": str(e)

        })


@app.route('/upload_har', methods=['GET', 'POST'])
def upload_har():
    global HEADERS_CACHE
    
    if not check_login():
        return redirect('/')

    if not is_superman_session():
        return redirect('/')

    if request.method == 'POST':
        file = request.files.get('har_file')
        if not file:
            return 'HAR file not found', 400

        try:
            har_data = json.load(file)
            headers = {}

            for entry in har_data.get("log", {}).get("entries", []):
                req_headers = {h['name']: h['value'] for h in entry['request']['headers']}

                if any(k.lower() == "authtoken" for k in req_headers.keys()):
                    headers = req_headers
                    break

            if not headers:
                return 'Authtoken not found in HAR', 400

            with HEADERS_LOCK:
                with open(HEADERS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(headers, f, indent=2, ensure_ascii=False)

                # reset cache supaya pakai header baru
                HEADERS_CACHE = dict(headers)

            return 'HAR uploaded dan header updated'

        except Exception as e:
            return f'Failed to parse HAR: {e}', 500

    return render_template('upload_har.html')

@app.route('/test_har', methods=['POST'])
def test_har():

    if not check_login():
        return jsonify({
            "success": False,
            "error": "Login required"
        }), 401

    if not is_superman_session():
        return jsonify({
            "success": False,
            "error": "HAR hanya bisa diakses akun SUPERMAN"
        }), 403

    try:

        with DB_LOCK:

            row = cursor.execute("""

            SELECT resi

            FROM scans

            WHERE resi IS NOT NULL
            AND TRIM(resi) != ''

            ORDER BY id DESC

            LIMIT 1

            """).fetchone()

        if not row:

            return jsonify({
                "success": False,
                "error": "Belum ada resi yang pernah discan"
            })

        resi = str(row[0]).strip()
        headers = load_headers()
        sensitive_headers = build_sensitive_headers(headers)

        payload_keyword = {
            "keywordList": [resi],
            "trackingTypeEnum": "WAYBILL",
            "countryId": "1"
        }

        payload_work = {
            "waybillNo": resi,
            "countryId": "1"
        }

        payload_sensitive = {
            "waybillNoList": [resi],
            "type": 15,
            "menuCode": "SEND_WAYBILL",
            "countryId": "1"
        }


        test_date = normalize_import_date()

        payload_return_management = {
            "current": 1,
            "size": 20,
            "waybillId": "",
            "waybillIds": [resi],
            "applyNetworkCode": "",
            "applyNetworkCodes": [],
            "countryId": "1",
            "dateType": 1,
            "startTime": f"{test_date} 00:00:00",
            "endTime": f"{test_date} 23:59:59",
            "exportType": 3,
            "expressTypeCodes": [],
            "orderSourceCodes": [],
            "selectTime": "",
            "status": "",
            "totalCount": 1,
            "type": 3
        }

        payload_package_numbers = ["LE42788354"]

        tests = [
            {
                "name": "queryWorkOrder",
                "method": "POST",
                "url": "https://jmsgw.jntexpress.id/operatingplatform/podTracking/queryWorkOrder",
                "json": payload_work
            },
            {
                "name": "keywordTracking",
                "method": "POST",
                "url": "https://jmsgw.jntexpress.id/operatingplatform/podTracking/inner/query/keywordList",
                "json": payload_keyword
            },
            {
                "name": "keywordTrackingNew",
                "method": "POST",
                "url": TRACKING_NEW_URL,
                "json": payload_keyword
            },
            {
                "name": "omsWaybillDetail",
                "method": "GET",
                "url": f"https://jmsgw.jntexpress.id/networkmanagement/omsWaybill/detail?waybillNo={resi}"
            },
            {
                "name": "commonWaybill",
                "method": "GET",
                "url": f"https://jmsgw.jntexpress.id/servicequality/thirdService/waybill/commonWaybillListByWaybillNos?waybillNos={resi}"
            },
            {
                "name": "sensitiveByWaybillNo",
                "method": "POST",
                "url": "https://jmsgw.jntexpress.id/networkmanagement/waybillapi/omsWaybill/sensitiveByWaybillNo",
                "json": payload_sensitive,
                "headers": sensitive_headers
            },
            {
                "name": "returnDateForPage",
                "method": "POST",
                "url": RETURN_DATE_FOR_PAGE_URL,
                "json": payload_return_management
            },
            {
                "name": "waybillIdsByPackageNumber",
                "method": "POST",
                "url": PACKAGE_WAYBILLS_URL,
                "json": payload_package_numbers
            }
        ]
        results = []

        for item in tests:

            try:

                if item["method"] == "POST":

                    response = req_session.post(
                        item["url"],
                        headers=item.get("headers") or headers,
                        json=item.get("json"),
                        timeout=10
                    )

                else:

                    response = req_session.get(
                        item["url"],
                        headers=item.get("headers") or headers,
                        timeout=10
                    )

                body = {}

                try:
                    body = response.json()
                except Exception:
                    body = {}

                ok = (
                    response.status_code == 200 and
                    body.get("fail") is not True and
                    body.get("code", 1) != 0
                )

                results.append({
                    "name": item["name"],
                    "status": response.status_code,
                    "success": ok,
                    "code": body.get("code"),
                    "msg": body.get("msg") or "-"
                })

            except Exception as api_error:

                results.append({
                    "name": item["name"],
                    "status": "-",
                    "success": False,
                    "code": "-",
                    "msg": str(api_error)
                })

        all_success = all(
            item["success"]
            for item in results
        )

        return jsonify({
            "success": all_success,
            "resi": resi,
            "results": results,
            "error": None if all_success else "Ada API yang gagal"
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/scan_settings', methods=['GET', 'POST'])
def scan_settings():

    if not check_login():
        return redirect('/')

    if request.method == 'POST':

        data = request.json or {}

        settings = {
            "badges": {},
            "fields": {}
        }

        for key in DEFAULT_SCAN_SETTINGS["badges"]:
            settings["badges"][key] = bool(
                data.get("badges", {}).get(key)
            )

        for key in DEFAULT_SCAN_SETTINGS["fields"]:
            settings["fields"][key] = bool(
                data.get("fields", {}).get(key)
            )

        save_scan_settings(settings)

        return jsonify({
            "success": True
        })

    if request.args.get("format") == "json":

        return jsonify({
            "success": True,
            "settings": load_scan_settings()
        })

    return render_template(
        'scan_settings.html',
        settings=load_scan_settings()
    )

@app.route('/mapping_implant', methods=['GET', 'POST'])
def mapping_implant():

    global MAPPING_CACHE

    if not check_login():
        return redirect('/')

    if request.method == 'POST':

        data = request.json or {}
        rows = data.get("rows", []) if isinstance(data, dict) else data
        saved_rows = save_mapping(rows)

        return jsonify({
            "success": True,
            "total": len(saved_rows),
            "rows": saved_rows
        })

    return render_template(
        'mapping_implant.html',
        mapping_rows=load_mapping()
    )

@app.route('/mapping_implant/import', methods=['POST'])
def mapping_implant_import():

    if not check_login():
        return jsonify({
            "success": False,
            "error": "Session expired"
        }), 401

    file = request.files.get("file")

    if not file:
        return jsonify({
            "success": False,
            "error": "File kosong"
        }), 400

    try:

        df = pd.read_excel(file)

        existing_rows = list(load_mapping())
        imported_rows = []

        def normalized_header(value):
            return re.sub(r"[^A-Z0-9]", "", str(value or "").strip().upper())

        column_lookup = {
            normalized_header(column): column
            for column in df.columns
        }
        sprinter_column = next((
            column_lookup[key]
            for key in (
                "SPRINTER", "NAMASPRINTER", "SPRINTERNAME", "KODESPRINTER",
                "KURIR", "NAMAKURIR", "COLLECTSTAFF", "DRIVER", "COURIER"
            )
            if key in column_lookup
        ), None)
        seller_column = next((
            column_lookup[key]
            for key in (
                "SELLER", "NAMASELLER", "SELLERNAME", "TOKO",
                "SENDER", "SENDERNAME", "CUSTOMER", "CUSTOMERNAME"
            )
            if key in column_lookup
        ), None)
        implant_column = next((
            column_lookup[key]
            for key in (
                "IMPLANT", "IMPLAN", "NAMAIMPLANT", "IMPLANTNAME",
                "IMPLANTID", "DROPPOINT", "GUDANG"
            )
            if key in column_lookup
        ), None)

        if not (sprinter_column or seller_column or implant_column):
            file.stream.seek(0)
            df = pd.read_excel(file, header=None)

        for _, row in df.iterrows():

            if sprinter_column or seller_column or implant_column:
                sprinter = (
                    str(row.get(sprinter_column, "")).strip()
                    if sprinter_column else ""
                )
                seller = (
                    str(row.get(seller_column, "")).strip()
                    if seller_column else ""
                )
                implant = (
                    str(row.get(implant_column, "-")).strip()
                    if implant_column else "-"
                )
            else:
                # Fallback file tanpa header yang dikenali.
                sprinter = str(row.iloc[0]).strip()
                if len(row) > 2:
                    seller = str(row.iloc[1]).strip()
                    implant = str(row.iloc[2]).strip()
                else:
                    seller = ""
                    implant = str(row.iloc[1]).strip() if len(row) > 1 else "-"

            if (
                (not sprinter or sprinter.lower() == "nan")
                and (not seller or seller.lower() == "nan")
            ):
                continue

            if seller.lower() == "nan":
                seller = ""

            if (
                not implant
                or implant.lower() == "nan"
            ):
                implant = "-"

            imported_rows.append({
                "sprinter": "" if sprinter.lower() == "nan" else sprinter,
                "seller": seller,
                "implant": implant
            })

        imported_rows = normalize_mapping_rows(imported_rows)

        if not imported_rows:
            return jsonify({
                "success": False,
                "error": (
                    "Tidak ada baris mapping yang valid. Pastikan file berisi "
                    "Sprinter atau Seller, serta kolom Implant."
                )
            }), 400
        deduplicated_import = []
        seen_sprinters = set()
        seen_sellers = set()

        # Jika file import memuat key yang sama beberapa kali, baris terakhir menang.
        for row in reversed(imported_rows):
            sprinter_key = normalize_mapping_key(row.get("sprinter"))
            seller_key = normalize_mapping_key(row.get("seller"))
            if (
                (sprinter_key and sprinter_key in seen_sprinters)
                or (seller_key and seller_key in seen_sellers)
            ):
                continue

            deduplicated_import.append(row)
            if sprinter_key:
                seen_sprinters.add(sprinter_key)
            if seller_key:
                seen_sellers.add(seller_key)

        imported_rows = list(reversed(deduplicated_import))
        imported_sprinters = {
            normalize_mapping_key(row.get("sprinter"))
            for row in imported_rows
            if normalize_mapping_key(row.get("sprinter"))
        }
        imported_sellers = {
            normalize_mapping_key(row.get("seller"))
            for row in imported_rows
            if normalize_mapping_key(row.get("seller"))
        }
        imported_references = imported_sprinters | imported_sellers

        retained_rows = []
        for row in existing_rows:
            old_sprinter = normalize_mapping_key(row.get("sprinter"))
            old_seller = normalize_mapping_key(row.get("seller"))
            if (
                (old_sprinter and old_sprinter in imported_references)
                or (old_seller and old_seller in imported_references)
            ):
                continue
            retained_rows.append(row)

        saved_rows = save_mapping(retained_rows + imported_rows)

        return jsonify({
            "success": True,
            "total": len(saved_rows),
            "imported": len(imported_rows),
            "rows": saved_rows,
            "detected_columns": {
                "sprinter": str(sprinter_column or ""),
                "seller": str(seller_column or ""),
                "implant": str(implant_column or "")
            }
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/mapping_implant/export')
def mapping_implant_export():

    if not check_login():
        return redirect('/')

    rows = []

    mapping_rows = sorted(
        load_mapping(),
        key=lambda row: (
            str(row.get("sprinter") or ""),
            str(row.get("seller") or "")
        )
    )

    for mapping_row in mapping_rows:

        sprinter = mapping_row.get("sprinter", "")
        seller = mapping_row.get("seller", "")
        implant = mapping_row.get("implant", "-")

        implant_value = str(implant).strip()

        if (
            not implant_value
            or implant_value == "-"
        ):
            implant_value = "-"
            status = "KOSONG"
        else:
            status = "TERISI"

        rows.append({
            "Sprinter": sprinter,
            "Seller": seller,
            "Implant": implant_value,
            "Keterangan": status
        })

    if not rows:
        rows.append({
            "Sprinter": "",
            "Seller": "",
            "Implant": "",
            "Keterangan": "ISI DATA DI BARIS INI"
        })

    df = pd.DataFrame(
        rows,
        columns=[
            "Sprinter",
            "Seller",
            "Implant",
            "Keterangan"
        ]
    )

    output = BytesIO()

    df.to_excel(
        output,
        index=False
    )

    output.seek(0)

    filename = "mapping_implant_export.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


import concurrent.futures

def process_bulk_resi(
    resi,
    spot="-",
    kode="-",
    import_date=None,
    bulk_import_id=None
):

    with app.test_request_context(

        json={
            "resi":resi,
            "spot":spot,
            "kode":kode,
            "import_date":import_date,
            "bulk_import_id":bulk_import_id,
            "bulk_import":True
        }

    ):

        result = scan_resi()

        return result


def update_bulk_resi_code(resi, kode, import_date=None, bulk_import_id=None):

    selected_date = normalize_import_date(import_date)
    day_start = f"{selected_date} 00:00:00"
    day_end = (
        datetime.strptime(selected_date, "%Y-%m-%d") + timedelta(days=1)
    ).strftime("%Y-%m-%d 00:00:00")

    with DB_LOCK:

        conn_local = configure_sqlite_connection(
            sqlite3.connect(DB_FILE, timeout=15)
        )

        try:
            row = conn_local.execute("""
                SELECT id
                FROM scans
                WHERE TRIM(resi)=TRIM(?)
                ORDER BY
                    CASE WHEN COALESCE(NULLIF(TRIM(kode), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                    waktu ASC,
                    id ASC
                LIMIT 1
            """, (resi,)).fetchone()

            if not row:
                return False

            if bulk_import_id:
                local_cursor = conn_local.cursor()
                _record_bulk_import_change(
                    local_cursor,
                    bulk_import_id,
                    row[0],
                    "UPDATE",
                    _scan_snapshot(local_cursor, row[0])
                )

            conn_local.execute(
                "UPDATE scans SET kode=? WHERE id=?",
                (kode, row[0])
            )
            conn_local.commit()
            return True
        finally:
            conn_local.close()


def code_used_by_other_resi_on_date(kode, resi, scan_date):

    if not _scan_value_filled(kode):
        return False

    selected_date = normalize_import_date(scan_date)
    day_start = f"{selected_date} 00:00:00"
    day_end = (
        datetime.strptime(selected_date, "%Y-%m-%d") + timedelta(days=1)
    ).strftime("%Y-%m-%d 00:00:00")

    with DB_LOCK:
        conn_local = configure_sqlite_connection(
            sqlite3.connect(DB_FILE, timeout=15)
        )
        try:
            row = conn_local.execute("""
                SELECT 1
                FROM scans
                WHERE waktu >= ? AND waktu < ?
                AND UPPER(TRIM(kode)) = UPPER(TRIM(?))
                AND UPPER(TRIM(resi)) != UPPER(TRIM(?))
                LIMIT 1
            """, (day_start, day_end, kode, resi)).fetchone()
            return bool(row)
        finally:
            conn_local.close()


def parse_jnt_datetime(value):

    text_value = str(value or "").strip()

    if not text_value:
        return None

    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M"
    ):
        try:
            return datetime.strptime(text_value[:19], date_format)
        except ValueError:
            continue

    return None


def collect_tracking_scan_details(value):

    details = []

    if isinstance(value, dict):

        if value.get("scanTime") and value.get("scanTypeName"):
            details.append(value)

        for child in value.values():
            details.extend(collect_tracking_scan_details(child))

    elif isinstance(value, list):

        for child in value:
            details.extend(collect_tracking_scan_details(child))

    return details


def save_bulk_tracking_times(
    resi,
    import_date,
    examine_time,
    scan_pack_seller_time="-",
    scan_delivery_time="-",
    scan_pack_time="-",
    scan_pack_code="-",
    scan_delivery_code="-",
    scan_pack_seller_code="-"
):

    selected_date = normalize_import_date(import_date)
    day_start = f"{selected_date} 00:00:00"
    day_end = (
        datetime.strptime(selected_date, "%Y-%m-%d") + timedelta(days=1)
    ).strftime("%Y-%m-%d 00:00:00")

    with DB_LOCK:

        conn_local = configure_sqlite_connection(
            sqlite3.connect(DB_FILE, timeout=15)
        )

        try:
            row = conn_local.execute("""
                SELECT id
                FROM scans
                WHERE TRIM(resi)=TRIM(?)
                ORDER BY
                    CASE WHEN COALESCE(NULLIF(TRIM(kode), ''), '-') != '-' THEN 1 ELSE 0 END DESC,
                    waktu ASC,
                    id ASC
                LIMIT 1
            """, (resi,)).fetchone()

            if not row:
                return False

            conn_local.execute("""
                UPDATE scans
                SET examine_time=?,
                    scan_pack_seller_time=?,
                    scan_delivery_time=?,
                    scan_pack_time=?,
                    scan_pack_code=?,
                    scan_delivery_code=?,
                    scan_pack_seller_code=?
                WHERE id=?
            """, (
                examine_time or "-",
                scan_pack_seller_time or "-",
                scan_delivery_time or "-",
                scan_pack_time or "-",
                scan_pack_code or "-",
                scan_delivery_code or "-",
                scan_pack_seller_code or "-",
                row[0]
            ))
            conn_local.commit()
            return True
        finally:
            conn_local.close()


def _tracking_package_number(scan_item):
    return str(scan_item[2].get("packageNumber") or "-").strip() or "-"


def classify_return_tracking_stages(qualified):
    pack_scans = sorted(
        (
            item for item in qualified
            if str(item[2].get("scanTypeName") or "").strip().lower() == "pack"
        ),
        key=lambda item: item[0]
    )
    delivery_scans = sorted(
        (
            item for item in qualified
            if str(item[2].get("scanTypeName") or "").strip().lower() == "scan delivery"
        ),
        key=lambda item: item[0]
    )

    first_pack = None
    seller_pack = None

    if pack_scans:
        earliest_pack = pack_scans[0]
        delivery_before_first = any(
            delivery[0] < earliest_pack[0]
            for delivery in delivery_scans
        )

        if delivery_before_first:
            # Pack yang didahului Delivery adalah Pack Seller dari alur sebelumnya.
            seller_pack = earliest_pack
        else:
            first_pack = earliest_pack

            seller_candidates = []
            for later_pack in pack_scans[1:]:
                delivery_between = any(
                    first_pack[0] < delivery[0] < later_pack[0]
                    for delivery in delivery_scans
                )
                different_date = later_pack[0].date() != first_pack[0].date()
                far_apart = (
                    later_pack[0] - first_pack[0]
                ).total_seconds() >= 3600

                if delivery_between or different_date or far_apart:
                    seller_candidates.append(later_pack)

            if seller_candidates:
                seller_pack = seller_candidates[-1]

    selected_delivery = None
    if first_pack:
        candidates = [item for item in delivery_scans if item[0] > first_pack[0]]
        if candidates:
            selected_delivery = candidates[-1]

    first_code = _tracking_package_number(first_pack) if first_pack else "-"
    seller_code = _tracking_package_number(seller_pack) if seller_pack else "-"

    return {
        "scan_pack_time": first_pack[1] if first_pack else "-",
        "scan_pack_code": first_code,
        "scan_delivery_time": selected_delivery[1] if selected_delivery else "-",
        "scan_delivery_code": first_code,
        "scan_pack_seller_time": seller_pack[1] if seller_pack else "-",
        "scan_pack_seller_code": seller_code,
        "latest_pack_package_number": (
            _tracking_package_number(pack_scans[-1]) if pack_scans else "-"
        )
    }


def enrich_bulk_return_tracking(resi, examine_time, import_date=None):

    empty_result = {
        "success": False,
        "scan_pack_seller_time": "-",
        "scan_delivery_time": "-",
        "scan_pack_time": "-",
        "scan_pack_code": "-",
        "scan_delivery_code": "-",
        "scan_pack_seller_code": "-",
        "latest_pack_package_number": "-"
    }
    cutoff = parse_jnt_datetime(examine_time)

    if not cutoff:
        save_bulk_tracking_times(resi, import_date, examine_time)
        empty_result["error"] = "examineTime tidak valid"
        return empty_result

    try:
        response = req_session.post(
            TRACKING_NEW_URL,
            headers=load_headers(),
            json={
                "keywordList": [resi],
                "trackingTypeEnum": "WAYBILL",
                "countryId": "1"
            },
            timeout=20
        )

        if response.status_code == 401:
            raise RuntimeError("Token expired")

        if response.status_code != 200:
            raise RuntimeError(f"podTrackingNew status {response.status_code}")

        data = response.json()

        if data.get("fail") is True or data.get("code", 1) == 0:
            raise RuntimeError(data.get("msg") or "podTrackingNew gagal")

        qualified = []

        for detail in collect_tracking_scan_details(data):
            scan_time_text = str(detail.get("scanTime") or "").strip()
            scan_time = parse_jnt_datetime(scan_time_text)
            scan_network = str(detail.get("scanNetworkName") or "").strip()

            if not scan_time or scan_time <= cutoff:
                continue

            if scan_network.upper() != "MODERN_PARK":
                continue

            qualified.append((scan_time, scan_time_text, detail))

        stages = classify_return_tracking_stages(qualified)

        save_bulk_tracking_times(
            resi,
            import_date,
            examine_time,
            stages["scan_pack_seller_time"],
            stages["scan_delivery_time"],
            stages["scan_pack_time"],
            stages["scan_pack_code"],
            stages["scan_delivery_code"],
            stages["scan_pack_seller_code"]
        )

        return {"success": True, **stages}

    except Exception as e:
        save_bulk_tracking_times(resi, import_date, examine_time)
        empty_result["error"] = str(e)
        return empty_result


def refresh_single_live_scan_tracking(
    resi,
    scan_date,
    status="",
    current_code="-"
):

    try:
        return_result = check_package_return_quota(
            [resi],
            import_date=scan_date
        )
        examine_times = return_result.get("examine_times", {})
        examine_time = examine_times.get(str(resi).strip().upper())

        if not examine_time:
            return {
                "success": False,
                "reason": "Bukan resi retur terkonfirmasi"
            }

        tracking_result = enrich_bulk_return_tracking(
            resi,
            examine_time,
            import_date=scan_date
        )

        if tracking_result.get("success"):
            package_number = tracking_result.get(
                "latest_pack_package_number",
                "-"
            )

            is_abnormal = str(status or "").strip().upper() == "INCOMING"
            can_assign_abnormal = (
                not is_abnormal or
                code_used_by_other_resi_on_date(
                    package_number,
                    resi,
                    scan_date
                )
            )

            code_changed = (
                _scan_value_filled(package_number) and
                str(package_number).strip().upper() !=
                str(current_code or "-").strip().upper()
            )

            if code_changed and can_assign_abnormal:
                update_bulk_resi_code(
                    resi,
                    package_number,
                    import_date=scan_date
                )

        return tracking_result

    except Exception as e:
        return {
            "success": False,
            "reason": str(e)
        }


TRACKING_REFRESH_STATE = {
    "running": False,
    "date": "",
    "processed": 0,
    "total": 0,
    "updated_at": ""
}
TRACKING_REFRESH_LOCK = threading.Lock()


def refresh_live_tracking_worker(scan_date):

    try:
        selected_date = normalize_import_date(scan_date)
        day_start = f"{selected_date} 00:00:00"
        day_end = (
            datetime.strptime(selected_date, "%Y-%m-%d") + timedelta(days=1)
        ).strftime("%Y-%m-%d 00:00:00")

        with DB_LOCK:
            conn_local = configure_sqlite_connection(
                sqlite3.connect(DB_FILE, timeout=15)
            )
            try:
                rows = conn_local.execute("""
                    SELECT resi, status, kode
                    FROM scans
                    WHERE waktu >= ? AND waktu < ?
                    AND resi IS NOT NULL
                    AND TRIM(resi) != ''
                    ORDER BY id DESC
                """, (day_start, day_end)).fetchall()
            finally:
                conn_local.close()

        unique_rows = []
        seen = set()

        for resi, status, kode in rows:
            key = str(resi or "").strip().upper()
            if not key or key in seen:
                continue
            seen.add(key)
            unique_rows.append((resi, status, kode))

        TRACKING_REFRESH_STATE.update({
            "date": selected_date,
            "processed": 0,
            "total": len(unique_rows)
        })

        for resi, status, kode in unique_rows:
            refresh_single_live_scan_tracking(
                resi,
                selected_date,
                status=status,
                current_code=kode
            )
            TRACKING_REFRESH_STATE["processed"] += 1

    finally:
        TRACKING_REFRESH_STATE["running"] = False
        TRACKING_REFRESH_STATE["updated_at"] = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )


@app.route('/refresh_live_tracking', methods=['GET', 'POST'])
def refresh_live_tracking():

    if request.method == 'GET':
        return jsonify({
            "success": True,
            "state": TRACKING_REFRESH_STATE
        })

    payload = request.get_json(silent=True) or {}
    scan_date = normalize_import_date(payload.get("date"))

    with TRACKING_REFRESH_LOCK:
        if TRACKING_REFRESH_STATE.get("running"):
            return jsonify({
                "success": True,
                "started": False,
                "state": TRACKING_REFRESH_STATE
            })

        TRACKING_REFRESH_STATE["running"] = True
        threading.Thread(
            target=refresh_live_tracking_worker,
            args=(scan_date,),
            daemon=True
        ).start()

    return jsonify({
        "success": True,
        "started": True,
        "state": TRACKING_REFRESH_STATE
    })

@app.route('/scan_resi', methods=['POST'])
def scan_resi():
    resi = request.json.get('resi', '').strip()
    spot = request.json.get('spot', '-')
    kode = request.json.get('kode', '-')
    scan_tracer_station = normalize_scan_station(
        request.json.get('station', '')
    )
    is_bulk_import = bool(request.json.get('bulk_import'))
    bulk_import_id = request.json.get('bulk_import_id')
    import_date = normalize_import_date(request.json.get('import_date'))
    scan_by = "import" if is_bulk_import else str(
        session.get("fullname") or session.get("username") or "-"
    ).strip()
    start_total = time.time()
    profile_started = time.perf_counter()
    print(f"RESI: {resi} | SPOT: {spot} | STATION: {scan_tracer_station or '-'}")
    if not resi or len(resi) < 10:
        return jsonify({'success': False, 'error': 'Resi tidak valid'}), 400

    try:
        print("=== START PROCESS ===")
        headers = load_headers()
        auth_headers = headers
        sensitive_headers = build_sensitive_headers(headers)

        
        payload_keyword = {"keywordList": [resi], "trackingTypeEnum": "WAYBILL", "countryId": "1"}
        payload_work = {"waybillNo": resi, "countryId": "1"}
        payload_sensitive = {
            "waybillNoList": [resi],
            "type": 15,
            "menuCode": "SEND_WAYBILL",
            "countryId": "1"
        }

        url_keyword = "https://jmsgw.jntexpress.id/operatingplatform/podTracking/inner/query/keywordList"
        url_work    = "https://jmsgw.jntexpress.id/operatingplatform/podTracking/queryWorkOrder"
        url_detail  = f"https://jmsgw.jntexpress.id/networkmanagement/omsWaybill/detail?waybillNo={resi}"
        url_common = f"https://jmsgw.jntexpress.id/servicequality/thirdService/waybill/commonWaybillListByWaybillNos?waybillNos={resi}"

        # Query VPS ini sebelumnya dijalankan setelah semua API J&T selesai
        # (menambah sekitar 0.5 dtk). Mulai sekarang bersamaan dengan API.
        manual_flags_future = SCAN_LOOKUP_EXECUTOR.submit(get_manual_flag_types, resi)
        results = {}

        if is_bulk_import:

            api_calls = [
                (
                    "queryWorkOrder",
                    req_session.post,
                    url_work,
                    {
                        "headers": auth_headers,
                        "json": payload_work,
                        "timeout": 8
                    }
                ),
                (
                    "keywordTracking",
                    req_session.post,
                    url_keyword,
                    {
                        "headers": auth_headers,
                        "json": payload_keyword,
                        "timeout": 8
                    }
                ),
                (
                    "omsWaybillDetail",
                    req_session.get,
                    url_detail,
                    {
                        "headers": auth_headers,
                        "timeout": 8
                    }
                ),
                (
                    "commonWaybill",
                    req_session.get,
                    url_common,
                    {
                        "headers": auth_headers,
                        "timeout": 8
                    }
                )
            ]

            for name, method, url, kwargs in api_calls:

                start_time = time.time()

                try:

                    results[name] = method(url, **kwargs)

                except Exception as e:

                    results[name] = None
                    print(
                        f"API {name} ERROR:",
                        e
                    )
                    continue

                print(f"[OK] {name} DONE, status:", results[name].status_code)
                duration = round(time.time() - start_time, 3)
                print(f"API {name}: {duration} detik")

        else:

            import concurrent.futures

            start_times = {}
            futures = {}

            start_times["queryWorkOrder"] = time.time()
            futures[EXECUTOR.submit(req_session.post, url_work, headers=auth_headers, json=payload_work, timeout=10)] = "queryWorkOrder"

            start_times["keywordTracking"] = time.time()
            futures[EXECUTOR.submit(req_session.post, url_keyword, headers=auth_headers, json=payload_keyword, timeout=10)] = "keywordTracking"

            start_times["omsWaybillDetail"] = time.time()
            futures[EXECUTOR.submit(req_session.get, url_detail, headers=auth_headers, timeout=10)] = "omsWaybillDetail"

            start_times["commonWaybill"] = time.time()
            futures[EXECUTOR.submit(req_session.get, url_common, headers=auth_headers, timeout=10)] = "commonWaybill"

            try:

                for future in concurrent.futures.as_completed(
                    futures,
                    timeout=15
                ):
                    name = futures[future]
                    try:

                        results[name] = future.result(timeout=1)

                    except Exception as e:

                        results[name] = None

                        print(
                            f"API {name} ERROR:",
                            e
                        )

                        continue
                    print(f"[OK] {name} DONE, status:", results[name].status_code)
                    duration = round(time.time() - start_times[name], 3)
                    print(f"API {name}: {duration} detik")

            except concurrent.futures.TimeoutError:

                print("API batch timeout, cancel pending futures")

                for future, name in futures.items():

                    if name not in results:
                        results[name] = None
                        future.cancel()

        profile_api_done = time.perf_counter()

        r_work = results["queryWorkOrder"]
        r_keyword = results["keywordTracking"]
        r_detail = results["omsWaybillDetail"]
        if not r_work or not r_keyword:
            return jsonify({'success': False, 'error': 'API utama J&T gagal diakses'}), 500

        if any(
            r and r.status_code == 401
            for r in [
                r_work,
                r_keyword,
                r_detail,
                results.get("commonWaybill")
            ]
        ):
            return jsonify({'success': False, 'error': 'Token expired. Silakan upload HAR baru.'}), 401

        data_keyword_raw = r_keyword.json().get('data', [])
        print("DEBUG keyword:", type(data_keyword_raw))
        details_list = data_keyword_raw[0].get('details', []) if data_keyword_raw else []
        
        data_work = r_work.json().get('data', [])
        print("DEBUG data_work:", type(data_work))

        if r_detail and r_detail.status_code == 200:
            data_detail = r_detail.json().get('data') or {}
        else:
            data_detail = {}
        print("DEBUG detail:", type(data_detail))
        profile_json_done = time.perf_counter()
        

        pickNetwork = "-"
        barang = "-"
        seller = "-"
        common_insured_amount = 0

        try:
            r_common = results.get("commonWaybill")

            if r_common and r_common.status_code == 200:
                data_common = r_common.json().get("data", [])
                if data_common:
                    common_detail = data_common[0]
                    pickNetwork = common_detail.get("pickNetworkName", "-")
                    barang = common_detail.get("goodsName", "-")
                    seller = common_detail.get("senderName") or "-"
                    common_insured_amount = common_detail.get("insuredAmount") or 0

        except:
            pass

        if seller == "-" or set(str(seller).strip()) == {"*"}:
            try:
                url_sensitive = "https://jmsgw.jntexpress.id/networkmanagement/waybillapi/omsWaybill/sensitiveByWaybillNo"
                r_sensitive = req_session.post(url_sensitive, headers=sensitive_headers, json=payload_sensitive, timeout=10)
                print("DEBUG sensitive status:", r_sensitive.status_code)

                if r_sensitive.status_code == 200:
                    data_sensitive = r_sensitive.json().get("data", [])
                    if data_sensitive:
                        seller = data_sensitive[0].get("senderName") or "-"

            except:
                pass

        profile_enrichment_done = time.perf_counter()

        goods_value = common_insured_amount or data_detail.get('insuredAmount') or 0

        detail_first = details_list[0] if details_list else {}
        
        collect_staff = "-"

        for item in details_list:
            if "Pickup" in item.get("scanTypeName", ""):
                collect_staff = item.get("scanByName", "-")
                break
        
        last_station = detail_first.get('scanNetworkName', "-")
        last_status = detail_first.get('scanTypeName', "-")
        waktu_scan = detail_first.get('scanTime', "-")

        akun_pu = "-"
        implant, mapping_source = resolve_implant_with_source(
            sprinter=collect_staff,
            seller=seller
        )

        waktu = f"{import_date} {datetime.now().strftime('%H:%M:%S')}"
        status = "OUTGOING" if pickNetwork == "MODERN_PARK" else "INCOMING"

        badges = [status]

        # ===== API J&T =====

        if any(w.get('workTypeName') == "Tiket Klaim" for w in data_work):
            badges.append("AUTOCLAIM")

        if any(w.get('workTypeName') in ["Tiket Platform", "Tiket Arbitrase"] for w in data_work):
            badges.append("COMPLAINT")

        # ===== HIGH VALUE =====

        if goods_value >= 1000000:
            badges.append("HIGH VALUE")

        # ===== INTERNAL DATABASE =====

        manual_flags_started = time.perf_counter()
        try:
            manual_types = manual_flags_future.result(timeout=5)
        except Exception as error:
            # Flag manual bukan alasan untuk menggagalkan scan yang datanya
            # sudah diterima dari J&T. Tetap catat agar masalah DB terlihat.
            print("MANUAL FLAGS LOOKUP ERROR:", error)
            manual_types = []
        profile_manual_flags_done = time.perf_counter()

        if (
            "COMPLAINT" in manual_types
            and
            "COMPLAINT" not in badges
        ):

            badges.append(
                "COMPLAINT"
            )

        if (

            "CLAIM_INTERNAL" in manual_types

            or

            "CLAIM INTERNAL" in manual_types

        ):

            if "CLAIM INTERNAL" not in badges:

                badges.append(
                    "CLAIM INTERNAL"
                )

        if "CLAIM" in manual_types:

            if "CLAIM" not in badges:

                badges.append(
                    "CLAIM"
                )
            
            
        if "COMPLAINT" in badges or "AUTOCLAIM" in badges:

            print("\nPROBLEM TERDETEKSI")
            print("RESI:", resi)
            print("SELLER:", seller)
            print("LAST STATUS:", last_status)
            print("BADGES:", badges)
            print("-----------------------------\n")    

        global SCAN_BUFFER, LAST_SCAN_TIME

        row = [
            resi,
            pickNetwork,
            seller,
            collect_staff,
            last_station,
            last_status,
            goods_value,
            waktu_scan,
            barang,
            waktu,
            status,
            spot
        ]

       
        scan_save_data = [
            resi,
            pickNetwork,
            seller,
            collect_staff,
            last_station,
            last_status,
            goods_value,
            barang,
            waktu_scan,
            waktu,
            status,
            spot,
            kode,
            ",".join(badges),
            scan_tracer_station,
            scan_by
        ]

        if "COMPLAINT" not in badges and "AUTOCLAIM" not in badges:

            with BUFFER_LOCK:
                SCAN_BUFFER.append([
                    resi,
                    pickNetwork,
                    seller,
                    collect_staff,
                    last_station,
                    last_status,
                    goods_value,
                    waktu_scan,
                    barang,
                    waktu,
                    status,
                    spot,
                    kode
                ])

        profile_logic_done = time.perf_counter()

        if is_bulk_import:
            # Jangan menunggu round-trip VPS per resi bulk. Hasilnya lebih
            # dulu dicatat aman di disk dan akan dikirim dalam satu tahap.
            stage_bulk_scan(bulk_import_id, scan_save_data)
        else:
            queue_scan_save(
                scan_save_data,
                refresh_args=(resi, import_date, status, kode)
            )
        profile_outbox_done = time.perf_counter()

        LAST_SCAN_TIME = time.time()
        
        print("TOTAL WAKTU:", round(time.time()-start_total,2), "detik")
        print(
            "PROFIL SCAN:",
            f"api_batch={profile_api_done-profile_started:.3f}s | ",
            f"json={profile_json_done-profile_api_done:.3f}s | ",
            f"enrichment={profile_enrichment_done-profile_json_done:.3f}s | ",
            f"manual_flags_db={profile_manual_flags_done-manual_flags_started:.3f}s | ",
            f"logic={profile_logic_done-profile_manual_flags_done:.3f}s | ",
            f"outbox={profile_outbox_done-profile_logic_done:.3f}s"
        )
        print("--------------------------------")
        print("DEBUG complaints type:", type(data_work))

        return jsonify({
            "success": True,
            "data": {
                "resi": resi,
                "dp_out": pickNetwork,
                "implant": implant,
                "mapping_source": mapping_source,
                "seller": seller,
                "collect_staff": collect_staff,
                "last_station": last_station,
                "last_status": last_status,
                "harga": goods_value,
                "waktu_scan": waktu_scan,
                "barang": barang,
                "waktu": waktu,
                "status": status,
                "station": dashboard_package_station(
                    scan_tracer_station,
                    "",
                    ""
                ),
                "badges": badges,
                "complaints": json.loads(json.dumps(data_work))
            }
        })

    except Exception as e:
        save_error(f"Gagal scan {resi}: {str(e)}")
        return jsonify({'success': False, 'error': 'Gagal mengambil data'}), 500

        
@app.route('/scan_common', methods=['POST'])
def scan_common():

    resi = request.json.get('resi','').strip()

    try:

        headers = load_headers()

        url_common = f"https://jmsgw.jntexpress.id/servicequality/thirdService/waybill/commonWaybillListByWaybillNos?waybillNos={resi}"

        r = req_session.get(url_common, headers=headers, timeout=10)

        data = r.json().get("data", [])

        details = data[0] if data else {}

        return jsonify({
            "success": True,
            "dp_out": details.get("pickNetworkName","-"),
            "barang": details.get("goodsName","-")
        })

    except Exception as e:
        return jsonify({"success":False})
        
@app.route('/get_complaints', methods=['POST'])
def get_complaints():
    resi = request.json.get('resi', '').strip()
    if not resi:
        return jsonify({'success': False, 'error': 'Resi tidak valid'}), 400

    try:
        headers = load_headers()
        auth_headers = {
            "authtoken": headers['authtoken'],
            "user-agent": headers['user-agent'],
            "cookie": headers['cookie'],
            "lang": headers['lang'],
            "langtype": headers['langtype'],
            "timezone": headers['timezone']
        }
        payload_work = {"waybillNo": resi, "countryId": "1"}

        r_work = requests.post("https://jmsgw.jntexpress.id/operatingplatform/podTracking/queryWorkOrder",
                               headers=auth_headers, json=payload_work, timeout=10)

        if r_work.status_code == 401:
            return jsonify({'success': False, 'error': 'Token expired. Silakan upload HAR baru.'}), 401

        data_work = r_work.json().get('data', [])
        print("DEBUG data_work:", type(data_work))
        return jsonify({'success': True, 'complaints': data_work})

    except Exception as e:
        print("ERROR ASLI:", str(e))
        import traceback
        traceback.print_exc()

        save_error(f"Gagal scan {resi}: {str(e)}")

        return jsonify({
            'success': False,
            'error': str(e)  # tampilkan error asli ke frontend
        }), 500

from flask import jsonify
import socket

@app.route("/get_ip")
def get_ip():
    hostname = socket.gethostname()
    ip = socket.gethostbyname(hostname)
    return jsonify({"ip": ip})
    
# @app.before_request
# def notify_dashboard():
#    try:
#        data = {
#            "app": "Scanner Resi",          # ubah sesuai nama app
#            "path": request.path,
#            "ip": request.remote_addr
#        }
#        requests.post(MONITOR_URL, json=data, timeout=1)
#    except:
#        pass
  

@app.route("/debug_scan", methods=["POST"])
def debug_scan():
    data = request.json.get("data", "")
    print("SCANNER INPUT:", data)
    return jsonify({"ok": True})
    
@app.route("/save_spot_photo", methods=["POST"])
def save_spot_photo():

    data = request.json
    image = data["image"]

    image = image.split(",")[1]

    folder = os.path.join(OUTPUT_DIR,"img")
    os.makedirs(folder,exist_ok=True)

    filename = datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"

    path = os.path.join(folder,filename)

    with open(path,"wb") as f:
        f.write(base64.b64decode(image))

    return jsonify({
        "success":True,
        "filename":filename
    })

@app.route("/upload_spot_photo", methods=["POST"])
def upload_spot_photo():

    file = request.files["photo"]

    today = datetime.now().strftime("%Y-%m-%d")

    folder = os.path.join(OUTPUT_DIR, "img", today)
    os.makedirs(folder, exist_ok=True)

    filename = datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"

    path = os.path.join(folder, filename)

    file.save(path)

    return jsonify({
        "success": True,
        "filename": f"{today}/{filename}"
    })   


@app.route("/check_resi", methods=["POST"])
def check_resi():

    resi = request.json.get("resi","").strip()

    print("CEK RESI:", resi)

    with DB_LOCK:

        cursor.execute("""
            SELECT spot FROM scans
            WHERE TRIM(resi)=TRIM(?)
            ORDER BY id DESC
            LIMIT 1
        """,(resi,))

        row = cursor.fetchone()

    if row:

        photo = "/output/img/" + row[0]

        return jsonify({
            "success":True,
            "resi":resi,
            "photo":photo
        })

    return jsonify({
        "success":False,
        "msg":"Resi tidak ditemukan"
    })

from flask import send_from_directory

@app.route('/output/img/<path:filename>')
def serve_image(filename):
    return send_from_directory(os.path.join(OUTPUT_DIR, "img"), filename)
    
@app.route("/export_zip")
def export_zip():

    try:

        tanggal = request.args.get("date")

        export_retur = (
            request.args.get("retur")
            == "true"
        )
        
        export_kode = (
            request.args.get("kode")
            == "true"
        )

        export_implant = (
            request.args.get("implant")
            == "true"
        )

        export_seller = (
            request.args.get("seller")
            == "true"
        )

        if not tanggal:
            return "Tanggal tidak valid"

        temp_dir = tempfile.mkdtemp()

        export_folder = os.path.join(
            temp_dir,
            f"REPORT_{tanggal}"
        )

        os.makedirs(export_folder, exist_ok=True)

        outgoing_folder = os.path.join(
            export_folder,
            "OUTGOING"
        )
        
        kode_folder = os.path.join(
            export_folder,
            "KODE"
        )

        implant_folder = os.path.join(
            export_folder,
            "IMPLANT"
        )
        
        form_retur_folder = os.path.join(
            export_folder,
            "FORM_RETUR"
        )

        os.makedirs(
            form_retur_folder,
            exist_ok=True
        )

        os.makedirs(outgoing_folder, exist_ok=True)
        
        os.makedirs(kode_folder, exist_ok=True)

        os.makedirs(implant_folder, exist_ok=True)

        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        # ALL DATA
        df = pd.read_sql_query("""

        SELECT

            resi,
            kode,
            dp_out,
            seller,
            collect_staff,
            last_status,
            waktu_scan,
            harga,
            barang,
            waktu,
            badges,
            received_at

        FROM scans

        WHERE DATE(waktu)=?

        """, conn_local, params=(tanggal,))

        if df.empty:

            return "Data kosong"

        df["Implant"] = df.apply(
            lambda row: resolve_implant(
                sprinter=row.get("collect_staff"),
                seller=row.get("seller")
            ),
            axis=1
        )

        # ===== FORMAT BADGE =====

        df["Complaint"] = df["badges"].fillna("").apply(
            lambda x:
            "YES"
            if "COMPLAINT" in x
            else ""
        )

        df["Auto Claim"] = df["badges"].fillna("").apply(
            lambda x:
            "YES"
            if "AUTOCLAIM" in x
            else ""
        )

        df["Claim Internal"] = df["badges"].fillna("").apply(
            lambda x:
            "YES"
            if "CLAIM INTERNAL" in x
            else ""
        )

        df["High Value"] = df["badges"].fillna("").apply(
            lambda x:
            "YES"
            if "HIGH VALUE" in x
            else ""
        )

        df["Status Retur"] = df["received_at"].apply(

            lambda x:

            "TERIMA"
            if pd.notna(x) and str(x).strip()
            else "PROSES"

        )

        df["Waktu Terima"] = df["received_at"].fillna("-")
        
        

        # ===== RENAME =====

        df = df.rename(columns={

            "resi":"Resi",
            "kode":"Kode",
            "dp_out":"DP Out",
            "seller":"Seller",
            "collect_staff":"Sprinter",
            "last_status":"Last Status",
            "waktu_scan":"Waktu Scan",
            "scan_by":"Scan By",
            "station":"Station",
            "harga":"Harga",
            "barang":"Barang",
            "waktu":"Waktu"

        })

        final_cols = [

            "Resi",
            "Kode",
            "DP Out",
            "Seller",
            "Sprinter",
            "Implant",
            "Last Status",
            "Waktu Scan",
            "Scan By",
            "Station",
            "Harga",
            "Barang",
            "Complaint",
            "Auto Claim",
            "Claim Internal",
            "High Value",
            "Status Retur",
            "Waktu Terima"

        ]

        df = df[final_cols]

        # ===== OUTGOING =====

        outgoing_df = pd.read_sql_query("""

        SELECT *

        FROM scans

        WHERE DATE(waktu)=?
        AND status='OUTGOING'

        """, conn_local, params=(tanggal,))
    
        seller_list = (
            outgoing_df["seller"]
            .dropna()
            .unique()
            .tolist()
        )
        
        # PER SELLER
        def process_seller_export(seller):

            seller_df = outgoing_df[
                outgoing_df["seller"] == seller
            ].copy()
            
            # REMOVE DUPLICATE RESI
            seller_df = seller_df.drop_duplicates(
                subset=["resi"]
            )

            seller_df["Implant"] = seller_df.apply(
                lambda row: resolve_implant(
                    sprinter=row.get("collect_staff"),
                    seller=row.get("seller")
                ),
                axis=1
            )

            safe_name = (
                str(seller)
                .replace("/","_")
                .replace("\\","_")
                .replace(":","_")
                .replace("*","_")
                .replace("?","_")
                .replace('"',"_")
                .replace("<","_")
                .replace(">","_")
                .replace("|","_")
                .strip(". ")
            )

            seller_file = os.path.join(
                outgoing_folder,
                f"{safe_name}.xlsx"
            )

            # ===== BADGES =====

            seller_df["Complaint"] = seller_df[
                "badges"
            ].fillna("").apply(

                lambda x:
                "YES"
                if "COMPLAINT" in x
                else ""

            )

            seller_df["Auto Claim"] = seller_df[
                "badges"
            ].fillna("").apply(

                lambda x:
                "YES"
                if "AUTOCLAIM" in x
                else ""

            )

            seller_df["Claim Internal"] = seller_df[
                "badges"
            ].fillna("").apply(

                lambda x:
                "YES"
                if "CLAIM INTERNAL" in x
                else ""

            )

            seller_df["High Value"] = seller_df[
                "badges"
            ].fillna("").apply(

                lambda x:
                "YES"
                if "HIGH VALUE" in x
                else ""

            )
            
            seller_df["Status Retur"] = seller_df["received_at"].apply(

                lambda x:

                "TERIMA"
                if pd.notna(x) and str(x).strip()
                else "PROSES"

            )

            seller_df["Waktu Terima"] = seller_df["received_at"].fillna("-")

            # ===== SLA =====

            seller_df["SLA"] = "-"

            # ===== RENAME =====

            seller_df = seller_df.rename(columns={

                "resi":"Resi",
                "kode":"Kode",
                "dp_out":"DP Out",
                "collect_staff":"Sprinter",
                "last_status":"Last Status",
                "waktu_scan":"Waktu Scan",
                "scan_by":"Scan By",
                "harga":"Harga",
                "barang":"Barang",
                "waktu":"Waktu"

            })

            # ===== FINAL COLUMN =====

            seller_df = seller_df[[

                "Resi",
                "Kode",
                "DP Out",
                "Sprinter",
                "Implant",
                "Last Status",
                "Waktu Scan",
                "Scan By",
                "Station",
                "Harga",
                "Barang",
                "SLA",
                "Complaint",
                "Auto Claim",
                "Claim Internal",
                "High Value",
                "Status Retur",
                "Waktu Terima"

            ]]

            # ===== EXPORT =====

            if export_seller:

                seller_df.to_excel(

                    seller_file,

                    index=False

                )
            
            if export_retur:
            
                # ====================================
                # GENERATE FORM RETUR
                # ====================================

                seller_folder = os.path.join(
                    form_retur_folder,
                    safe_name
                )

                os.makedirs(
                    seller_folder,
                    exist_ok=True
                )

                resi_list = seller_df["Resi"].tolist()

                # 5 kolom x 20 row
                chunk_size = 100

                chunks = [

                    resi_list[i:i + chunk_size]

                    for i in range(
                        0,
                        len(resi_list),
                        chunk_size
                    )

                ]

                for batch_index, chunk in enumerate(chunks, start=1):

                    safe_batch_seller = (
                        str(seller)
                        .replace(" ","")
                        .replace("/","")
                        .replace("\\","")
                        .replace(":","")
                        .replace("*","")
                        .replace("?","")
                        .replace('"',"")
                        .replace("<","")
                        .replace(">","")
                        .replace("|","")
                        .strip(". ")
                        .upper()
                    )

                    batch_code = (

                        f"RET-"
                        f"{tanggal.replace('-','')}-"
                        f"{safe_batch_seller}-"
                        f"{batch_index:03}"

                    )

                    with DB_LOCK:

                        conn_insert = configure_sqlite_connection(
                            sqlite3.connect(
                                DB_FILE,
                                timeout=15
                            )
                        )

                        try:

                            conn_insert.execute("""

                            INSERT INTO retur_batches (

                                batch_code,
                                seller,
                                tanggal,
                                total_resi,
                                resi_data,
                                created_at,
                                status

                            )

                            VALUES (?,?,?,?,?,?,?)

                            """, (

                                batch_code,
                                seller,
                                tanggal,
                                len(chunk),
                                json.dumps(chunk),

                                datetime.now().strftime(
                                    "%Y-%m-%d %H:%M:%S"
                                ),

                                "PENDING"

                            ))

                            conn_insert.commit()

                        finally:

                            conn_insert.close()


                    print(
                        "TEMPLATE =",
                        os.path.abspath(TEMPLATE_RETUR)
                    )

                    wb = load_workbook(
                        TEMPLATE_RETUR
                    )

                    ws = wb.active

                    # ======================
                    # HEADER
                    # ======================

                    ws["B4"] = tanggal

                    ws["C4"] = seller

                    # ======================
                    # BARCODE
                    # ======================

                    barcode_path = os.path.join(
                        tempfile.gettempdir(),
                        f"{batch_code}.png"
                    )

                    barcode_obj = Code128(
                        batch_code,
                        writer=ImageWriter()
                    )

                    with open(barcode_path, "wb") as f:

                        barcode_obj.write(
                            f,
                            {
                                "write_text": False
                            }
                        )

                    img = XLImage(barcode_path)

                    img.width = 260
                    img.height = 55

                    ws.add_image(
                        img,
                        "A7"
                    )

                    # ======================
                    # RESI
                    # ======================

                    start_row = 12

                    col_list = [
                        "A",
                        "B",
                        "C",
                        "D",
                        "E"
                    ]

                    for idx, resi in enumerate(chunk):

                        col_index = idx // 20

                        row_offset = idx % 20

                        if col_index >= len(col_list):
                            break

                        cell = f"{col_list[col_index]}{start_row + row_offset}"

                        ws[cell] = str(resi)

                    batch_file = os.path.join(

                        seller_folder,

                        f"{batch_code}.xlsx"

                    )

                    wb.save(batch_file)
                    
        with concurrent.futures.ThreadPoolExecutor(
            max_workers=4
        ) as executor:
            
            if export_kode:

                kode_df = df.copy()

                kode_df["Kode"] = kode_df["Kode"].fillna("-")

                kode_df.loc[
                    kode_df["Kode"].astype(str).str.strip() == "",
                    "Kode"
                ] = "-"

                kode_list = (

                    kode_df["Kode"]

                    .astype(str)

                    .unique()

                    .tolist()

                )

                for kode in kode_list:

                    data_kode = kode_df[
                        kode_df["Kode"] == kode
                    ]

                    safe_kode = (
                        str(kode)
                        .replace("/","_")
                        .replace("\\","_")
                        .replace(":","_")
                        .replace("*","_")
                        .replace("?","_")
                        .replace('"',"_")
                        .replace("<","_")
                        .replace(">","_")
                        .replace("|","_")
                    )

                    file_kode = os.path.join(
                        kode_folder,
                        f"{safe_kode}.xlsx"
                    )

                    data_kode.to_excel(
                        file_kode,
                        index=False
                    )

            if export_implant:

                implant_df = df.copy()

                implant_df["Implant"] = implant_df["Implant"].fillna("-")

                implant_df.loc[
                    implant_df["Implant"].astype(str).str.strip() == "",
                    "Implant"
                ] = "-"

                implant_list = (

                    implant_df["Implant"]

                    .astype(str)

                    .unique()

                    .tolist()

                )

                for implant in implant_list:

                    data_implant = implant_df[
                        implant_df["Implant"] == implant
                    ]

                    safe_implant = (
                        str(implant)
                        .replace("/","_")
                        .replace("\\","_")
                        .replace(":","_")
                        .replace("*","_")
                        .replace("?","_")
                        .replace('"',"_")
                        .replace("<","_")
                        .replace(">","_")
                        .replace("|","_")
                    )

                    file_implant = os.path.join(
                        implant_folder,
                        f"{safe_implant}.xlsx"
                    )

                    data_implant.to_excel(
                        file_implant,
                        index=False
                    )

            if (
                export_seller or
                export_retur
            ):

                list(
                    executor.map(
                        process_seller_export,
                        seller_list
                    )
                )

        conn_local.commit()
        

        # ===== MAIN FILE =====

        df.to_excel(
            os.path.join(
                export_folder,
                "ALL_DATA.xlsx"
            ),
            index=False
        )

        # ===== ZIP =====

        zip_path = os.path.join(
            temp_dir,
            f"REPORT_{tanggal}.zip"
        )

        with zipfile.ZipFile(
            zip_path,
            "w",
            zipfile.ZIP_DEFLATED
        ) as zipf:

            for root, dirs, files in os.walk(export_folder):

                for file in files:

                    file_path = os.path.join(root, file)

                    arcname = os.path.relpath(
                        file_path,
                        export_folder
                    )

                    zipf.write(file_path, arcname)

        conn_local.close()

        return send_file(

            zip_path,

            as_attachment=True,

            download_name=
                f"REPORT_{tanggal}.zip"

        )

    except Exception as e:

        return str(e)
        
@app.route('/admin_complaint')
def admin_complaint():
    return render_template('admin_complaint.html')


@app.route('/admin_claim')
def admin_claim():
    return render_template('admin_claim.html')


@app.route('/save_manual_flag', methods=['POST'])
def save_manual_flag():

    try:

        data = request.form

        type_flag = data.get("type")

        seller = data.get("seller","-")

        nominal = data.get("nominal",0)

        tanggal = data.get("tanggal")

        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        photo = "-"

        if 'photo' in request.files:

            file = request.files['photo']

            if file.filename:

                upload_dir = os.path.join(BASE_DIR, "static/uploads")

                os.makedirs(upload_dir, exist_ok=True)

                filename = f"{int(time.time())}_{file.filename}"

                file.save(os.path.join(upload_dir, filename))

                photo = filename

        conn = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )
        cursor = conn.cursor()

        if type_flag == "COMPLAINT":

            resi_list = data.get("resi","").splitlines()

            for r in resi_list:

                r = r.strip()

                if not r:
                    continue

                cursor.execute("""
                INSERT INTO manual_flags
                (resi, type, seller, nominal, tanggal, photo, created_at)
                VALUES (?,?,?,?,?,?,?)
                """, (
                    r,
                    "COMPLAINT",
                    seller,
                    0,
                    tanggal,
                    photo,
                    created_at
                ))

        elif type_flag == "CLAIM_INTERNAL":

            cursor.execute("""
            INSERT INTO manual_flags
            (resi, type, seller, nominal, tanggal, photo, created_at)
            VALUES (?,?,?,?,?,?,?)
            """, (
                data.get("resi"),
                "CLAIM_INTERNAL",
                "-",
                nominal,
                tanggal,
                photo,
                created_at
            ))

        conn.commit()
        conn.close()

        return jsonify({
            "success":True
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        })


@app.route("/retur_page")
def retur_page():
    
    if not check_login():
        return redirect('/')

    return render_template(
        "retur_page.html"
    )


@app.route("/lookup_retur_receipt", methods=["POST"])
def lookup_retur_receipt():

    payload = request.get_json(silent=True) or {}
    value = str(payload.get("value") or "").strip()

    if not value:
        return jsonify({
            "success": False,
            "error": "Resi atau kode wajib diisi"
        }), 400

    conn_local = configure_sqlite_connection(
        sqlite3.connect(DB_FILE, timeout=15)
    )

    try:
        batch_row = conn_local.execute("""
            SELECT batch_code, seller, resi_data, status
            FROM retur_batches
            WHERE UPPER(TRIM(batch_code)) = UPPER(TRIM(?))
            ORDER BY id DESC
            LIMIT 1
        """, (value,)).fetchone()

        if batch_row:
            try:
                raw_resi = json.loads(batch_row[2] or "[]")
            except (TypeError, ValueError, json.JSONDecodeError):
                raw_resi = []

            resi_list = list(dict.fromkeys(
                str(item).strip()
                for item in raw_resi
                if str(item).strip()
            ))

            if not resi_list:
                return jsonify({
                    "success": False,
                    "error": "Batch ditemukan tetapi tidak berisi resi"
                }), 404

            return jsonify({
                "success": True,
                "source_type": "BATCH",
                "source_value": batch_row[0],
                "batch_code": batch_row[0],
                "seller": batch_row[1] or "-",
                "total": len(resi_list),
                "resi_list": resi_list,
                "status": batch_row[3] or "PROCESS"
            })

        resi_row = conn_local.execute("""
            SELECT resi, kode, seller, received_at
            FROM scans
            WHERE UPPER(TRIM(resi)) = UPPER(TRIM(?))
            ORDER BY id DESC
            LIMIT 1
        """, (value,)).fetchone()

        if resi_row:
            return jsonify({
                "success": True,
                "source_type": "RESI",
                "source_value": resi_row[0],
                "batch_code": "",
                "seller": resi_row[2] or "-",
                "total": 1,
                "resi_list": [resi_row[0]],
                "status": "RECEIVED" if resi_row[3] else "PROCESS"
            })

        code_rows = conn_local.execute("""
            WITH code_scans AS (
                SELECT
                    resi,
                    seller,
                    received_at,
                    ROW_NUMBER() OVER (
                        PARTITION BY UPPER(TRIM(resi))
                        ORDER BY id DESC
                    ) AS row_rank
                FROM scans
                WHERE UPPER(TRIM(kode)) = UPPER(TRIM(?))
                AND resi IS NOT NULL
                AND TRIM(resi) != ''
            )
            SELECT resi, seller, received_at
            FROM code_scans
            WHERE row_rank = 1
            ORDER BY resi
        """, (value,)).fetchall()

        if not code_rows:
            return jsonify({
                "success": False,
                "error": "Resi atau kode tidak ditemukan"
            }), 404

        resi_list = [row[0] for row in code_rows]
        sellers = sorted({
            str(row[1]).strip()
            for row in code_rows
            if str(row[1] or "").strip() not in ("", "-")
        })
        received_count = sum(1 for row in code_rows if row[2])

        if received_count == len(code_rows):
            status = "RECEIVED"
        elif received_count:
            status = "PARTIAL"
        else:
            status = "PROCESS"

        return jsonify({
            "success": True,
            "source_type": "KODE",
            "source_value": value,
            "batch_code": "",
            "seller": sellers[0] if len(sellers) == 1 else (
                f"{len(sellers)} SELLER" if sellers else "-"
            ),
            "total": len(resi_list),
            "resi_list": resi_list,
            "status": status,
            "received_count": received_count
        })
    except Exception as error:
        print("LOOKUP TANDA TERIMA ERROR:", error)
        return jsonify({
            "success": False,
            "error": "Gagal membaca data tanda terima"
        }), 500
    finally:
        conn_local.close()
    
@app.route("/get_retur_batch")
def get_retur_batch():

    batch_code = request.args.get(
        "batch_code"
    )

    conn_local = configure_sqlite_connection(
        sqlite3.connect(
            DB_FILE,
            timeout=15
        )
    )

    row = conn_local.execute("""

    SELECT

        seller,
        total_resi,
        resi_data,
        status

    FROM retur_batches

    WHERE batch_code=?

    """,(batch_code,)).fetchone()

    conn_local.close()

    if not row:

        return jsonify({

            "success":False

        })

    return jsonify({

        "success":True,

        "seller":row[0],

        "total":row[1],

        "resi_list":
            json.loads(row[2]),

        "status":row[3]

    })
    
@app.route(
    "/submit_retur",
    methods=["POST"]
)
def submit_retur():

    try:

        batch_code = request.form.get(
            "batch_code",
            ""
        )

        raw_resi_list = json.loads(

            request.form.get(
                "resi_list",
                "[]"
            )

        )

        if not isinstance(raw_resi_list, list):
            raw_resi_list = []

        resi_list = list(dict.fromkeys(
            str(item).strip()
            for item in raw_resi_list
            if str(item).strip()
        ))

        if len(resi_list) > 10000:
            return jsonify({
                "success":False,
                "error":"Maksimal 10.000 resi sekali tanda terima"
            }), 400

        photo = request.files.get(
            "photo"
        )

        if not photo:

            return jsonify({
                "success":False,
                "error":"Foto wajib"
            })

        if not resi_list:

            return jsonify({
                "success":False,
                "error":"Resi kosong"
            })

        # =========================
        # SAVE PHOTO
        # =========================

        folder = os.path.join(

            BASE_DIR,

            "static",

            "retur_photo"

        )

        os.makedirs(
            folder,
            exist_ok=True
        )

        filename = f"{int(time.time() * 1000)}_{secrets.token_hex(4)}.jpg"

        save_path = os.path.join(
            folder,
            filename
        )

        photo.save(save_path)

        conn_local = configure_sqlite_connection(
            sqlite3.connect(
                DB_FILE,
                timeout=15
            )
        )

        # =========================
        # UPDATE RESI
        # =========================

        updated_rows = 0

        for r in resi_list:

            print("UPDATE RETUR:", r)

            conn_local.execute("""

            UPDATE scans

            SET

                badges =

                    CASE

                        WHEN COALESCE(NULLIF(TRIM(badges), ''), '-') = '-'
                        THEN 'TANDA TERIMA RETUR'

                        WHEN badges LIKE '%TANDA TERIMA RETUR%'
                        THEN badges

                        ELSE
                            badges
                            || ',TANDA TERIMA RETUR'

                    END,

                received_at=?,

                received_photo=?

            WHERE UPPER(TRIM(resi))=UPPER(TRIM(?))

            """,(

                datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),

                filename,

                r

            ))
            updated_rows += conn_local.execute(
                "SELECT changes()"
            ).fetchone()[0]

        if updated_rows == 0:
            conn_local.rollback()
            conn_local.close()
            try:
                os.remove(save_path)
            except OSError:
                pass
            return jsonify({
                "success":False,
                "error":"Tidak ada resi yang cocok untuk ditanda terima"
            }), 404

        # =========================
        # UPDATE BATCH
        # =========================

        if batch_code:

            conn_local.execute("""

            UPDATE retur_batches

            SET

                received_at=?,
                received_photo=?,
                status='RECEIVED'

            WHERE UPPER(TRIM(batch_code))=UPPER(TRIM(?))

            """,(

                datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),

                filename,

                batch_code

            ))

        conn_local.commit()

        conn_local.close()

        return jsonify({

            "success":True,
            "updated":updated_rows

        })

    except Exception as e:

        return jsonify({

            "success":False,
            "error":str(e)

        })
       

@app.route(
    "/bulk_headers",
    methods=["POST"]
)
def bulk_headers():

    try:

        file = request.files.get("file")

        if not file:

            return jsonify({

                "success":False,
                "error":"File kosong"

            })

        df = pd.read_excel(
            file,
            header=None,
            nrows=1
        )

        headers = []

        if not df.empty:

            for index, value in enumerate(df.iloc[0].tolist()):

                name = ""

                if not pd.isna(value):

                    name = str(value).strip()

                if not name:

                    name = f"Kolom {index + 1}"

                headers.append({

                    "index":index,
                    "name":name

                })

        return jsonify({

            "success":True,
            "headers":headers

        })

    except Exception as e:

        return jsonify({

            "success":False,
            "error":str(e)

        })

@app.route(
    "/bulk_upload",
    methods=["POST"]
)
def bulk_upload():

    try:

        file = request.files.get("file")
        use_kode = (
            request.form.get("use_kode")
            == "true"
        )
        import_mode = request.form.get("import_mode") or ("package_code" if use_kode else "waybill")
        import_date = normalize_import_date(
            request.form.get("import_date")
        )

        resi_items = []

        if import_mode == "package_code":

            raw_codes = request.form.get("package_codes", "")
            package_codes = [
                item.strip()
                for item in re.split(r"[\s,;]+", raw_codes)
                if item.strip()
            ]

            if not package_codes:

                return jsonify({
                    "success": False,
                    "error": "Isi kode dulu"
                })

            seen_codes = set()

            for code in package_codes:

                key = code.upper()

                if key in seen_codes:
                    continue

                seen_codes.add(key)
                resi_items.append({
                    "package_code": code,
                    "import_date": import_date
                })

        else:

            if not file:

                return jsonify({
                    "success":False,
                    "error":"File kosong"
                })

            filepath = os.path.join(
                UPLOAD_FOLDER,
                secure_filename(file.filename)
            )

            file.save(filepath)

            df = pd.read_excel(
                filepath,
                header=None
            )

            for row_index, row in df.iterrows():

                val = row.iloc[0]

                if pd.isna(val):
                    continue

                val = str(val).strip()

                if not val:
                    continue

                header_key = (
                    val.lower()
                    .replace(" ", "")
                    .replace(".", "")
                    .replace("_", "")
                    .replace("-", "")
                )

                if header_key in (
                    "nowaybill",
                    "waybill",
                    "resi",
                    "awb"
                ):
                    continue

                resi_items.append({
                    "resi":val,
                    "kode":"-",
                    "import_date":import_date
                })

        if not resi_items:
            return jsonify({
                "success": False,
                "error": "Tidak ada data yang bisa diimport"
            }), 400

        if bulk_job_get("running"):
            return jsonify({
                "success": False,
                "error": "Masih ada proses bulk import yang berjalan"
            }), 409

        batch_id = (
            f"{import_date}-{int(time.time() * 1000)}-"
            f"{secrets.token_hex(4)}"
        )

        # Simpan seluruh daftar dan posisi awal secara lokal sebelum worker
        # dimulai, sehingga listrik mati tidak memaksa import dari awal.
        write_bulk_checkpoint(batch_id, resi_items, 0)

        with DB_LOCK:
            conn.execute("""
                INSERT INTO bulk_import_batches
                (batch_id, import_date, import_mode, created_at,
                 status, total_items, processed)
                VALUES (?, ?, ?, ?, 'RUNNING', ?, 0)
            """, (
                batch_id,
                import_date,
                import_mode,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(resi_items)
            ))
            conn.commit()

        update_bulk_job(
            running=True,
            total=len(resi_items),
            processed=0,
            success=0,
            failed=0,
            updated=0,
            tracking_failed=0,
            skipped=0,
            skip_details=[],
            import_date=import_date,
            batch_id=batch_id,
            current_item="",
            current_step="Menyiapkan import",
            started_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            done=False,
            cancel=False
        )

        threading.Thread(
            target=bulk_worker,
            args=(resi_items, batch_id),
            daemon=True
        ).start()

        return jsonify({
            "success":True,
            "total":len(resi_items),
            "batch_id":batch_id
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        })

def bulk_worker(
    resi_items,
    batch_id,
    start_index=0
):

    global BULK_JOB

    BULK_JOB.setdefault("skipped", 0)
    BULK_JOB.setdefault("skip_details", [])
    BULK_JOB.setdefault("updated", 0)
    BULK_JOB.setdefault("tracking_failed", 0)
    processed_package_waybills = set()

    for item_index, item in enumerate(resi_items[start_index:], start=start_index):

        if bulk_job_get("cancel"):
            break

        try:

            if isinstance(item, dict):

                package_code = item.get("package_code")
                resi = item.get("resi")
                kode = item.get("kode", "-")
                import_date = item.get("import_date")

            else:

                package_code = None
                resi = item
                kode = "-"
                import_date = None

            current_item_label = package_code or resi or "-"

            update_bulk_job(
                current_item=str(current_item_label),
                current_step="Mulai proses"
            )

            if package_code:

                update_bulk_job(
                    current_item=str(package_code),
                    current_step="Ambil resi dari kode"
                )

                package_result = get_waybills_by_package_number(package_code)

                if bulk_job_get("cancel"):
                    break

                if not package_result.get("success"):

                    bulk_job_skip(
                        package_code,
                        "-",
                        package_result.get("message", "Kode gagal dibaca")
                    )
                    bulk_job_add("processed")
                    update_bulk_job(current_step="Skip kode gagal")
                    write_bulk_checkpoint(batch_id, resi_items, item_index + 1)
                    continue

                package_waybills = package_result.get("waybill_ids", [])

                update_bulk_job(
                    current_item=str(package_code),
                    current_step="Cek kuota retur"
                )

                quota_result = check_package_return_quota(
                    package_waybills,
                    import_date=import_date
                )

                if bulk_job_get("cancel"):
                    break

                if not quota_result.get("allowed"):

                    bulk_job_skip(
                        package_code,
                        "MODERN_PARK",
                        quota_result.get("reason", "Retur kurang dari 5")
                    )
                    bulk_job_add("processed")
                    update_bulk_job(current_step="Skip kuota retur")
                    write_bulk_checkpoint(batch_id, resi_items, item_index + 1)
                    continue

                package_waybills = quota_result.get("return_waybills", [])
                examine_times = quota_result.get("examine_times", {})

                for package_resi in package_waybills:

                    if bulk_job_get("cancel"):
                        break

                    resi_key = str(package_resi or "").strip().upper()

                    if resi_key in processed_package_waybills:
                        # Scan pertama sudah tersimpan di staging; perubahan
                        # kode duplikat diproses setelah sinkronisasi VPS.
                        continue

                    update_bulk_job(
                        current_item=str(package_resi),
                        current_step="Scan resi dari kode"
                    )

                    process_bulk_resi(
                        package_resi,
                        kode=package_code,
                        import_date=import_date,
                        bulk_import_id=batch_id
                    )

                    processed_package_waybills.add(resi_key)
                    bulk_job_add("success")

                if bulk_job_get("cancel"):
                    break

                bulk_job_add("processed")
                update_bulk_job(current_step="Selesai kode")
                write_bulk_checkpoint(batch_id, resi_items, item_index + 1)

                continue

            update_bulk_job(
                current_item=str(resi),
                current_step="Filter jaringan retur"
            )

            filter_result = check_bulk_return_network(
                resi,
                import_date=import_date
            )

            if not filter_result.get("allowed"):

                reason = filter_result.get("reason", "Not MODERN_PARK")
                network = filter_result.get("network", "-")

                print(
                    "BULK SKIP:",
                    resi,
                    reason
                )

                bulk_job_skip(resi, network, reason)

            else:

                update_bulk_job(
                    current_item=str(resi),
                    current_step="Scan resi"
                )

                process_bulk_resi(
                    resi,
                    kode=kode,
                    import_date=import_date,
                    bulk_import_id=batch_id
                )

                bulk_job_add("success")

        except Exception as e:

            print(
                "BULK ERROR:",
                e
            )

            bulk_job_add("failed")

        bulk_job_add("processed")
        update_bulk_job(current_step="Selesai item")
        write_bulk_checkpoint(batch_id, resi_items, item_index + 1)

    cancelled = bool(bulk_job_get("cancel"))
    if cancelled:
        write_bulk_checkpoint(batch_id, resi_items, bulk_job_get("processed", 0), "PAUSED")
        update_bulk_job(running=False, done=True, current_item="", current_step="Dibatalkan")
    else:
        update_bulk_job(current_item="", current_step="Simpan semua data ke VPS")
        write_bulk_checkpoint(batch_id, resi_items, len(resi_items), "FLUSHING")
        sync_ok, synced = flush_staged_bulk_scans(batch_id)
        if sync_ok:
            folder, manifest = _bulk_stage_paths(batch_id)
            manifest.unlink(missing_ok=True)
            try:
                folder.rmdir()
            except OSError:
                pass
            update_bulk_job(running=False, done=True, current_step=f"Selesai, {synced} data tersimpan ke VPS")
            # Data sudah tersedia di VPS; tracking retur dikerjakan setelahnya
            # agar tidak memperlambat tahap ambil seluruh bulk.
            batch_date = normalize_import_date(
                (resi_items[0].get("import_date") if resi_items and isinstance(resi_items[0], dict) else None)
            )
            EXECUTOR.submit(refresh_live_tracking_worker, batch_date)
        else:
            update_bulk_job(running=False, done=False, current_step="Sinkronisasi VPS tertunda - akan dilanjutkan")

    with DB_LOCK:
        conn.execute("""
            UPDATE bulk_import_batches
            SET status=?, processed=?
            WHERE batch_id=?
        """, (
            "CANCELLED" if cancelled else ("COMPLETED" if bulk_job_get("done") else "SYNC_PENDING"),
            bulk_job_get("processed", 0),
            batch_id
        ))
        conn.commit()


@app.route("/bulk_undo", methods=["POST"])
def bulk_undo():
    selected_date = normalize_import_date(
        (request.get_json(silent=True) or {}).get("date")
    )

    if BULK_JOB.get("running"):
        return jsonify({
            "success": False,
            "error": "Tunggu bulk import selesai atau batalkan prosesnya dulu"
        }), 409

    with DB_LOCK:
        conn_local = configure_sqlite_connection(
            sqlite3.connect(DB_FILE, timeout=15)
        )

        try:
            conn_local.execute("BEGIN IMMEDIATE")
            batch = conn_local.execute("""
                SELECT batch_id
                FROM bulk_import_batches
                WHERE import_date=?
                AND status IN ('COMPLETED', 'CANCELLED')
                ORDER BY created_at DESC, rowid DESC
                LIMIT 1
            """, (selected_date,)).fetchone()

            if not batch:
                conn_local.rollback()
                return jsonify({
                    "success": False,
                    "error": "Belum ada riwayat import yang bisa di-undo pada tanggal ini"
                }), 404

            batch_id = batch[0]
            changes = conn_local.execute("""
                SELECT scan_id, action, old_data
                FROM bulk_import_changes
                WHERE batch_id=?
                ORDER BY id DESC
            """, (batch_id,)).fetchall()

            scan_columns = {
                row[1]
                for row in conn_local.execute("PRAGMA table_info(scans)").fetchall()
            }
            deleted = 0
            restored = 0

            for scan_id, action, old_data in changes:
                if action == "INSERT":
                    deleted += conn_local.execute(
                        "DELETE FROM scans WHERE id=?",
                        (scan_id,)
                    ).rowcount
                    continue

                snapshot = json.loads(old_data or "{}")
                restore_columns = [
                    name for name in snapshot
                    if name != "id" and name in scan_columns
                ]
                if not restore_columns:
                    continue

                assignments = ", ".join(
                    f'"{name}"=?' for name in restore_columns
                )
                values = [snapshot[name] for name in restore_columns]
                result = conn_local.execute(
                    f'UPDATE scans SET {assignments} WHERE id=?',
                    values + [scan_id]
                )
                restored += result.rowcount

            conn_local.execute("""
                UPDATE bulk_import_batches
                SET status='UNDONE', undone_at=?
                WHERE batch_id=?
            """, (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                batch_id
            ))
            conn_local.commit()

            return jsonify({
                "success": True,
                "batch_id": batch_id,
                "deleted": deleted,
                "restored": restored,
                "total": deleted + restored,
                "message": "Import terakhir berhasil di-undo"
            })
        except Exception as error:
            conn_local.rollback()
            return jsonify({
                "success": False,
                "error": str(error)
            }), 500
        finally:
            conn_local.close()

@app.route("/bulk_cancel", methods=["POST"])
def bulk_cancel():

    was_running = bool(bulk_job_get("running"))
    update_bulk_job(cancel=True)

    return jsonify({
        "success": True,
        "running": was_running,
        "message": "Proses pembatalan dikirim" if was_running else "Tidak ada proses aktif"
    })

@app.route("/bulk_progress")
def bulk_progress():

    progress = bulk_job_snapshot()
    last_update = progress.get("last_update") or ""

    try:
        progress["stalled_seconds"] = int(
            (
                datetime.now() -
                datetime.strptime(last_update, "%Y-%m-%d %H:%M:%S")
            ).total_seconds()
        )
    except Exception:
        progress["stalled_seconds"] = 0

    return jsonify(progress)


@app.route("/scanner_debug", methods=["POST"])
def scanner_debug():

    data = request.json

    print(
        "\n[SCANNER DEBUG]",
        data.get("msg"),
        flush=True
    )

    return {"success": True}
    


@app.route("/send_account_message", methods=["POST"])
def send_account_message():

    recipient_username = request.form.get(
        "recipient_username",
        ""
    ).strip()
    message = request.form.get(
        "message",
        ""
    ).strip()
    photo = request.files.get("photo")

    if not recipient_username:

        return jsonify({
            "success":False,
            "error":"Tujuan akun belum dipilih"
        }), 400

    if not message and not photo:

        return jsonify({
            "success":False,
            "error":"Isi pesan atau upload foto dulu"
        }), 400

    recipient = cursor.execute("""

        SELECT
            fullname,
            username
        FROM users
        WHERE username=?

    """, (
        recipient_username,
    )).fetchone()

    if not recipient:

        return jsonify({
            "success":False,
            "error":"Akun tujuan tidak ditemukan"
        }), 404

    image_path = ""

    if photo and photo.filename:

        original_name = secure_filename(
            photo.filename
        )
        extension = os.path.splitext(
            original_name
        )[1].lower()

        if extension not in (
            ".jpg",
            ".jpeg",
            ".png",
            ".webp",
            ".gif"
        ):

            return jsonify({
                "success":False,
                "error":"Format foto harus JPG, PNG, WEBP, atau GIF"
            }), 400

        filename = "{}_{}{}".format(
            int(time.time() * 1000),
            secrets.token_hex(6),
            extension
        )
        photo.save(
            os.path.join(
                MESSAGE_UPLOAD_FOLDER,
                filename
            )
        )
        image_path = "/static/message_uploads/{}".format(
            filename
        )

    created_at = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    sender_username = session.get(
        "username",
        ""
    )
    sender_fullname = session.get(
        "fullname",
        sender_username
    )

    with DB_LOCK:
        cursor.execute(
            "DELETE FROM account_messages WHERE recipient_username=?",
            (recipient_username,)
        )
        cursor.execute("""

            INSERT INTO account_messages (
                sender_username,
                sender_fullname,
                recipient_username,
                recipient_fullname,
                message,
                image_path,
                message_type,
                app_version,
                created_at,
                read_at
            )
            VALUES (?,?,?,?,?,?,?,'',?,NULL)

        """, (
            sender_username,
            sender_fullname,
            recipient_username,
            recipient[0] or recipient_username,
            message,
            image_path,
            "direct_message",
            created_at
        ))

        conn.commit()

    return jsonify({
        "success":True,
        "recipient_username": recipient_username,
        "message": "Pesan hanya dikirim ke @{}".format(recipient_username)
    })

@app.route("/account_message_poll")
def account_message_poll():

    recipient_username = session.get(
        "username",
        ""
    )

    if not recipient_username:

        return jsonify({
            "success":True,
            "message":None
        })

    row = cursor.execute("""

        SELECT
            id,
            sender_username,
            sender_fullname,
            message,
            image_path,
            created_at,
            read_at
        FROM account_messages
        WHERE recipient_username=?
        ORDER BY id DESC
        LIMIT 1

    """, (
        recipient_username,
    )).fetchone()

    if not row or row[6]:

        return jsonify({
            "success":True,
            "message":None
        })

    return jsonify({
        "success":True,
        "message":{
            "id":row[0],
            "sender_username":row[1] or "",
            "sender_fullname":row[2] or row[1] or "Admin",
            "text":row[3] or "",
            "image_path":row[4] or "",
            "created_at":row[5] or ""
        }
    })

@app.route("/account_message_ack", methods=["POST"])
def account_message_ack():

    data = request.get_json(
        silent=True
    ) or {}
    message_id = data.get("id")
    recipient_username = session.get(
        "username",
        ""
    )

    if not message_id or not recipient_username:

        return jsonify({
            "success":False
        }), 400

    cursor.execute("""

        UPDATE account_messages
        SET read_at=?
        WHERE id=?
          AND recipient_username=?

    """, (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        message_id,
        recipient_username
    ))

    conn.commit()

    return jsonify({
        "success":True
    })

@app.route('/add_user', methods=['POST'])
def add_user():

    try:

        data = request.json

        fullname = data.get("fullname","").strip()
        username = data.get("username","").strip()
        password = data.get("password","").strip()
        level = data.get("level","ADMIN").strip()

        if is_reserved_superman_account(username, level):
            return jsonify({
                "success":False,
                "error":"Akun SUPERMAN tidak bisa dibuat dari menu tambah akun"
            }), 400

        attendance_bypass = 1 if truthy_flag(
            data.get("attendance_bypass", 0)
        ) else 0
        iphone_user = 1 if truthy_flag(
            data.get("iphone_user", 0)
        ) else 0
        disable_location_lock = 1 if truthy_flag(
            data.get("disable_location_lock", 0)
        ) else 0
        permission_items = normalize_permissions(
            data.get("permissions", [])
        )
        permission_items = [
            item for item in permission_items
            if item != ATTENDANCE_BYPASS_PERMISSION
        ]
        if attendance_bypass:
            permission_items.append(
                ATTENDANCE_BYPASS_PERMISSION
            )
        permissions = ",".join(
            sanitize_user_permissions(
                permission_items
            )
        )

        cursor.execute("""

        INSERT INTO users(

            fullname,
            username,
            password,
            level,
            permissions,
            attendance_bypass,
            iphone_user,
            disable_location_lock

        )

        VALUES(?,?,?,?,?,?,?,?)

        """,(

            fullname,
            username,
            password,
            level,
            permissions,
            attendance_bypass,
            iphone_user,
            disable_location_lock

        ))

        conn.commit()

        return jsonify({
            "success":True
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        })

@app.route('/import_users', methods=['POST'])
def import_users():

    try:

        file = request.files.get(
            "file"
        )

        if not file:

            return jsonify({
                "success":False,
                "error":"File Excel belum dipilih"
            }), 400

        workbook = load_workbook(
            BytesIO(
                file.read()
            ),
            read_only=True,
            data_only=True
        )

        sheet = workbook.active

        rows = []

        for row in sheet.iter_rows(
            min_row=1,
            max_col=1,
            values_only=True
        ):

            fullname = str(
                row[0] or ""
            ).strip()

            if fullname:

                rows.append(
                    fullname
                )

        if not rows:

            return jsonify({
                "success":False,
                "error":"Kolom A belum berisi nama"
            }), 400

        imported = []

        for fullname in rows:

            username = generate_unique_username(
                fullname,
                cursor
            )

            password = generate_import_password()

            cursor.execute("""

            INSERT INTO users(

                fullname,
                username,
                password,
                level,
                permissions,
                must_change_credentials

            )

            VALUES(?,?,?,?,?,?)

            """,(

                fullname,
                username,
                password,
                "ADMIN",
                ATTENDANCE_ONLY_PERMISSION,
                1

            ))

            imported.append({
                "fullname":fullname,
                "username":username,
                "password":password
            })

        conn.commit()

        return jsonify({
            "success":True,
            "imported":imported,
            "count":len(imported)
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        }), 500
        
@app.route('/delete_user', methods=['POST'])
def delete_user():

    try:

        user_id = request.json.get("id")

        target_user = cursor.execute("""
            SELECT username, level
            FROM users
            WHERE id=?
        """,(user_id,)).fetchone()

        if target_user and is_reserved_superman_account(target_user[0], target_user[1]):
            return jsonify({
                "success":False,
                "error":"User SUPERMAN tidak bisa dihapus"
            }), 400

        cursor.execute("""

        DELETE FROM users

        WHERE id=?

        """,(user_id,))

        conn.commit()

        return jsonify({
            "success":True
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        })

@app.route('/delete_attendance_user', methods=['POST'])
def delete_attendance_user():

    try:

        data = request.json or {}

        user_id = str(
            data.get(
                "id",
                ""
            )
        ).strip()

        username = str(
            data.get(
                "username",
                ""
            )
        ).strip()

        target_date = str(
            data.get(
                "date",
                ""
            )
        ).strip()

        target_user = None

        if user_id:

            target_user = cursor.execute("""

            SELECT
                id,
                username,
                level

            FROM users

            WHERE id=?

            """,(user_id,)).fetchone()

        if (
            not target_user and
            username
        ):

            target_user = cursor.execute("""

            SELECT
                id,
                username,
                level

            FROM users

            WHERE username=?

            """,(username,)).fetchone()

        target_username = (
            target_user[1]
            if target_user
            else username
        )

        if not target_username:

            return jsonify({
                "success":False,
                "error":"User tidak ditemukan"
            })

        if (
            str(target_username).lower() == "superman" or
            (
                target_user and
                str(target_user[2] or "").upper() == "SUPERMAN"
            )
        ):

            return jsonify({
                "success":False,
                "error":"User SUPERMAN tidak bisa dihapus"
            })

        if target_date:

            attendance_row = cursor.execute("""

            SELECT
                id,
                clock_out

            FROM attendance

            WHERE username=?
            AND tanggal=?

            ORDER BY id DESC
            LIMIT 1

            """,(
                target_username,
                target_date
            )).fetchone()

            if attendance_row and str(attendance_row[1] or "").strip():

                cursor.execute("""

                UPDATE attendance

                SET
                    clock_out='',
                    clock_out_at='',
                    clock_out_photo='',
                    clock_out_latitude='',
                    clock_out_longitude='',
                    clock_out_address=''

                WHERE id=?

                """,(attendance_row[0],))

                conn.commit()

                return jsonify({
                    "success":True,
                    "mode":"date",
                    "deleted_stage":"clock_out"
                })

            if attendance_row:

                cursor.execute("""

                DELETE FROM attendance

                WHERE id=?

                """,(attendance_row[0],))

                cursor.execute("""

                DELETE FROM attendance_leave

                WHERE username=?
                AND tanggal=?

                """,(
                    target_username,
                    target_date
                ))

                conn.commit()

                return jsonify({
                    "success":True,
                    "mode":"date",
                    "deleted_stage":"clock_in"
                })

            cursor.execute("""

            DELETE FROM attendance_leave

            WHERE username=?
            AND tanggal=?

            """,(
                target_username,
                target_date
            ))

            conn.commit()

            return jsonify({
                "success":True,
                "mode":"date",
                "deleted_stage":"leave"
            })

        cursor.execute("""

        DELETE FROM attendance

        WHERE username=?

        """,(target_username,))

        cursor.execute("""

        DELETE FROM attendance_leave

        WHERE username=?

        """,(target_username,))

        if target_user:

            cursor.execute("""

            DELETE FROM users

            WHERE id=?

            """,(target_user[0],))

        conn.commit()

        return jsonify({
            "success":True
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        })
        
        
@app.route('/update_user', methods=['POST'])
def update_user():

    try:

        data = request.json

        user_id = data.get("id")

        fullname = data.get("fullname")
        username = data.get("username")
        password = data.get("password")
        level = data.get("level")

        if is_reserved_superman_account(username, level):
            return jsonify({
                "success":False,
                "error":"Akun SUPERMAN tidak bisa dibuat atau diedit dari account management"
            }), 400

        target_user = cursor.execute("""
            SELECT username, level
            FROM users
            WHERE id=?
        """,(user_id,)).fetchone()

        if target_user and is_reserved_superman_account(target_user[0], target_user[1]):
            return jsonify({
                "success":False,
                "error":"Akun SUPERMAN tidak bisa diedit"
            }), 400

        attendance_bypass = 1 if truthy_flag(
            data.get("attendance_bypass", 0)
        ) else 0
        iphone_user = 1 if truthy_flag(
            data.get("iphone_user", 0)
        ) else 0
        disable_location_lock = 1 if truthy_flag(
            data.get("disable_location_lock", 0)
        ) else 0
        permission_items = normalize_permissions(
            data.get("permissions", [])
        )
        permission_items = [
            item for item in permission_items
            if item != ATTENDANCE_BYPASS_PERMISSION
        ]
        if attendance_bypass:
            permission_items.append(
                ATTENDANCE_BYPASS_PERMISSION
            )
        permissions = ",".join(
            sanitize_user_permissions(
                permission_items
            )
        )

        cursor.execute("""

        UPDATE users

        SET

            fullname=?,
            username=?,
            password=?,
            level=?,
            permissions=?,
            attendance_bypass=?,
            iphone_user=?,
            disable_location_lock=?

        WHERE id=?

        """,(

            fullname,
            username,
            password,
            level,
            permissions,
            attendance_bypass,
            iphone_user,
            disable_location_lock,
            user_id

        ))

        conn.commit()

        return jsonify({
            "success":True
        })

    except Exception as e:

        return jsonify({
            "success":False,
            "error":str(e)
        })
        



@app.route("/api/check_attendance")
def check_attendance_status():

    username = session.get("username")

    if not requires_attendance_for_level(
        session.get(
            "level"
        ),
        session.get(
            "attendance_bypass",
            0
        )
    ):

        return jsonify({
            "attended":True,
            "attendance_only":
                is_attendance_only_session()
        })

    if not username:

        return jsonify({
            "attended":False
        })

    attendance = get_today_attendance(
        username
    )

    return jsonify({

        "attended":
        attendance is not None,

        "attendance_only":
            is_attendance_only_session(),

        "attendance":
                attendance,

            "redirect":(
                "/attendance_success"
                if attendance_only
                else "/attendance_dashboard"
            )

        })
    
    
@app.route("/api/permissions")
def api_permissions():

    return jsonify({

        "permissions":
            session.get(
                "permissions",
                []
            ),

        "level":
            session.get(
                "level",
                ""
            ),

        "username":
            session.get(
                "username",
                ""
            )

    })


def resume_pending_bulk_imports():
    """Continue the one unfinished local bulk checkpoint after restart."""
    for manifest in sorted(BULK_STAGE_DIR.glob("*/manifest.json")):
        try:
            with manifest.open("r", encoding="utf-8") as file:
                state = json.load(file)
            if state.get("state") not in {"RUNNING", "FLUSHING"}:
                continue
            items = state.get("items") or []
            batch_id = state.get("batch_id")
            start_index = min(max(int(state.get("next_index", 0)), 0), len(items))
            if not batch_id or not items:
                continue
            update_bulk_job(
                running=True, total=len(items), processed=start_index,
                success=0, failed=0, skipped=0, skip_details=[],
                batch_id=batch_id, current_item="", current_step="Melanjutkan bulk yang terinterupsi",
                done=False, cancel=False
            )
            threading.Thread(
                target=bulk_worker, args=(items, batch_id, start_index), daemon=True
            ).start()
            print(f"BULK RESUME: {batch_id} dari item {start_index + 1}/{len(items)}")
            break
        except Exception as error:
            print("BULK RESUME ERROR:", error)


if __name__ == '__main__':

    replay_pending_scan_outbox()
    resume_pending_bulk_imports()

    app.run(
        debug=False,
        host='0.0.0.0',
        port=5010,
        threaded=True,
        use_reloader=False
    )























































































